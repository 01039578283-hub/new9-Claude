from __future__ import annotations

import csv
import hashlib
import html
import json
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, Iterable, Iterator

from openpyxl import load_workbook


SITE = Path(__file__).resolve().parents[1]
COMMON = SITE.parent / "참고자료" / "공통자료"
COMMON_CSV = COMMON / "센터정보 정리.csv"
SOURCE_DIR = SITE.parents[1] / "새 폴더"
PARENT = "과목별학원"
EXPECTED_ROWS = 371
MIN_SENTENCE_CHARS = 60
SHINGLE_WORDS = 12
MAX_EXAMPLES = 8


@dataclass(frozen=True)
class Category:
    slug: str
    grade: str
    subject: str
    source_name: str


CATEGORIES = (
    Category("중1수학학원", "중1", "수학", "중1 수학학원 원고.xlsx"),
    Category("중1영어학원", "중1", "영어", "중1 영어학원 원고.xlsx"),
    Category("중2수학학원", "중2", "수학", "중2 수학학원 원고.xlsx"),
    Category("중2영어학원", "중2", "영어", "중2 영어학원 원고.xlsx"),
)


BLOCK_TAGS = {
    "address",
    "article",
    "aside",
    "blockquote",
    "br",
    "dd",
    "details",
    "div",
    "dl",
    "dt",
    "figcaption",
    "figure",
    "footer",
    "h1",
    "h2",
    "h3",
    "h4",
    "h5",
    "h6",
    "header",
    "hr",
    "li",
    "main",
    "nav",
    "ol",
    "p",
    "section",
    "summary",
    "table",
    "tbody",
    "td",
    "tfoot",
    "th",
    "thead",
    "tr",
    "ul",
}
VOID_TAGS = {
    "area",
    "base",
    "br",
    "col",
    "embed",
    "hr",
    "img",
    "input",
    "link",
    "meta",
    "param",
    "source",
    "track",
    "wbr",
}
SKIP_TAGS = {"script", "style", "template", "noscript"}
RISK_EXCLUDED_CLASS_FRAGMENTS = (
    "breadcrumb",
    "card-grid",
    "link-grid",
    "related-links",
    "subject-nav",
)

TOKEN_RE = re.compile(r"[가-힣]+|[a-z0-9]+(?:[-._][a-z0-9]+)*", re.I)
SENTENCE_SPLIT_RE = re.compile(r"(?:\r?\n)+|(?<=[.!?。！？])\s+")
GRADE_SHORT_RE = re.compile(r"(?<![가-힣A-Za-z0-9])(초|중|고)\s*([1-6])(?![가-힣A-Za-z0-9])")
GRADE_LONG_RE = re.compile(
    r"(?<![가-힣A-Za-z0-9])(초등학교|중학교|고등학교)\s*([1-6])\s*학년(?![가-힣A-Za-z0-9])"
)

OVERCLAIM_PATTERNS = (
    (
        "성적·점수 상승 단정",
        re.compile(
            r"(?:성적|점수|등급)(?:이|가|을|를|도|은|는|의|으로|에서|까지|\s)*"
            r"(?:바로\s*)?(?:오르|올리|상승|향상|개선|회복|높이|끌어올리|완성|보장)"
        ),
    ),
    (
        "결과 보장 표현",
        re.compile(
            r"(?:100\s*%|무조건|반드시|확실히|완벽(?:히|하게)?|단기간에?).{0,30}"
            r"(?:성적|점수|등급|향상|상승|합격|해결|보장)"
            r"|(?:성적|점수|등급|합격).{0,30}(?:100\s*%|무조건|반드시|보장)"
        ),
    ),
    ("하락 방지 단정", re.compile(r"(?:성적|점수)\s*하락을\s*(?:막|방지)")),
)

TEACHER_FEATURE_RE = re.compile(r"(?:선생님|교사|강사)\s*(?:의\s*)?특징")
AUTHORING_PATTERNS = (
    ("D열 제작 메모", re.compile(r"(?<![A-Za-z])D열(?![A-Za-z])", re.I)),
    ("원고 제작 메모", re.compile(r"(?:이|본|해당|첨부|기존)?\s*원고(?:를|의|에|는|에서|로|가)?")),
    ("검색 최적화 제작어", re.compile(r"(?<![A-Za-z])(?:SEO|AEO|GEO)(?![A-Za-z])", re.I)),
    ("JSON-LD 제작어", re.compile(r"JSON\s*-?\s*LD", re.I)),
    ("메타 제작어", re.compile(r"메타\s*(?:설명|요소|태그|데이터)")),
    ("생성 프롬프트 흔적", re.compile(r"프롬프트|생성형\s*AI|챗\s*GPT|ChatGPT", re.I)),
    ("복사·재작성 메모", re.compile(r"(?:다른\s*사이트|기존\s*글|복사한\s*내용|재작성한\s*내용)")),
    ("저자 메타 화법", re.compile(r"(?:이번|이)\s*글에서는|독자(?:분|님)|검색\s*엔진")),
)

FORBIDDEN_SCHEMA_TYPES = {"Review", "AggregateRating"}
FORBIDDEN_SCHEMA_KEYS = {
    "aggregateRating",
    "bestRating",
    "ratingCount",
    "ratingValue",
    "review",
    "reviewCount",
    "reviewRating",
    "worstRating",
}


@dataclass(frozen=True)
class Frame:
    tag: str
    hidden: bool
    in_main: bool
    risk_excluded: bool


class PublicTextParser(HTMLParser):
    """Collect public text while excluding scripts and visually hidden nodes."""

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.stack: list[Frame] = []
        self.main_count = 0
        self.all_parts: list[str] = []
        self.main_parts: list[str] = []
        self.risk_parts: list[str] = []

    @staticmethod
    def _is_hidden(tag: str, attrs: dict[str, str]) -> bool:
        style = attrs.get("style", "").replace(" ", "").lower()
        return (
            tag in SKIP_TAGS
            or "hidden" in attrs
            or attrs.get("aria-hidden", "").strip().lower() == "true"
            or "display:none" in style
            or "visibility:hidden" in style
        )

    @staticmethod
    def _is_risk_excluded(tag: str, attrs: dict[str, str]) -> bool:
        if tag in {"nav", "footer"}:
            return True
        classes = attrs.get("class", "").lower()
        return any(fragment in classes for fragment in RISK_EXCLUDED_CLASS_FRAGMENTS)

    def _append(self, value: str, frame: Frame | None = None) -> None:
        current = frame or (self.stack[-1] if self.stack else None)
        if current is not None and current.hidden:
            return
        self.all_parts.append(value)
        if current is not None and current.in_main:
            self.main_parts.append(value)
            if not current.risk_excluded:
                self.risk_parts.append(value)

    def handle_starttag(self, tag: str, attrs_raw: list[tuple[str, str | None]]) -> None:
        tag = tag.lower()
        attrs = {key.lower(): value or "" for key, value in attrs_raw}
        parent = self.stack[-1] if self.stack else None
        hidden = (parent.hidden if parent else False) or self._is_hidden(tag, attrs)
        in_main = (parent.in_main if parent else False) or tag == "main"
        risk_excluded = (
            (parent.risk_excluded if parent else False)
            or self._is_risk_excluded(tag, attrs)
        )
        frame = Frame(tag, hidden, in_main, risk_excluded)
        if tag == "main":
            self.main_count += 1
        if tag in BLOCK_TAGS:
            self._append("\n", frame)
        if tag not in VOID_TAGS:
            self.stack.append(frame)

    def handle_startendtag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        self.handle_starttag(tag, attrs)
        if tag.lower() not in VOID_TAGS:
            self.handle_endtag(tag)

    def handle_endtag(self, tag: str) -> None:
        tag = tag.lower()
        match_index = next(
            (index for index in range(len(self.stack) - 1, -1, -1) if self.stack[index].tag == tag),
            None,
        )
        if match_index is None:
            return
        frame = self.stack[match_index]
        if tag in BLOCK_TAGS:
            self._append("\n", frame)
        del self.stack[match_index:]

    def handle_data(self, data: str) -> None:
        if not data:
            return
        self._append(data)

    @property
    def all_text(self) -> str:
        return normalize_public_text(" ".join(self.all_parts))

    @property
    def main_text(self) -> str:
        return normalize_public_text(" ".join(self.main_parts))

    @property
    def risk_text(self) -> str:
        return normalize_public_text(" ".join(self.risk_parts))


@dataclass(frozen=True)
class SourceRef:
    category: str
    row: int
    local: str

    def label(self) -> str:
        return f"{self.category} row={self.row} local={self.local}"


class Audit:
    def __init__(self) -> None:
        self.counts: Counter[str] = Counter()
        self.examples: dict[str, list[str]] = defaultdict(list)
        self.metrics: Counter[str] = Counter()

    def fail(self, code: str, path: Path | str, detail: str) -> None:
        self.counts[code] += 1
        if len(self.examples[code]) < MAX_EXAMPLES:
            try:
                label = str(Path(path).resolve().relative_to(SITE))
            except (OSError, ValueError):
                label = str(path)
            self.examples[code].append(f"{label}: {detail}")

    @property
    def failed(self) -> bool:
        return bool(self.counts)

    def report(self) -> int:
        print(
            "[INFO] middle source rewrite audit "
            f"sources={self.metrics['source_docs']} "
            f"pages={self.metrics['generated_pages']} "
            f"sentence_fingerprints={self.metrics['source_sentences']} "
            f"shingle_fingerprints={self.metrics['source_shingles']}"
        )
        if not self.failed:
            print(
                "[PASS] 4개 중등 카테고리의 공개 main이 원고 재사용, 위험 표현, "
                "학교 목록, 후기 스키마 및 본문 중복 감사를 통과했습니다."
            )
            return 0
        print(f"[FAIL] failure_count={sum(self.counts.values())} codes={len(self.counts)}")
        for code in sorted(self.counts):
            print(f"- {code}: {self.counts[code]}")
            for example in self.examples[code]:
                print(f"  * {example}")
        return 1


def normalize_public_text(value: str) -> str:
    value = unicodedata.normalize("NFKC", html.unescape(value or ""))
    value = value.replace("\u200b", "").replace("\ufeff", "")
    value = re.sub(r"[ \t\f\v]+", " ", value)
    value = re.sub(r" *\n *", "\n", value)
    value = re.sub(r"\n{2,}", "\n", value)
    return value.strip()


def token_fingerprint(value: str) -> str:
    return " ".join(token.lower() for token in TOKEN_RE.findall(value))


def split_sentences(value: str) -> list[str]:
    result: list[str] = []
    for part in SENTENCE_SPLIT_RE.split(value):
        normalized = normalize_public_text(part)
        if normalized:
            result.append(normalized)
    return result


def split_items(value: str) -> list[str]:
    """Split the middle-school cells, which mix dots, middots and spaces."""
    result: list[str] = []
    for part in re.split(r"[,/·.\s]+", value or ""):
        name = part.strip()
        if name.endswith("중학교"):
            name = name[:-2]
        if name and name not in result:
            result.append(name)
    return result


def slug_ko(value: str) -> str:
    value = re.sub(r"\s+", "", value.strip())
    return re.sub(r'[\\/:*?"<>|#%&+]', "", value)


def read_common(audit: Audit) -> list[dict[str, str]]:
    if not COMMON_CSV.is_file():
        audit.fail("common_csv_missing", COMMON_CSV, "센터정보 정리.csv 없음")
        return []
    with COMMON_CSV.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    if len(rows) != EXPECTED_ROWS:
        audit.fail("common_row_count", COMMON_CSV, f"expected={EXPECTED_ROWS} actual={len(rows)}")
    slugs = [slug_ko(row.get("근처 수업가능 동네", "")) for row in rows]
    if any(not slug for slug in slugs):
        audit.fail("common_empty_local", COMMON_CSV, "빈 지역 또는 slug 존재")
    duplicate_slugs = [slug for slug, count in Counter(slugs).items() if count > 1]
    if duplicate_slugs:
        audit.fail("common_duplicate_slug", COMMON_CSV, repr(duplicate_slugs[:10]))
    return rows


def parse_public_html(source: str) -> PublicTextParser:
    parser = PublicTextParser()
    parser.feed(source)
    parser.close()
    return parser


def build_geo_pattern(rows: list[dict[str, str]]) -> re.Pattern[str] | None:
    terms: set[str] = set()
    for row in rows:
        for key in ("근처 수업가능 동네", "지역", "시or구", "센터명"):
            value = normalize_public_text(row.get(key, ""))
            if value:
                terms.add(value)
    if not terms:
        return None
    alternatives = "|".join(re.escape(term) for term in sorted(terms, key=len, reverse=True))
    return re.compile(rf"(?<![가-힣A-Za-z0-9])(?:{alternatives})(?![가-힣A-Za-z0-9])", re.I)


def mask_copy_entities(value: str, geo_pattern: re.Pattern[str] | None) -> str:
    value = normalize_public_text(value)
    value = re.sub(
        r"(?:와와학습코칭센터|스터디와와|코칭아카데미|코칭센터)(?:\s*[가-힣A-Za-z0-9()]+점)?",
        " ORGTOKEN ",
        value,
        flags=re.I,
    )
    value = GRADE_LONG_RE.sub(" GRADETOKEN ", value)
    value = GRADE_SHORT_RE.sub(" GRADETOKEN ", value)
    value = re.sub(r"(?<![가-힣])(?:국어|영어|수학|과학|사회|영수)(?![가-힣])", " SUBJECTTOKEN ", value)
    if geo_pattern is not None:
        value = geo_pattern.sub(" LOCTOKEN ", value)
    return token_fingerprint(value)


def shingle_hash(tokens: tuple[str, ...]) -> int:
    payload = "\x1f".join(tokens).encode("utf-8")
    return int.from_bytes(hashlib.blake2b(payload, digest_size=8).digest(), "big")


def iter_shingles(fingerprint: str) -> Iterator[tuple[int, str]]:
    words = fingerprint.split()
    if len(words) < SHINGLE_WORDS:
        return
    for index in range(len(words) - SHINGLE_WORDS + 1):
        shingle = tuple(words[index : index + SHINGLE_WORDS])
        yield shingle_hash(shingle), " ".join(shingle)


def load_sources(
    rows: list[dict[str, str]],
    geo_pattern: re.Pattern[str] | None,
    audit: Audit,
) -> tuple[dict[str, SourceRef], dict[str, SourceRef], dict[int, SourceRef]]:
    raw_sentences: dict[str, SourceRef] = {}
    masked_sentences: dict[str, SourceRef] = {}
    shingles: dict[int, SourceRef] = {}
    for category in CATEGORIES:
        path = SOURCE_DIR / category.source_name
        if not path.is_file():
            audit.fail("source_xlsx_missing", path, "원고 파일 없음")
            continue
        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook.active
        values = [
            (excel_row, row[0])
            for excel_row, row in enumerate(worksheet.iter_rows(values_only=True), 1)
            if row and isinstance(row[0], str) and row[0].strip()
        ]
        workbook.close()
        if len(values) != EXPECTED_ROWS:
            audit.fail(
                "source_nonempty_row_count",
                path,
                f"expected={EXPECTED_ROWS} actual={len(values)}",
            )
        for position, (excel_row, fragment) in enumerate(values, 1):
            local = rows[position - 1].get("근처 수업가능 동네", "") if position <= len(rows) else ""
            ref = SourceRef(category.slug, excel_row, local)
            parser = parse_public_html(fragment)
            public = parser.main_text or parser.all_text
            if not public:
                audit.fail("source_public_text_empty", path, f"row={excel_row} local={local}")
                continue
            audit.metrics["source_docs"] += 1
            if "봉담3지구" in public:
                audit.metrics["source_bongdam3_docs"] += 1
            for sentence in split_sentences(public):
                raw = token_fingerprint(sentence)
                masked = mask_copy_entities(sentence, geo_pattern)
                if len(raw) >= MIN_SENTENCE_CHARS:
                    raw_sentences.setdefault(raw, ref)
                if len(masked) >= MIN_SENTENCE_CHARS:
                    masked_sentences.setdefault(masked, ref)
            document = mask_copy_entities(public, geo_pattern)
            for digest, _ in iter_shingles(document):
                shingles.setdefault(digest, ref)
    audit.metrics["source_sentences"] = len(set(raw_sentences) | set(masked_sentences))
    audit.metrics["source_shingles"] = len(shingles)
    if audit.metrics["source_docs"] == 0:
        audit.fail("source_output_missing", SOURCE_DIR, "비교 가능한 원고 공개 텍스트가 0개")
    return raw_sentences, masked_sentences, shingles


def json_values(source: str) -> tuple[list[Any], list[str]]:
    values: list[Any] = []
    errors: list[str] = []
    pattern = re.compile(r"<script\b([^>]*)>(.*?)</script\s*>", re.I | re.S)
    for attrs, payload in pattern.findall(source):
        if not re.search(r"\btype\s*=\s*([\"'])application/ld\+json\1", attrs, re.I):
            continue
        try:
            values.append(json.loads(payload.strip()))
        except (TypeError, ValueError) as exc:
            errors.append(str(exc))
    return values, errors


def json_walk(value: Any) -> Iterator[Any]:
    yield value
    if isinstance(value, dict):
        for child in value.values():
            yield from json_walk(child)
    elif isinstance(value, list):
        for child in value:
            yield from json_walk(child)


def schema_types(node: dict[str, Any]) -> set[str]:
    value = node.get("@type")
    if isinstance(value, str):
        return {value}
    if isinstance(value, list):
        return {item for item in value if isinstance(item, str)}
    return set()


def canonical_school(value: str) -> str:
    value = normalize_public_text(value)
    return value[:-2] if value.endswith("중학교") else value


def build_school_pattern(rows: list[dict[str, str]]) -> tuple[re.Pattern[str] | None, set[str]]:
    names = {
        canonical_school(name)
        for row in rows
        for name in split_items(row.get("타깃학교\n(중)", ""))
        if name.strip()
    }
    names.discard("")
    if not names:
        return None, names
    alternatives = "|".join(re.escape(name) for name in sorted(names, key=len, reverse=True))
    pattern = re.compile(
        rf"(?<![가-힣A-Za-z0-9])({alternatives})(?:학교)?(?![가-힣A-Za-z0-9])"
    )
    return pattern, names


def mentioned_middle_schools(
    value: str,
    known_pattern: re.Pattern[str] | None,
) -> tuple[set[str], set[str]]:
    known = {canonical_school(match.group(1)) for match in known_pattern.finditer(value)} if known_pattern else set()
    explicit = {
        canonical_school(match.group(1))
        for match in re.finditer(
            r"(?<![가-힣A-Za-z0-9])([가-힣A-Za-z0-9·]{2,20}중학교)(?![가-힣A-Za-z0-9])",
            value,
        )
    }
    return known, explicit - known


def grade_tokens(value: str) -> set[str]:
    result = {f"{match.group(1)}{match.group(2)}" for match in GRADE_SHORT_RE.finditer(value)}
    level = {"초등학교": "초", "중학교": "중", "고등학교": "고"}
    result.update(f"{level[match.group(1)]}{match.group(2)}" for match in GRADE_LONG_RE.finditer(value))
    return result


def disclosure_ok(value: str) -> bool:
    not_review = bool(
        re.search(r"실제.{0,40}(?:이용\s*)?후기.{0,60}(?:아니|않)", value)
        or re.search(r"후기.{0,40}(?:아니|않).{0,80}예시", value)
    )
    virtual_example = bool(re.search(r"(?:가상|상담\s*상황)\s*예시", value))
    variable_result = bool(re.search(r"학습\s*결과.{0,80}(?:달라질|다를)\s*수", value))
    return not_review and virtual_example and variable_result


def validate_review_schema(source: str, path: Path, audit: Audit) -> None:
    values, errors = json_values(source)
    if not values:
        audit.fail("jsonld_output_missing", path, "JSON-LD가 없어 Review 금지 여부를 검증할 수 없음")
    for error in errors:
        audit.fail("jsonld_parse_error", path, error)
    for value in values:
        for node in json_walk(value):
            if not isinstance(node, dict):
                continue
            bad_types = schema_types(node) & FORBIDDEN_SCHEMA_TYPES
            if bad_types:
                audit.fail("review_schema_forbidden", path, f"types={sorted(bad_types)}")
            bad_keys = node.keys() & FORBIDDEN_SCHEMA_KEYS
            if bad_keys:
                audit.fail("review_schema_forbidden", path, f"keys={sorted(bad_keys)}")


def validate_copy(
    public: str,
    path: Path,
    geo_pattern: re.Pattern[str] | None,
    raw_sentences: dict[str, SourceRef],
    masked_sentences: dict[str, SourceRef],
    source_shingles: dict[int, SourceRef],
    audit: Audit,
) -> None:
    sentence_hits: list[str] = []
    for sentence in split_sentences(public):
        raw = token_fingerprint(sentence)
        masked = mask_copy_entities(sentence, geo_pattern)
        ref = None
        if len(raw) >= MIN_SENTENCE_CHARS:
            ref = raw_sentences.get(raw)
        if ref is None and len(masked) >= MIN_SENTENCE_CHARS:
            ref = masked_sentences.get(masked)
        if ref is not None:
            sentence_hits.append(f"{ref.label()} :: {sentence[:120]}")
    if sentence_hits:
        audit.fail(
            "source_sentence_reuse",
            path,
            f"matches={len(sentence_hits)} sample={sentence_hits[:2]!r}",
        )

    shingle_hits: dict[int, tuple[SourceRef, str]] = {}
    fingerprint = mask_copy_entities(public, geo_pattern)
    for digest, sample in iter_shingles(fingerprint):
        ref = source_shingles.get(digest)
        if ref is not None:
            shingle_hits.setdefault(digest, (ref, sample))
    if shingle_hits:
        examples = [f"{ref.label()} :: {sample}" for ref, sample in list(shingle_hits.values())[:2]]
        audit.fail(
            "source_12word_shingle_reuse",
            path,
            f"matches={len(shingle_hits)} sample={examples!r}",
        )


def validate_risk_text(value: str, category: Category, path: Path, audit: Audit) -> None:
    other_grades = sorted(grade_tokens(value) - {category.grade})
    if other_grades:
        audit.fail("other_grade_exposure", path, f"target={category.grade} found={other_grades}")
    for name, pattern in OVERCLAIM_PATTERNS:
        match = pattern.search(value)
        if match:
            audit.fail("outcome_guarantee_exposure", path, f"{name}: {match.group(0)!r}")
    match = TEACHER_FEATURE_RE.search(value)
    if match:
        audit.fail("teacher_feature_exposure", path, match.group(0))
    for name, pattern in AUTHORING_PATTERNS:
        match = pattern.search(value)
        if match:
            audit.fail("authoring_term_exposure", path, f"{name}: {match.group(0)!r}")
    if "봉담3지구" in value:
        audit.fail("source_bongdam3_exposure", path, "공통자료는 봉담2지구인데 봉담3지구가 노출됨")


def validate_disclosure(value: str, path: Path, audit: Audit) -> None:
    if not disclosure_ok(value):
        audit.fail(
            "consultation_example_disclaimer",
            path,
            "비실제후기 + 가상예시 + 결과 개인차 고지 중 하나 이상 누락",
        )
    unsafe_review_sentences = [
        sentence
        for sentence in split_sentences(value)
        if "후기" in sentence and not re.search(r"(?:아니|않|예시)", sentence)
    ]
    if unsafe_review_sentences:
        audit.fail(
            "visible_review_claim",
            path,
            f"sample={unsafe_review_sentences[:2]!r}",
        )


def validate_schools(
    value: str,
    row: dict[str, str],
    known_pattern: re.Pattern[str] | None,
    path: Path,
    audit: Audit,
) -> None:
    expected = {canonical_school(name) for name in split_items(row.get("타깃학교\n(중)", ""))}
    expected.discard("")
    mentioned, unknown_explicit = mentioned_middle_schools(value, known_pattern)
    unexpected = sorted(mentioned - expected)
    missing = sorted(expected - mentioned)
    if unexpected or unknown_explicit:
        audit.fail(
            "middle_school_whitelist",
            path,
            f"unexpected={unexpected} unknown_explicit={sorted(unknown_explicit)} expected={sorted(expected)}",
        )
    if missing:
        audit.fail("middle_school_missing", path, f"missing={missing}")


def expected_page_path(category: Category, row: dict[str, str]) -> Path:
    local = row.get("근처 수업가능 동네", "")
    return SITE / PARENT / category.slug / slug_ko(local) / "index.html"


def main() -> int:
    audit = Audit()
    rows = read_common(audit)
    if not rows:
        return audit.report()
    geo_pattern = build_geo_pattern(rows)
    school_pattern, _ = build_school_pattern(rows)
    raw_sentences, masked_sentences, source_shingles = load_sources(rows, geo_pattern, audit)

    body_groups: dict[str, list[Path]] = defaultdict(list)
    seen_pages: set[Path] = set()
    for category in CATEGORIES:
        category_dir = SITE / PARENT / category.slug
        actual = {
            path.resolve()
            for path in category_dir.glob("*/index.html")
            if path.parent != category_dir
        } if category_dir.is_dir() else set()
        expected = {expected_page_path(category, row).resolve() for row in rows}
        missing = sorted(expected - actual)
        unexpected = sorted(actual - expected)
        if len(actual) != EXPECTED_ROWS or missing or unexpected:
            audit.fail(
                "generated_page_set",
                category_dir,
                f"expected={EXPECTED_ROWS} actual={len(actual)} missing={len(missing)} "
                f"unexpected={len(unexpected)} missing_sample={[str(p.parent.name) for p in missing[:5]]}",
            )

        for row in rows:
            path = expected_page_path(category, row)
            if not path.is_file():
                continue
            resolved = path.resolve()
            if resolved in seen_pages:
                audit.fail("generated_page_repeated", path, "같은 파일이 두 번 매핑됨")
                continue
            seen_pages.add(resolved)
            source = path.read_text(encoding="utf-8")
            parser = parse_public_html(source)
            if parser.main_count != 1:
                audit.fail("main_count", path, f"count={parser.main_count}")
            public = parser.main_text
            risk = parser.risk_text
            if not public:
                audit.fail("main_output_empty", path, "공개 main 텍스트가 비어 있음")
                continue
            if not risk:
                audit.fail("risk_output_empty", path, "내부링크를 제외한 공개 main 텍스트가 비어 있음")
                continue
            audit.metrics["generated_pages"] += 1
            validate_copy(
                public,
                path,
                geo_pattern,
                raw_sentences,
                masked_sentences,
                source_shingles,
                audit,
            )
            validate_risk_text(risk, category, path, audit)
            validate_disclosure(public, path, audit)
            validate_review_schema(source, path, audit)
            validate_schools(public, row, school_pattern, path, audit)
            body = token_fingerprint(public)
            if not body:
                audit.fail("page_body_fingerprint_empty", path, "본문 지문 생성 실패")
            else:
                body_groups[body].append(path)

    duplicate_groups = [paths for paths in body_groups.values() if len(paths) > 1]
    for paths in duplicate_groups:
        audit.fail(
            "exact_page_body_duplicate",
            paths[0],
            f"count={len(paths)} paths={[str(path.relative_to(SITE)) for path in paths[:6]]}",
        )
    if audit.metrics["generated_pages"] == 0:
        audit.fail(
            "generated_output_missing",
            SITE / PARENT,
            "4개 신규 카테고리에서 검사 가능한 공개 main 출력이 0개",
        )
    return audit.report()


if __name__ == "__main__":
    sys.exit(main())
