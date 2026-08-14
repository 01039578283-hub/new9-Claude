from __future__ import annotations

import argparse
import hashlib
import html
import math
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from html.parser import HTMLParser
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import load_workbook

import generate_grade1_math_pages as generator


EXPECTED_PAGE_COUNT = 371
CONTENT_DATE = "2026-08-14"
SOURCE_DIR = Path.home() / "Desktop" / "새 폴더"


@dataclass(frozen=True)
class CategoryConfig:
    category: str
    grade_label: str
    grade_number: int
    grade_en: str
    subject: str
    subject_en: str
    focus_label: str
    workbook_name: str
    representative_offset: int
    school_level: str = "중학교"
    school_key: str = "타깃학교\n(중)"

    @property
    def subject_label(self) -> str:
        return f"{self.grade_label} {self.subject}"

    @property
    def workbook_path(self) -> Path:
        return SOURCE_DIR / self.workbook_name

    @property
    def is_elementary(self) -> bool:
        return self.school_level == "초등학교"


@dataclass(frozen=True)
class TopicSignal:
    code: str
    label: str
    keywords: tuple[str, ...]
    check: str
    evidence: str
    practice: str
    home_action: str


@dataclass(frozen=True)
class SourceSelection:
    primary: TopicSignal
    secondary: TopicSignal
    support: TopicSignal

    @property
    def codes(self) -> tuple[str, str, str]:
        return (self.primary.code, self.secondary.code, self.support.code)


CONFIGS: tuple[CategoryConfig, ...] = (
    CategoryConfig(
        category="중1수학학원",
        grade_label="중1",
        grade_number=1,
        grade_en="GRADE 7",
        subject="수학",
        subject_en="MATH",
        focus_label="개념 연결·오답 재학습",
        workbook_name="중1 수학학원 원고.xlsx",
        representative_offset=0,
    ),
    CategoryConfig(
        category="중1영어학원",
        grade_label="중1",
        grade_number=1,
        grade_en="GRADE 7",
        subject="영어",
        subject_en="ENGLISH",
        focus_label="어휘·문장 구조·독해 근거",
        workbook_name="중1 영어학원 원고.xlsx",
        representative_offset=93,
    ),
    CategoryConfig(
        category="중2수학학원",
        grade_label="중2",
        grade_number=2,
        grade_en="GRADE 8",
        subject="수학",
        subject_en="MATH",
        focus_label="개념 적용·서술형·누적 오답",
        workbook_name="중2 수학학원 원고.xlsx",
        representative_offset=186,
    ),
    CategoryConfig(
        category="중2영어학원",
        grade_label="중2",
        grade_number=2,
        grade_en="GRADE 8",
        subject="영어",
        subject_en="ENGLISH",
        focus_label="어휘 누적·문법 적용·독해 근거",
        workbook_name="중2 영어학원 원고.xlsx",
        representative_offset=279,
    ),
    CategoryConfig(
        category="중3수학학원",
        grade_label="중3",
        grade_number=3,
        grade_en="GRADE 9",
        subject="수학",
        subject_en="MATH",
        focus_label="개념 통합·서술형·고등 기초 연결",
        workbook_name="중3 수학학원 원고.xlsx",
        representative_offset=31,
    ),
    CategoryConfig(
        category="중3영어학원",
        grade_label="중3",
        grade_number=3,
        grade_en="GRADE 9",
        subject="영어",
        subject_en="ENGLISH",
        focus_label="어휘·구문·독해 근거·서술형",
        workbook_name="중3 영어학원 원고.xlsx",
        representative_offset=124,
    ),
    CategoryConfig(
        category="초3수학학원",
        grade_label="초3",
        grade_number=3,
        grade_en="GRADE 3",
        subject="수학",
        subject_en="MATH",
        focus_label="연산 정확도·개념 설명·문장제",
        workbook_name="초3 수학학원 원고.xlsx",
        representative_offset=217,
        school_level="초등학교",
        school_key="타깃학교\n(초)",
    ),
    CategoryConfig(
        category="초3영어학원",
        grade_label="초3",
        grade_number=3,
        grade_en="GRADE 3",
        subject="영어",
        subject_en="ENGLISH",
        focus_label="파닉스·기초 어휘·문장 읽기",
        workbook_name="초3 영어학원.xlsx",
        representative_offset=62,
        school_level="초등학교",
        school_key="타깃학교\n(초)",
    ),
    CategoryConfig(
        category="초4수학학원",
        grade_label="초4",
        grade_number=4,
        grade_en="GRADE 4",
        subject="수학",
        subject_en="MATH",
        focus_label="연산 원리·분수·문장제 풀이",
        workbook_name="초4 수학학원 원고.xlsx",
        representative_offset=155,
        school_level="초등학교",
        school_key="타깃학교\n(초)",
    ),
    CategoryConfig(
        category="초4영어학원",
        grade_label="초4",
        grade_number=4,
        grade_en="GRADE 4",
        subject="영어",
        subject_en="ENGLISH",
        focus_label="어휘·문장 구조·독해 근거",
        workbook_name="초4 영어학원 원고.xlsx",
        representative_offset=248,
        school_level="초등학교",
        school_key="타깃학교\n(초)",
    ),
    CategoryConfig(
        category="초5수학학원",
        grade_label="초5",
        grade_number=5,
        grade_en="GRADE 5",
        subject="수학",
        subject_en="MATH",
        focus_label="분수·소수·도형·문장제",
        workbook_name="초5 수학학원 원고.xlsx",
        representative_offset=341,
        school_level="초등학교",
        school_key="타깃학교\n(초)",
    ),
    CategoryConfig(
        category="초5영어학원",
        grade_label="초5",
        grade_number=5,
        grade_en="GRADE 5",
        subject="영어",
        subject_en="ENGLISH",
        focus_label="어휘 확장·문장 구조·독해 근거",
        workbook_name="초5 영어학원 원고.xlsx",
        representative_offset=15,
        school_level="초등학교",
        school_key="타깃학교\n(초)",
    ),
    CategoryConfig(
        category="초6수학학원",
        grade_label="초6",
        grade_number=6,
        grade_en="GRADE 6",
        subject="수학",
        subject_en="MATH",
        focus_label="분수·소수 연산·비와 비율·문제 해결",
        workbook_name="초6 수학학원 원고.xlsx",
        representative_offset=108,
        school_level="초등학교",
        school_key="타깃학교\n(초)",
    ),
    CategoryConfig(
        category="초6영어학원",
        grade_label="초6",
        grade_number=6,
        grade_en="GRADE 6",
        subject="영어",
        subject_en="ENGLISH",
        focus_label="어휘·문법 연결·독해·짧은 쓰기",
        workbook_name="초6 영어학원 원고.xlsx",
        representative_offset=201,
        school_level="초등학교",
        school_key="타깃학교\n(초)",
    ),
)


MATH_SIGNALS: tuple[TopicSignal, ...] = (
    TopicSignal(
        "concept_connection",
        "개념 연결",
        ("개념", "정의", "원리", "연결"),
        "앞 단원의 정의를 말로 설명한 뒤 현재 단원의 식에 적용할 수 있는지",
        "기본 문제와 조건이 바뀐 문제 사이에서 풀이 근거가 유지되는지",
        "정의 한 줄, 대표 예제, 변형 문제를 한 묶음으로 다시 확인하기",
        "하루 뒤 해설 없이 같은 개념을 설명하고 한 문제를 재풀이하기",
    ),
    TopicSignal(
        "calculation_accuracy",
        "계산 정확도",
        ("계산", "부호", "실수", "연산"),
        "부호와 괄호를 옮기는 단계에서 같은 계산 실수가 반복되는지",
        "오답지의 중간 계산과 답을 고친 뒤의 재풀이가 일치하는지",
        "계산 과정을 한 줄씩 분리하고 틀린 단계에 이유를 표시하기",
        "짧은 계산 세트를 시간보다 정확도 기준으로 다시 풀기",
    ),
    TopicSignal(
        "equation_expression",
        "식과 방정식 표현",
        ("방정식", "부등식", "식", "문자"),
        "문제의 조건을 문자와 식으로 바꾸는 첫 단계가 정확한지",
        "식을 세우지 못한 문제와 식은 맞지만 풀이가 끊긴 문제를 구분했는지",
        "조건에 밑줄을 긋고 문자 정의부터 식 완성까지 순서대로 적기",
        "유사 조건의 문제를 골라 식을 세우는 과정만 다시 연습하기",
    ),
    TopicSignal(
        "function_graph",
        "함수와 그래프 해석",
        ("함수", "그래프", "좌표", "변화"),
        "식, 표, 그래프가 같은 관계를 나타낸다는 점을 연결해 설명하는지",
        "좌표를 잘못 읽은 경우와 변화 관계를 이해하지 못한 경우를 나눴는지",
        "한 관계를 식과 표와 그래프로 각각 바꾸어 표현하기",
        "그래프의 축과 단위를 먼저 확인한 뒤 변화 이유를 한 문장으로 적기",
    ),
    TopicSignal(
        "geometry_reasoning",
        "도형 추론",
        ("도형", "각", "삼각형", "증명", "기하"),
        "그림에 주어진 조건과 스스로 추정한 내용을 구분하는지",
        "사용한 성질의 이름과 그 성질이 적용되는 위치를 짚을 수 있는지",
        "조건 표시, 사용할 성질, 결론의 순서로 풀이를 구성하기",
        "도형을 다시 그린 뒤 조건만 보고 필요한 성질을 말로 설명하기",
    ),
    TopicSignal(
        "word_problem",
        "문장제 조건 해석",
        ("문장제", "조건", "해석", "상황"),
        "긴 문장에서 수량과 관계를 나눠 읽고 필요한 조건만 고르는지",
        "문제를 읽지 못한 오답과 식을 잘못 세운 오답을 따로 표시했는지",
        "조건을 표나 짧은 식으로 바꾼 뒤 풀이를 시작하기",
        "문장제 한 문제를 읽고 조건, 구할 것, 사용할 식을 따로 적기",
    ),
    TopicSignal(
        "written_solution",
        "서술형 풀이 과정",
        ("서술형", "풀이과정", "과정", "근거"),
        "정답뿐 아니라 필요한 조건과 계산 근거가 풀이에 남아 있는지",
        "감점된 답안에서 빠진 식과 설명을 채점 기준에 맞춰 찾았는지",
        "풀이를 식, 이유, 결론의 세 단계로 나누어 쓰기",
        "완성한 답안을 소리 내어 설명하며 빠진 근거를 보완하기",
    ),
    TopicSignal(
        "error_relearning",
        "오답 재학습",
        ("오답", "재풀이", "복습", "틀린"),
        "답만 고친 문제와 원인을 설명하고 다시 푼 문제를 구분했는지",
        "같은 유형을 며칠 뒤 다시 풀었을 때 풀이가 유지되는지",
        "오답 원인을 개념, 해석, 계산으로 나눈 뒤 유사 문제를 재풀이하기",
        "다시 볼 날짜를 정하고 해설 없이 해결됐는지 기록하기",
    ),
    TopicSignal(
        "time_management",
        "시험 시간 배분",
        ("시간", "시험", "속도", "배분"),
        "아는 문제에서도 시간이 길어진 구간과 막혀서 멈춘 구간이 어디인지",
        "문항별 풀이 시간과 마지막 검산에 남긴 시간이 기록되어 있는지",
        "문제 난도를 나눠 1차 풀이와 재검토 순서를 연습하기",
        "짧은 범위 시험을 풀고 문항별 소요 시간과 이유를 함께 적기",
    ),
)


ENGLISH_SIGNALS: tuple[TopicSignal, ...] = (
    TopicSignal(
        "vocabulary_accumulation",
        "어휘 누적",
        ("어휘", "단어", "암기", "뜻"),
        "단어의 한글 뜻만 외웠는지 문장 안 쓰임까지 구분하는지",
        "교과서와 시험지에서 반복해 막힌 단어가 별도로 기록되어 있는지",
        "새 단어, 예문, 누적 복습 날짜를 한 묶음으로 관리하기",
        "짧은 간격으로 뜻과 쓰임을 번갈아 확인하고 틀린 단어만 다시 보기",
    ),
    TopicSignal(
        "sentence_structure",
        "문장 구조와 문법 적용",
        ("문법", "문장", "구문", "구조"),
        "주어와 동사를 찾은 뒤 수식 관계를 끊어 읽을 수 있는지",
        "문법 규칙을 아는 문제와 실제 문장에서 적용하지 못한 문제를 나눴는지",
        "한 문장을 구조 표시, 해석, 문법 근거의 순서로 분석하기",
        "교과서 한 문장을 바꿔 쓰며 같은 문법이 유지되는지 확인하기",
    ),
    TopicSignal(
        "reading_evidence",
        "독해 근거 찾기",
        ("독해", "지문", "근거", "주제"),
        "답을 고른 문장 근거와 선택지를 지운 이유를 함께 말할 수 있는지",
        "어휘 부족, 구조 해석, 내용 추론 가운데 독해가 끊긴 지점을 표시했는지",
        "문단의 핵심 문장과 선택지 근거를 연결해 표시하기",
        "짧은 지문을 다시 읽고 정답 근거를 한 문장으로 설명하기",
    ),
    TopicSignal(
        "school_material",
        "교과서·학교 자료",
        ("교과서", "본문", "학교", "프린트"),
        "교과서 본문과 학교 프린트에서 시험 범위가 어떻게 겹치는지",
        "본문 암기와 문장 변형 문제의 준비 수준이 구분되어 있는지",
        "본문 표현, 학교 프린트, 시험 변형 문항을 순서대로 연결하기",
        "학교 자료에서 바뀐 표현을 찾아 원문과 나란히 비교하기",
    ),
    TopicSignal(
        "written_response",
        "서술형과 영작",
        ("서술형", "영작", "쓰기", "수행평가"),
        "요구 조건을 모두 반영해 문장을 완성하고 스스로 검토하는지",
        "내용 오류와 어휘 선택, 문법 오류를 서로 다른 표시로 남겼는지",
        "조건 확인, 초안 작성, 문법 검토의 세 단계로 답안을 만들기",
        "틀린 문장을 고쳐 쓴 뒤 같은 구조로 새 문장을 한 개 만들기",
    ),
    TopicSignal(
        "listening_detail",
        "듣기 세부 정보",
        ("듣기", "발음", "대화", "청취"),
        "소리를 놓친 구간과 표현을 알아도 의미를 연결하지 못한 구간이 다른지",
        "오답 문항에서 핵심 표현과 답의 단서가 들린 위치를 찾았는지",
        "짧게 듣고 핵심어를 적은 뒤 대본으로 놓친 이유를 확인하기",
        "같은 대화를 다시 들으며 시간, 장소, 의도를 구분해 기록하기",
    ),
    TopicSignal(
        "error_pattern",
        "영어 오답 유형",
        ("오답", "틀린", "유형", "약점"),
        "틀린 문제가 어휘, 문법, 독해 중 어느 원인에서 시작됐는지",
        "답을 고친 뒤 같은 근거로 유사 문제를 해결할 수 있는지",
        "오답을 원인별로 묶고 가장 자주 나온 유형부터 재학습하기",
        "며칠 뒤 해설 없이 다시 풀어 같은 판단 오류가 남았는지 확인하기",
    ),
    TopicSignal(
        "study_routine",
        "누적 복습 습관",
        ("복습", "누적", "습관", "계획"),
        "새 진도와 지난 범위 복습이 한 주 계획에 함께 들어 있는지",
        "계획한 분량과 실제 완료한 분량의 차이가 기록되어 있는지",
        "새 학습과 누적 복습의 분량을 나누고 완료 표시를 남기기",
        "주말에 미완료 항목과 다음 주 첫 복습 대상을 정리하기",
    ),
)


ELEMENTARY_MATH_SIGNALS: tuple[TopicSignal, ...] = (
    TopicSignal(
        "place_value_calculation",
        "수 감각과 계산 원리",
        ("연산", "계산", "수 감각", "자릿값"),
        "자릿값과 계산 순서를 말로 설명한 뒤 같은 원리를 새 문제에 적용하는지",
        "답만 고친 것이 아니라 계산 과정에서 멈춘 단계가 표시되어 있는지",
        "계산 과정을 한 줄씩 나누고 각 단계의 이유를 짧게 설명하기",
        "하루 뒤 비슷한 계산 두 문제를 풀고 틀린 이유를 한 문장으로 남기기",
    ),
    TopicSignal(
        "multiplication_division",
        "곱셈·나눗셈 관계",
        ("곱셈", "나눗셈", "구구단", "몫", "나머지"),
        "곱셈과 나눗셈의 관계를 그림이나 식으로 바꾸어 설명할 수 있는지",
        "계산 실수와 문제 상황을 잘못 이해한 경우가 따로 표시되어 있는지",
        "같은 수 관계를 곱셈식과 나눗셈식으로 각각 써 보기",
        "생활 속 수량 한 가지를 골라 곱셈식과 나눗셈식으로 바꾸어 보기",
    ),
    TopicSignal(
        "fraction_foundation",
        "분수 개념과 연산",
        ("분수", "분모", "분자", "전체", "부분"),
        "전체와 부분의 관계를 그림, 말, 분수로 차례대로 나타낼 수 있는지",
        "분모와 분자의 뜻을 외운 답과 실제 그림에 적용한 답이 일치하는지",
        "한 그림을 여러 분수 표현으로 바꾸고 이유를 말로 설명하기",
        "집에서 물건이나 그림을 나누어 분수 한 가지를 직접 만들어 보기",
    ),
    TopicSignal(
        "decimal_operations",
        "소수 개념과 연산",
        ("소수", "소수점", "소수의", "자릿수"),
        "소수점 위치와 자릿값을 설명한 뒤 계산 결과의 크기를 예상할 수 있는지",
        "소수 계산에서 자릿값을 맞춘 기록과 계산 뒤 검산한 흔적이 남아 있는지",
        "소수점과 같은 자릿수를 세로로 맞추고 계산 이유를 한 줄씩 적기",
        "생활 속 소수 자료 하나를 골라 크기를 비교하고 계산 과정을 설명해 보기",
    ),
    TopicSignal(
        "ratio_proportion",
        "비와 비율",
        ("비율", "비례", "비의", "백분율", "비례식"),
        "두 양의 관계를 비로 나타내고 기준량과 비교하는 양을 구분할 수 있는지",
        "비율 문제에서 무엇을 기준으로 비교했는지 식과 설명에 함께 표시했는지",
        "기준량, 비교하는 양, 비율의 순서로 조건을 표와 식에 정리하기",
        "생활 속 비율 한 가지를 찾아 기준량과 비교하는 양을 말로 설명해 보기",
    ),
    TopicSignal(
        "geometry_measurement",
        "도형과 측정",
        ("도형", "길이", "각도", "들이", "무게", "측정"),
        "단위와 도형의 성질을 구분하고 문제에 맞는 기준을 고를 수 있는지",
        "측정값의 단위를 빠뜨린 경우와 계산 자체가 틀린 경우를 구분했는지",
        "문제의 단위에 표시하고 그림이나 표로 조건을 다시 정리하기",
        "주변 물건 하나를 재거나 관찰한 뒤 사용한 단위와 이유를 적기",
    ),
    TopicSignal(
        "elementary_word_problem",
        "문장제 해석",
        ("문장제", "서술", "조건", "문제 해결", "응용"),
        "문장에서 주어진 수와 구할 내용을 나누어 표시할 수 있는지",
        "식을 세우기 전 문제 상황을 자신의 말이나 그림으로 바꾸었는지",
        "조건, 구할 내용, 식, 답의 순서로 풀이를 정리하기",
        "짧은 문장제 한 문제를 읽고 식을 세운 이유까지 말해 보기",
    ),
    TopicSignal(
        "elementary_explanation",
        "풀이 설명",
        ("설명", "풀이", "과정", "이유"),
        "정답과 함께 어떤 순서로 해결했는지를 말로 설명할 수 있는지",
        "답을 맞힌 문제에서도 빠진 조건이나 설명이 없는지",
        "풀이를 계산, 이유, 답의 세 부분으로 나누어 적기",
        "가장 자신 있는 문제 하나를 골라 풀이 이유를 가족에게 설명해 보기",
    ),
    TopicSignal(
        "elementary_math_review",
        "오답 재확인 습관",
        ("오답", "복습", "다시", "습관", "확인"),
        "고친 문제를 며칠 뒤 해설 없이 다시 해결할 수 있는지",
        "틀린 이유와 다시 확인할 날짜가 학습 기록에 함께 남아 있는지",
        "오답마다 원인 한 가지와 다음 확인일 한 가지를 적기",
        "주말에 한 주 오답 가운데 두 문제만 골라 스스로 다시 풀기",
    ),
)


ELEMENTARY_ENGLISH_SIGNALS: tuple[TopicSignal, ...] = (
    TopicSignal(
        "phonics_sound",
        "소리와 철자 연결",
        ("파닉스", "소리", "철자", "알파벳", "발음"),
        "글자를 보고 소리를 말하거나 들은 소리를 철자로 고를 수 있는지",
        "외운 단어와 처음 보는 비슷한 단어에서 같은 소리를 찾는지",
        "같은 소리가 나는 단어를 묶고 소리 내어 읽기",
        "짧은 단어 세 개를 듣고 철자와 소리를 함께 확인하기",
    ),
    TopicSignal(
        "elementary_vocabulary",
        "어휘 확장",
        ("단어", "어휘", "뜻", "암기"),
        "단어의 뜻뿐 아니라 그림이나 짧은 문장에서 쓰임을 설명하는지",
        "바로 기억한 단어와 며칠 뒤에도 읽고 뜻을 말한 단어가 구분되어 있는지",
        "단어, 그림, 짧은 예문을 한 묶음으로 정리하기",
        "하루 뒤 단어를 보고 뜻과 짧은 문장을 하나씩 말해 보기",
    ),
    TopicSignal(
        "basic_sentence",
        "기본 문장 구조",
        ("문장", "어순", "주어", "동사", "문법"),
        "누가 무엇을 하는지 기본 어순을 구분해 문장을 읽을 수 있는지",
        "단어 뜻은 알지만 문장 순서를 바꾸면 해석이 멈추는지",
        "짧은 문장을 주어, 동사, 나머지 말로 나누어 표시하기",
        "교과서 문장 하나의 낱말을 바꾸어 새 문장을 만들어 보기",
    ),
    TopicSignal(
        "elementary_reading",
        "글의 흐름과 독해 근거",
        ("읽기", "독해", "본문", "내용", "이해"),
        "짧은 글을 읽고 중심 인물과 일어난 일을 자신의 말로 설명하는지",
        "답을 고른 근거가 글의 어느 문장에 있는지 표시했는지",
        "문단마다 핵심 낱말 한 개와 내용을 보여 주는 문장 한 개를 고르기",
        "짧은 글을 소리 내어 읽고 가장 중요한 내용을 한 문장으로 말하기",
    ),
    TopicSignal(
        "listening_speaking",
        "듣기와 말하기",
        ("듣기", "말하기", "대화", "표현"),
        "짧은 안내나 대화를 듣고 필요한 정보를 구분해 말할 수 있는지",
        "들은 표현을 그대로 따라 한 경우와 상황에 맞게 바꾼 경우가 구분되는지",
        "짧은 대화를 듣고 핵심 낱말과 대답을 차례로 적기",
        "배운 표현 한 가지를 가족과 짧은 질문·대답으로 연습하기",
    ),
    TopicSignal(
        "elementary_writing",
        "짧은 문장 쓰기",
        ("쓰기", "영작", "문장 만들기", "일기"),
        "배운 낱말을 활용해 뜻이 분명한 짧은 문장을 완성할 수 있는지",
        "대문자, 띄어쓰기, 문장부호와 어순 오류를 따로 표시했는지",
        "말하고 싶은 내용을 정한 뒤 낱말 배열과 문장부호를 확인하기",
        "교과서 문장을 참고해 자신의 이야기로 한 문장 바꾸어 쓰기",
    ),
    TopicSignal(
        "elementary_english_review",
        "짧은 반복 학습",
        ("복습", "반복", "습관", "학습 계획"),
        "읽기, 듣기, 쓰기 가운데 어떤 활동을 다시 해야 하는지 기록되어 있는지",
        "계획한 짧은 활동과 실제로 마친 활동이 구분되어 있는지",
        "읽기, 단어, 문장 활동을 짧게 나누고 완료 표시를 남기기",
        "주말에 가장 어려웠던 표현 한 가지를 다시 읽고 말해 보기",
    ),
)


UPPER_ELEMENTARY_MATH_SIGNALS: tuple[TopicSignal, ...] = (
    TopicSignal(
        "upper_fraction_decimal",
        "분수·소수 연산",
        ("분수", "소수", "약분", "통분", "분모", "분자", "소수점", "연산"),
        "분수와 소수의 크기를 먼저 예상하고 계산 원리를 단계별로 설명할 수 있는지",
        "약분·통분이나 소수점 위치에서 생긴 오류와 계산 실수를 따로 표시했는지",
        "계산 전 예상값을 적고 각 단계의 근거와 검산 결과를 차례로 남기기",
        "비슷한 분수·소수 문제 두 개를 골라 풀이 방법의 공통점을 설명해 보기",
    ),
    TopicSignal(
        "upper_ratio_proportion",
        "비와 비율",
        ("비율", "비례", "비의", "백분율", "비례식", "기준량"),
        "기준량과 비교하는 양을 구분한 뒤 두 양의 관계를 식으로 나타낼 수 있는지",
        "비율 문제에서 기준이 달라진 경우와 계산만 틀린 경우를 나누어 기록했는지",
        "기준량, 비교하는 양, 비율을 표에 정리한 뒤 식과 답을 연결하기",
        "생활 속 비율 자료 하나를 골라 기준량과 비교하는 양을 말로 설명해 보기",
    ),
    TopicSignal(
        "upper_geometry_measurement",
        "도형과 측정 관계",
        ("도형", "각도", "넓이", "부피", "둘레", "단위", "측정"),
        "도형의 성질과 길이·넓이·부피의 단위를 문제 조건에 맞게 연결할 수 있는지",
        "공식을 잘못 고른 경우와 단위 변환에서 막힌 경우를 구분해 표시했는지",
        "그림에 조건과 단위를 표시하고 사용한 공식의 이유를 한 문장으로 적기",
        "주변 물체 하나를 골라 어떤 값을 재고 어떤 단위를 쓸지 설명해 보기",
    ),
    TopicSignal(
        "upper_word_problem",
        "조건 해석과 문제 해결",
        ("문장제", "조건", "문제 해결", "응용", "서술", "해석"),
        "긴 문제에서 주어진 조건과 구할 내용을 분리하고 필요한 식을 세울 수 있는지",
        "조건을 놓친 문제와 풀이 전략을 정하지 못한 문제를 서로 구분했는지",
        "조건, 구할 내용, 풀이 계획, 계산, 답의 순서로 해결 과정을 정리하기",
        "문장제 한 문제를 골라 계산보다 먼저 해결 순서를 말로 설명해 보기",
    ),
    TopicSignal(
        "upper_math_explanation",
        "풀이 근거 설명",
        ("설명", "풀이", "과정", "이유", "근거"),
        "정답과 함께 어떤 개념과 순서로 해결했는지를 문장으로 설명할 수 있는지",
        "맞힌 문제에서도 생략한 조건이나 근거가 없는 풀이를 표시했는지",
        "풀이마다 사용한 개념, 계산 단계, 결론을 짧은 문장으로 연결하기",
        "가장 어려웠던 문제 하나의 풀이 근거를 해설 없이 다시 설명해 보기",
    ),
    TopicSignal(
        "upper_math_review",
        "누적 오답 재확인",
        ("오답", "복습", "다시", "누적", "확인", "재풀이"),
        "고친 문제를 며칠 뒤 해설 없이 다시 풀고 같은 오류가 줄었는지 확인하는지",
        "오답 원인과 재풀이 날짜, 두 번째 결과가 한 기록에 함께 남아 있는지",
        "오답을 개념, 조건 해석, 계산으로 나누고 다음 확인일을 정하기",
        "주말에 누적 오답 두 문제를 골라 풀이 근거까지 다시 적어 보기",
    ),
)


UPPER_ELEMENTARY_ENGLISH_SIGNALS: tuple[TopicSignal, ...] = (
    TopicSignal(
        "upper_vocabulary_context",
        "문맥 속 어휘",
        ("어휘", "단어", "뜻", "문맥", "암기"),
        "단어 뜻을 외우는 데서 그치지 않고 문장 안 쓰임과 함께 설명할 수 있는지",
        "바로 기억한 단어와 며칠 뒤 문맥 속에서도 이해한 단어를 구분했는지",
        "단어, 뜻, 문맥 단서, 새 예문을 한 묶음으로 정리하기",
        "새 단어 세 개를 골라 교과서와 다른 짧은 문장에 적용해 보기",
    ),
    TopicSignal(
        "upper_grammar_application",
        "문법의 문장 적용",
        ("문법", "문장", "어순", "동사", "시제", "구문"),
        "문법 규칙을 설명한 뒤 처음 보는 문장의 구조와 의미에 적용할 수 있는지",
        "규칙은 알지만 문장 해석이나 쓰기에서 적용하지 못한 사례를 표시했는지",
        "문장의 핵심 구조를 나누고 같은 규칙으로 새 문장 한 개를 만들기",
        "교과서 문장 하나의 시제나 주어를 바꾸고 달라진 형태를 확인해 보기",
    ),
    TopicSignal(
        "upper_reading_evidence",
        "독해 흐름과 답의 근거",
        ("독해", "읽기", "근거", "중심", "요지", "추론", "본문"),
        "글의 중심 내용과 문단 관계를 설명하고 답을 고른 근거 문장을 찾을 수 있는지",
        "어휘, 문장 구조, 내용 추론 가운데 독해가 끊긴 지점을 구분했는지",
        "문단마다 핵심 문장과 연결 표현을 표시한 뒤 답의 근거를 한 줄로 적기",
        "짧은 글을 읽고 중심 내용과 그 근거가 된 문장을 함께 말해 보기",
    ),
    TopicSignal(
        "upper_sentence_writing",
        "문장 쓰기와 서술형",
        ("쓰기", "영작", "서술형", "작문", "문장 만들기"),
        "주어진 어휘와 문법을 활용해 뜻이 분명한 문장을 완성할 수 있는지",
        "내용 오류와 어순·시제·문장부호 오류를 따로 표시했는지",
        "쓸 내용을 먼저 정하고 문장 구조와 표현을 차례로 점검하기",
        "배운 표현 한 가지를 활용해 자신의 경험을 짧은 문장으로 써 보기",
    ),
    TopicSignal(
        "upper_listening_speaking",
        "듣기와 말하기 연결",
        ("듣기", "말하기", "대화", "표현", "발표"),
        "대화의 목적과 핵심 정보를 듣고 상황에 맞는 표현으로 답할 수 있는지",
        "들은 표현을 따라 한 경우와 자신의 문장으로 바꾼 경우를 구분했는지",
        "대화의 핵심 정보와 대답에 필요한 표현을 차례로 정리하기",
        "배운 표현을 활용해 짧은 질문과 대답을 소리 내어 연습해 보기",
    ),
    TopicSignal(
        "upper_english_review",
        "누적 오답과 복습 계획",
        ("오답", "복습", "누적", "반복", "학습 계획", "확인"),
        "어휘, 문법, 독해 오답을 원인별로 나누고 다시 확인할 순서를 정했는지",
        "계획한 복습과 실제로 마친 활동, 다음 확인일이 함께 기록되어 있는지",
        "오답을 어휘, 구조, 근거 찾기로 나누고 우선 복습 항목을 정하기",
        "주말에 가장 자주 틀린 유형 하나를 골라 근거까지 다시 확인해 보기",
    ),
)


class _VisibleTextParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.parts: list[str] = []
        self._ignored_depth = 0

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag.lower() in {"script", "style", "noscript"}:
            self._ignored_depth += 1

    def handle_endtag(self, tag: str) -> None:
        if tag.lower() in {"script", "style", "noscript"} and self._ignored_depth:
            self._ignored_depth -= 1

    def handle_data(self, data: str) -> None:
        if not self._ignored_depth:
            self.parts.append(data)


def visible_source_text(raw_html: str) -> str:
    parser = _VisibleTextParser()
    parser.feed(raw_html)
    parser.close()
    return re.sub(r"\s+", " ", html.unescape(" ".join(parser.parts))).strip()


def stable_number(*parts: object) -> int:
    payload = "|".join(str(part) for part in parts).encode("utf-8")
    return int.from_bytes(hashlib.sha256(payload).digest()[:8], "big")


def choose(items: Sequence[str], *seed_parts: object) -> str:
    if not items:
        raise ValueError("선택할 문장 후보가 없습니다.")
    return items[stable_number(*seed_parts) % len(items)]


def has_batchim(value: str) -> bool:
    """Return whether the spoken final character takes a consonant-form particle."""
    text = re.sub(r"[\s'\"’”〉》」』)\]}]+$", "", (value or "").strip())
    if not text:
        return True
    char = text[-1]
    code = ord(char)
    if 0xAC00 <= code <= 0xD7A3:
        return (code - 0xAC00) % 28 != 0
    if char.isdigit():
        # 영, 일, 삼, 육, 칠, 팔 have a final consonant; 이, 사, 오, 구 do not.
        return int(char) in {0, 1, 3, 6, 7, 8}
    return True


def eun_neun(value: str) -> str:
    return "은" if has_batchim(value) else "는"


def i_ga(value: str) -> str:
    return "이" if has_batchim(value) else "가"


def eul_reul(value: str) -> str:
    return "을" if has_batchim(value) else "를"


def gwa_wa(value: str) -> str:
    return "과" if has_batchim(value) else "와"


def ina_na(value: str) -> str:
    return "이나" if has_batchim(value) else "나"


def read_workbook_cells(path: Path) -> list[str]:
    if not path.exists():
        raise FileNotFoundError(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        values = [
            str(row[0]).strip()
            for row in worksheet.iter_rows(min_col=1, max_col=1, values_only=True)
            if row[0] is not None and str(row[0]).strip()
        ]
    finally:
        workbook.close()
    if len(values) != EXPECTED_PAGE_COUNT:
        raise ValueError(f"A열 비어 있지 않은 셀 개수 오류: {path.name}={len(values)}")
    return values


def signals_for(config: CategoryConfig) -> tuple[TopicSignal, ...]:
    if config.is_elementary and config.grade_number >= 5 and config.subject == "수학":
        return UPPER_ELEMENTARY_MATH_SIGNALS
    if config.is_elementary and config.grade_number >= 4 and config.subject == "영어":
        return UPPER_ELEMENTARY_ENGLISH_SIGNALS
    if config.is_elementary and config.subject == "수학":
        return ELEMENTARY_MATH_SIGNALS
    if config.is_elementary and config.subject == "영어":
        return ELEMENTARY_ENGLISH_SIGNALS
    if config.subject == "수학":
        return MATH_SIGNALS
    if config.subject == "영어":
        return ENGLISH_SIGNALS
    raise ValueError(f"지원하지 않는 과목: {config.subject}")


def select_signals(raw_html: str, config: CategoryConfig, local: str) -> SourceSelection:
    # The reused workbook contributes only aggregate keyword signals.  No title,
    # locality, paragraph, sentence, or markup from it is carried into a page.
    source_text = visible_source_text(raw_html).casefold()
    ranked: list[tuple[int, int, TopicSignal]] = []
    for signal in signals_for(config):
        score = sum(source_text.count(keyword.casefold()) for keyword in signal.keywords)
        tie_breaker = stable_number(config.category, local, signal.code, "source-signal")
        ranked.append((-score, tie_breaker, signal))
    ranked.sort(key=lambda item: (item[0], item[1]))
    if config.is_elementary and config.subject == "수학":
        process_codes = {
            "elementary_explanation",
            "elementary_math_review",
            "upper_math_explanation",
            "upper_math_review",
        }
        curriculum = [item[2] for item in ranked if item[2].code not in process_codes]
        process = [item[2] for item in ranked if item[2].code in process_codes]
        chosen = [*curriculum[:2], process[0]]
    else:
        chosen = [item[2] for item in ranked[:3]]
    return SourceSelection(chosen[0], chosen[1], chosen[2])


def middle_school_names(row: dict[str, str]) -> list[str]:
    """Normalize the middle-school column without inventing a school reference."""
    raw = row.get("타깃학교\n(중)", "").strip()
    if not raw:
        return []
    # A few source cells use spaces or periods instead of commas.  Join a spaced
    # school-type suffix before tokenizing, then normalize the long suffix.
    raw = re.sub(r"([^\s,/·.]+)\s+중학교(?=$|[\s,/·.])", r"\1중학교", raw)
    result: list[str] = []
    for item in re.split(r"[,/·.\s]+", raw):
        name = re.sub(r"중학교$", "중", item.strip())
        if len(name) < 2 or not name.endswith("중") or name in result:
            continue
        result.append(name)
    return result


def elementary_school_names(row: dict[str, str]) -> list[str]:
    """Normalize the elementary-school column without inventing a school reference."""
    raw = row.get("타깃학교\n(초)", "").strip()
    if not raw:
        return []
    raw = re.sub(r"([^\s,/·.]+)\s+초등학교(?=$|[\s,/·.])", r"\1초등학교", raw)
    result: list[str] = []
    for item in re.split(r"[,/·.\s]+", raw):
        name = re.sub(r"초등학교$", "초", item.strip())
        if len(name) < 2 or not name.endswith("초") or name in result:
            continue
        result.append(name)
    return result


def school_names_for(row: dict[str, str], config: CategoryConfig) -> list[str]:
    if config.is_elementary:
        return elementary_school_names(row)
    return middle_school_names(row)


def fit_meta(candidates: Iterable[str], title: str) -> str:
    for candidate in candidates:
        value = re.sub(r"\s+", " ", candidate).strip()
        if 80 <= len(value) <= 155:
            return value
    lengths = [(len(re.sub(r"\s+", " ", item).strip()), item) for item in candidates]
    raise ValueError(f"메타설명 길이 후보 오류: {title}, lengths={[length for length, _ in lengths]}")


def optional_third(
    first: str,
    second: str,
    third: str,
    *,
    config: CategoryConfig,
    local: str,
    section_code: str,
) -> list[str]:
    paragraphs = [first, second]
    if stable_number(config.category, local, section_code, "paragraph-count") % 2:
        paragraphs.append(third)
    return paragraphs


CONTEXT_EVIDENCE: tuple[str, ...] = (
    "틀린 문제 옆에 적은 원인 메모",
    "해설을 가린 채 다시 해결한 결과",
    "교재에 남긴 질문 표시",
    "학교 범위표와 현재 진도",
    "한 주 과제의 완료 기록",
    "문항별 풀이 시간을 적은 표",
    "서술 과정에서 빠진 근거",
    "학생이 말로 설명한 해결 순서",
    "교과서 예문을 바꿔 쓴 결과",
    "며칠 뒤 재풀이한 기록",
    "시험지에서 고친 답의 이유",
    "수업 뒤 정리한 질문 목록",
    "단원별 자신감을 표시한 표",
    "같은 유형을 두 번째 해결한 흔적",
    "학교 안내문에 적힌 평가 범위",
    "가정에서 완료한 분량",
    "막힌 시점에 표시한 기호",
    "답을 고르며 남긴 근거 문장",
    "문제를 시작하기 전에 적은 조건",
    "주간 계획과 실제 수행량의 차이",
    "교재별 마지막 완료 지점",
    "다시 설명하기 어려웠던 부분",
    "채점 뒤 고쳐 쓴 과정",
    "학생이 스스로 정한 다음 질문",
    "처음 답과 고친 답 사이의 변화",
    "복습 날짜별로 달라진 정답 근거",
    "한 문항을 풀며 멈춘 단계",
    "수업 전에 표시해 둔 낯선 표현",
    "문제 유형별로 나눈 오류 원인",
    "누락된 가정 복습 항목",
    "시험 직후 기억나는 어려운 문항",
    "설명 없이 해결할 수 있었던 범위",
)

MATH_CONTEXT_EVIDENCE: tuple[str, ...] = tuple(
    {
        "교과서 예문을 바꿔 쓴 결과": "조건을 바꿔 다시 푼 결과",
        "수업 전에 표시해 둔 낯선 표현": "수업 전에 표시해 둔 낯선 조건",
    }.get(item, item)
    for item in CONTEXT_EVIDENCE
)

ENGLISH_CONTEXT_EVIDENCE: tuple[str, ...] = tuple(
    {
        "문제를 시작하기 전에 적은 조건": "지문을 읽으며 표시한 핵심 표현",
    }.get(item, item)
    for item in CONTEXT_EVIDENCE
)


def elementary_context_text(value: str) -> str:
    replacements = (
        ("학교 범위표", "학교 진도표"),
        ("시험지", "단원평가 자료"),
        ("시험 직후", "단원평가 뒤"),
        ("시험 전에", "다음 점검 전까지"),
        ("시험 범위", "학교 학습 범위"),
        ("시험일까지", "다음 점검일까지"),
        ("시험까지", "다음 점검까지"),
        ("학교 범위", "학교 진도"),
    )
    for source, target in replacements:
        value = value.replace(source, target)
    return value


ELEMENTARY_MATH_CONTEXT_EVIDENCE: tuple[str, ...] = tuple(
    elementary_context_text(item) for item in MATH_CONTEXT_EVIDENCE
)
ELEMENTARY_ENGLISH_CONTEXT_EVIDENCE: tuple[str, ...] = tuple(
    elementary_context_text(item) for item in ENGLISH_CONTEXT_EVIDENCE
)

CONTEXT_DECISIONS: tuple[str, ...] = (
    "첫 설명을 시작할 지점",
    "혼자 연습할 분량",
    "피드백을 받을 시점",
    "다음 질문의 우선순위",
    "시험 전에 끝낼 범위",
    "과제량을 줄이거나 늘릴 기준",
    "보충 설명이 필요한 개념",
    "다시 풀 문제의 개수",
    "학교 자료를 확인할 순서",
    "주간 최소 학습량",
    "첫 점검 날짜",
    "새 진도로 넘어갈 조건",
    "가정에서 확인할 한 가지 행동",
    "교재를 바꿀지 판단할 기준",
    "질문을 모아 전달할 방식",
    "반복 오류를 확인할 날짜",
    "시험 범위 안에서의 학습 비중",
    "직접 설명해 볼 항목",
    "주말에 보완할 작은 목표",
    "다음 확인 때 비교할 변화",
    "학교 진도와 복습의 비율",
    "시간을 더 배정할 유형",
    "현재 교재에서 남길 범위",
    "학부모에게 공유할 핵심 항목",
    "수업 후 바로 확인할 행동",
    "재학습이 끝났다고 볼 기준",
    "시험일까지 유지할 루틴",
    "학생이 선택할 복습 순서",
    "설명과 연습을 나눌 비중",
    "다음 주 학습량을 조절할 기준",
    "꾸준히 기록할 항목",
    "등록 전에 확인할 항목",
)

CONTEXT_MOMENTS: tuple[str, ...] = (
    "준비한 자료를 펼치기 전",
    "첫 설명을 들은 직후",
    "이틀 뒤 다시 볼 때",
    "주간 계획을 조정할 때",
    "학교 범위가 공지된 날",
    "한 단원의 연습을 마친 뒤",
    "과제를 제출하기 전",
    "다음 수업을 준비할 때",
    "시험지를 다시 펼쳤을 때",
    "가정 학습을 끝낸 저녁",
    "주말 복습을 시작할 때",
    "상담 답변을 정리하는 날",
    "같은 유형을 다시 만났을 때",
    "교재 진도를 표시한 뒤",
    "메모를 정리하기 전",
    "시험까지 남은 날을 셀 때",
    "피드백 내용을 옮겨 적을 때",
    "이번 기록을 마무리할 때",
    "학습량이 부담스럽다고 느낄 때",
    "한 주 실행 기록을 모은 뒤",
)

ELEMENTARY_CONTEXT_DECISIONS: tuple[str, ...] = tuple(
    elementary_context_text(item) for item in CONTEXT_DECISIONS
)
ELEMENTARY_CONTEXT_MOMENTS: tuple[str, ...] = tuple(
    elementary_context_text(item) for item in CONTEXT_MOMENTS
)


CONTEXT_PURPOSE_SLOT: dict[str, int] = {
    "intro": 0,
    "direct-answer": 2,
    "evidence-diagnosis": 5,
    "recommended-student": 8,
    "conditional-four-week": 11,
    "feedback-home": 14,
    "consult-checklist": 17,
    "faq": 20,
    "consultation-scenario": 25,
    "quick-answer": 27,
}


def naturalize_context_start(
    sentence: str,
    *,
    config: CategoryConfig,
    local: str,
    purpose: str,
    ordinal: int,
    sentence_index: int,
) -> str:
    """Vary page context without repeating a keyword label at every sentence start."""
    consultation = f"{local} {config.grade_label} {config.subject} 상담"
    learning = f"{local} {config.grade_label} {config.subject} 학습"
    rotation = stable_number(
        config.category,
        local,
        purpose,
        ordinal,
        sentence_index,
        "natural-context-start",
    )

    consultation_locative = (
        f"이때 {local} {config.grade_label} {config.subject} 상담에서 ",
        f"{local} 학부모가 확인하는 {config.grade_label} {config.subject} 상담에서 ",
        f"학생 자료를 다루는 {local} {config.grade_label} {config.subject} 상담에서 ",
        f"다음 계획을 정하는 {local} {config.grade_label} {config.subject} 상담에서 ",
        f"가정 기록을 확인하는 {local} {config.grade_label} {config.subject} 상담에서 ",
        f"학교 자료를 비교하는 {local} {config.grade_label} {config.subject} 상담에서 ",
        f"첫 진단을 준비하는 {local} {config.grade_label} {config.subject} 상담에서 ",
        f"복습 순서를 논의하는 {local} {config.grade_label} {config.subject} 상담에서 ",
        f"학부모 질문을 정리하는 {local} {config.grade_label} {config.subject} 상담에서 ",
        f"수업 조건을 비교하는 {local} {config.grade_label} {config.subject} 상담에서 ",
    )
    consultation_before = (
        f"{local}에서 {config.grade_label} {config.subject} 상담을 시작하기 전에는 ",
        f"{config.grade_label} {config.subject} 상담을 준비하는 {local} 학부모는 ",
        f"{config.subject} 상담 자료를 {local}의 {config.grade_label} 학생과 정리할 때는 ",
        f"방문 전 {local}에서 {config.grade_label} {config.subject} 상담 자료를 모을 때는 ",
    )
    consultation_possessive = (
        f"이번 {local} {config.grade_label} {config.subject} 상담의 ",
        f"학생 자료를 다루는 {local} {config.grade_label} {config.subject} 상담의 ",
        f"계획을 정리하는 {local} {config.grade_label} {config.subject} 상담의 ",
        f"학부모 질문을 모은 {local} {config.grade_label} {config.subject} 상담의 ",
        f"학교 일정을 반영할 {local} {config.grade_label} {config.subject} 상담의 ",
        f"첫 진단을 준비하는 {local} {config.grade_label} {config.subject} 상담의 ",
    )
    learning_locative = (
        f"이때 {local} {config.grade_label} {config.subject} 학습에서 ",
        f"{local} 학생의 {config.grade_label} {config.subject} 학습에서 ",
        f"복습 기록을 남기는 {local} {config.grade_label} {config.subject} 학습에서 ",
        f"다음 계획을 세우는 {local} {config.grade_label} {config.subject} 학습에서 ",
        f"가정에서 이어 가는 {local} {config.grade_label} {config.subject} 학습에서 ",
        f"학교 진도를 반영한 {local} {config.grade_label} {config.subject} 학습에서 ",
        f"학생 설명을 확인하는 {local} {config.grade_label} {config.subject} 학습에서 ",
        f"오답 기록을 점검하는 {local} {config.grade_label} {config.subject} 학습에서 ",
        f"주간 분량을 조정하는 {local} {config.grade_label} {config.subject} 학습에서 ",
        f"다음 질문을 남기는 {local} {config.grade_label} {config.subject} 학습에서 ",
    )
    learning_possessive = (
        f"이번 {local} {config.grade_label} {config.subject} 학습의 ",
        f"복습 기록을 남기는 {local} {config.grade_label} {config.subject} 학습의 ",
        f"다음 계획을 세우는 {local} {config.grade_label} {config.subject} 학습의 ",
        f"학교 진도를 반영한 {local} {config.grade_label} {config.subject} 학습의 ",
        f"학생 설명을 확인하는 {local} {config.grade_label} {config.subject} 학습의 ",
        f"주간 분량을 조정하는 {local} {config.grade_label} {config.subject} 학습의 ",
    )
    learning_record = (
        f"{local}에서 작성한 {config.grade_label} {config.subject} 학습 기록에 ",
        f"{config.grade_label} 학생이 {local}에서 남긴 {config.subject} 기록에 ",
        f"{config.subject} 학습을 위해 {local} {config.grade_label} 학생이 만든 기록에 ",
        f"학부모가 {local} {config.grade_label} {config.subject} 과정에서 확인할 기록에 ",
    )
    learning_material = (
        f"{local}에서 모은 {config.grade_label} {config.subject} 학습 자료에서 ",
        f"{config.grade_label} 학생이 {local}에서 준비한 {config.subject} 자료에서 ",
        f"{config.subject} 학습을 위해 {local} {config.grade_label} 학생이 정리한 자료에서 ",
        f"학부모가 {local} {config.grade_label} {config.subject} 과정에서 살펴볼 자료에서 ",
    )

    def replace_prefix(prefix: str, replacements: tuple[str, ...]) -> str | None:
        if not sentence.startswith(prefix):
            return None
        replacement = replacements[rotation % len(replacements)]
        return f"{replacement}{sentence[len(prefix):]}"

    rules: tuple[tuple[str, tuple[str, ...]], ...] = (
        (f"{consultation} 전에는 ", consultation_before),
        (f"{consultation} 전에 ", consultation_before),
        (
            f"{consultation} 뒤에는 ",
            (f"{local}에서 {config.grade_label} {config.subject} 상담을 마친 뒤에는 ",),
        ),
        (f"{consultation}에서는 ", consultation_locative),
        (f"{consultation}에서 ", consultation_locative),
        (
            f"{consultation}{eun_neun(consultation)} ",
            (f"{local}에서 {config.grade_label} {config.subject} 상담 내용을 살펴보면, ",),
        ),
        (f"{consultation}의 ", consultation_possessive),
        (
            f"{consultation}{i_ga(consultation)} ",
            (f"{local}에서 진행한 {config.grade_label} {config.subject} 상담이 ",),
        ),
        (
            f"{consultation}{eul_reul(consultation)} ",
            (f"{local}에서 {config.grade_label} {config.subject} 상담을 ",),
        ),
        (f"{learning} 기록에 ", learning_record),
        (f"{learning} 자료에서 ", learning_material),
        (f"{learning}에서는 ", learning_locative),
        (f"{learning}에서 ", learning_locative),
        (
            f"{learning}{eun_neun(learning)} ",
            (f"{local}에서 {config.grade_label} {config.subject} 학습 흐름을 살펴보면, ",),
        ),
        (f"{learning}의 ", learning_possessive),
        (
            f"{learning}{i_ga(learning)} ",
            (f"{local}에서 이어 가는 {config.grade_label} {config.subject} 학습이 ",),
        ),
        (
            f"{learning}{eul_reul(learning)} ",
            (f"{local}에서 {config.grade_label} {config.subject} 학습을 ",),
        ),
        (
            f"{local}의 {config.grade_label} {config.subject} 학습에서는 ",
            learning_locative,
        ),
        (
            f"{local} {config.grade_label} {config.subject} 자료에서 ",
            learning_material,
        ),
        (
            f"{consultation} ",
            (f"{local}에서 진행할 {config.grade_label} {config.subject} 상담 ",),
        ),
        (
            f"{learning} ",
            (f"{local}의 {config.grade_label} 학생을 위한 {config.subject} 학습 ",),
        ),
    )
    for prefix, replacements in rules:
        rewritten = replace_prefix(prefix, replacements)
        if rewritten is not None:
            return rewritten
    return sentence


def context_variation_sentence(
    config: CategoryConfig,
    local: str,
    purpose: str,
    ordinal: int,
) -> str:
    """Add one concrete, natural record-to-decision sentence per paragraph."""
    seed = (config.category, local, purpose, ordinal)
    slot = CONTEXT_PURPOSE_SLOT.get(purpose, 0) + ordinal
    page_rotation = stable_number(config.category, local, "context-rotation")
    if config.is_elementary:
        evidence_bank = (
            ELEMENTARY_MATH_CONTEXT_EVIDENCE
            if config.subject == "수학"
            else ELEMENTARY_ENGLISH_CONTEXT_EVIDENCE
        )
        decision_bank = ELEMENTARY_CONTEXT_DECISIONS
        moment_bank = ELEMENTARY_CONTEXT_MOMENTS
    else:
        evidence_bank = (
            MATH_CONTEXT_EVIDENCE
            if config.subject == "수학"
            else ENGLISH_CONTEXT_EVIDENCE
        )
        decision_bank = CONTEXT_DECISIONS
        moment_bank = CONTEXT_MOMENTS
    evidence = evidence_bank[(page_rotation + slot) % len(evidence_bank)]
    decision = decision_bank[
        (page_rotation // 3 + slot * 5) % len(decision_bank)
    ]
    moment = moment_bank[(page_rotation // 7 + slot * 3) % len(moment_bank)]
    consultation = f"{local} {config.grade_label} {config.subject} 상담"
    learning = f"{local} {config.grade_label} {config.subject} 학습"
    frames = (
        f"{consultation}에서는 {evidence}{eul_reul(evidence)} 살핀 뒤 {decision}{eul_reul(decision)} 구체화할 수 있습니다.",
        f"{consultation}에서 {moment} {evidence}{i_ga(evidence)} 무엇을 보여 주는지 파악하면 {decision}{eul_reul(decision)} 정하기 쉽습니다.",
        f"{learning}에서 {decision}{eul_reul(decision)} 정할 때는 {evidence}{eul_reul(evidence)} 먼저 찾아보는 편이 좋습니다.",
        f"{consultation}에서 {evidence}{eul_reul(evidence)} 확인한 뒤 이를 참고해 {decision}{eul_reul(decision)} 정할 수 있습니다.",
        f"{learning} 기록에 {evidence}, {decision}{eul_reul(decision)} 나란히 적어 두세요.",
        f"{local}의 {config.grade_label} {config.subject} 학습에서는 {moment} 결과만 보지 말고 {evidence}{eul_reul(evidence)} 바탕으로 {decision}{eul_reul(decision)} 정해 보세요.",
        f"{consultation}에서 {decision}{eul_reul(decision)} 정하려면 {evidence}{eun_neun(evidence)} 구체적인 판단 자료가 됩니다.",
        f"{learning}에서 {evidence}{eul_reul(evidence)} 살펴본 뒤 {decision}{eul_reul(decision)} 정하면 이후 실행 계획을 구체화할 수 있습니다.",
        f"{consultation}에서 {evidence}{eul_reul(evidence)} 함께 살펴본 뒤 {decision}{eul_reul(decision)} 조율해 보세요.",
        f"{learning}에서는 {evidence}{eul_reul(evidence)} 한 번 확인하는 데 그치지 말고 {moment} {decision}{eul_reul(decision)} 재검토해야 합니다.",
        f"{consultation} 전에 {evidence}{eul_reul(evidence)} 확인해 두면 {decision}{eul_reul(decision)} 더 분명하게 논의할 수 있습니다.",
        f"{local} {config.grade_label} {config.subject} 자료에서 {moment} 살펴볼 내용은 {evidence}이며, 이를 바탕으로 {decision}{eul_reul(decision)} 정할 수 있습니다.",
        f"{consultation}에서는 {decision}{eul_reul(decision)} 바로 확정하기보다 {evidence}{eul_reul(evidence)} 바탕으로 다시 살펴보세요.",
        f"{learning}에서 {evidence}{eul_reul(evidence)} 다시 확인해 {decision}도 유지할지 점검해 보세요.",
        f"{consultation}에서 {evidence}{eul_reul(evidence)} 확인한 뒤 {decision}{eul_reul(decision)} 정해 보세요.",
        f"{learning}에서 {decision}{eul_reul(decision)} 정할 때에는 {evidence}{i_ga(evidence)} 기록에 남아 있는지도 살펴보세요.",
    )
    sentence = choose(frames, *seed, "frame")
    sentence = naturalize_context_start(
        sentence,
        config=config,
        local=local,
        purpose=purpose,
        ordinal=ordinal,
        sentence_index=97,
    )
    if not all(token in sentence for token in (local, config.grade_label, config.subject)):
        raise ValueError(
            f"문맥 문장 필수 요소 누락: {config.category}/{local}/{purpose}/{ordinal}"
        )
    return sentence


def diversify_paragraphs(
    paragraphs: Sequence[str],
    *,
    config: CategoryConfig,
    local: str,
    purpose: str,
) -> list[str]:
    diversified: list[str] = []
    for ordinal, paragraph in enumerate(paragraphs):
        localized = localize_sentences(
            paragraph,
            config=config,
            local=local,
            purpose=purpose,
            ordinal=ordinal,
        )
        diversified.append(localized)
    return diversified


ANCHOR_PURPOSE_ORDER = (
    "quick-answer",
    "intro",
    "direct-answer",
    "evidence-diagnosis",
    "recommended-student",
    "conditional-four-week",
    "feedback-home",
    "consult-checklist",
    "faq",
)
ANCHOR_SECTION_CODES = {
    "direct-answer": "direct-answer",
    "evidence-diagnosis": "evidence-diagnosis",
    "recommended-student": "recommended-student",
    "conditional-four-week": "four-week-plan",
    "feedback-home": "feedback-home",
    "consult-checklist": "consult-checklist",
}


def rendered_paragraph_count(config: CategoryConfig, local: str, purpose: str) -> int:
    if purpose == "quick-answer":
        return 1
    if purpose == "intro":
        return 2
    if purpose == "faq":
        return 5
    section_code = ANCHOR_SECTION_CODES[purpose]
    return 2 + (
        stable_number(config.category, local, section_code, "paragraph-count") % 2
    )


def paragraph_anchor_variant(
    config: CategoryConfig,
    local: str,
    purpose: str,
    ordinal: int,
    *,
    kind: str,
) -> int:
    """Choose varied anchors without repeating across paragraph boundaries."""
    step = 1 if kind == "consultation" else 2
    previous_end: int | None = None
    for candidate_purpose in ANCHOR_PURPOSE_ORDER:
        start = stable_number(
            config.category,
            local,
            candidate_purpose,
            kind,
            "paragraph-anchor",
        ) % 5
        if previous_end is not None and start == previous_end:
            start = (start + 1) % 5
        count = rendered_paragraph_count(config, local, candidate_purpose)
        if candidate_purpose == purpose:
            return (start + ordinal * step) % 5
        previous_end = (start + (count - 1) * step) % 5
    raise ValueError(f"알 수 없는 문단 목적: {purpose}")


def localize_sentences(
    value: str,
    *,
    config: CategoryConfig,
    local: str,
    purpose: str,
    ordinal: int,
) -> str:
    """Replace repeated full search labels with readable stage references."""
    sentences = re.split(r"(?<=[.!?])\s+", value.strip())
    if not all(token in value for token in (local, config.grade_label, config.subject)):
        raise ValueError(
            "문단 문맥 누락: "
            f"{config.category}/{local}/{purpose}/{ordinal}"
        )
    consultation = f"{local} {config.grade_label} {config.subject} 상담"
    learning = f"{local} {config.grade_label} {config.subject} 학습"
    consultation_variant = paragraph_anchor_variant(
        config,
        local,
        purpose,
        ordinal,
        kind="consultation",
    )
    learning_variant = paragraph_anchor_variant(
        config,
        local,
        purpose,
        ordinal,
        kind="learning",
    )
    consultation_forms = (
        dict(before_long="상담 전에는", before="상담 전에", after="상담 뒤에는", loc_topic="상담에서는", loc="상담에서", gen="상담의", material="상담 자료", class_time="센터 수업 시간", preparing_family="상담을 준비하는 가정", topic="상담은", subject="상담이", object="상담을", space="상담 "),
        dict(before_long="이 상담 전에는", before="이 상담 전에", after="이 상담을 마친 뒤에는", loc_topic="이 상담에서는", loc="이 상담에서", gen="이 상담의", material="이 상담 자료", class_time="확인할 수업 시간", preparing_family="이 상담을 준비하는 가정", topic="이 상담은", subject="이 상담이", object="이 상담을", space="이 상담 "),
        dict(before_long="상담에 앞서", before="상담 전에", after="상담을 마친 뒤에는", loc_topic="상담 자리에서는", loc="상담 자리에서", gen="상담의", material="상담 자료", class_time="실제 수업 시간", preparing_family="상담을 앞둔 가정", topic="상담 자리는", subject="상담 자리가", object="상담 자리를", space="상담 자리 "),
        dict(before_long="첫 상담 전에는", before="첫 상담 전에", after="첫 상담 뒤에는", loc_topic="첫 상담에서는", loc="첫 상담에서", gen="첫 상담의", material="첫 상담 자료", class_time="첫 수업 시간", preparing_family="첫 상담을 준비하는 가정", topic="첫 상담은", subject="첫 상담이", object="첫 상담을", space="첫 상담 "),
        dict(before_long="진단 상담 전에는", before="진단 상담 전에", after="진단 상담 뒤에는", loc_topic="진단 상담에서는", loc="진단 상담에서", gen="진단 상담의", material="진단 상담 자료", class_time="안내받은 수업 시간", preparing_family="진단 상담을 준비하는 가정", topic="진단 상담은", subject="진단 상담이", object="진단 상담을", space="진단 상담 "),
    )[consultation_variant]
    consultation_fixed_forms = {
        "priority": "학습 우선순위",
        "diagnostic_material": "진단 자료",
        "conditional_example": "조건부 계획 예시",
        "actual_amount": "실제 학습 분량",
        "schedule_amount": "제안된 주차별 분량",
        "week_four": "계획의 넷째 주",
        "availability": "수업 가능 여부",
        "class_assignment": "실제 반 편성",
        "feedback_cycle": "수업 피드백 주기",
        "proposed_plan": "제안된 계획이",
        "suitability": "수업 적합성",
        "suitability_decision": "수업 적합 여부",
    }
    consultation_forms.update(consultation_fixed_forms)
    learning_forms = (
        dict(record="학습 기록에", material="학습 자료에서", feedback="학습 피드백", evidence_table="학습 근거표", next_progress="다음 학습 진도", home_review_topic="가정 복습은", loc_topic="학습에서는", loc="학습에서", gen="학습의", topic="학습은", subject="학습이", object="학습을", space="학습 "),
        dict(record="이 학습 기록에", material="이 학습 자료에서", feedback="이 학습의 피드백", evidence_table="이 학습의 근거표", next_progress="이후 학습 진도", home_review_topic="집에서 이어 갈 복습은", loc_topic="이 학습에서는", loc="이 학습에서", gen="이 학습의", topic="이 학습은", subject="이 학습이", object="이 학습을", space="이 학습 "),
        dict(record="수업 기록에", material="수업 자료에서", feedback="수업 피드백", evidence_table="수업 기록표", next_progress="다음 수업 진도", home_review_topic="수업 뒤 가정 복습은", loc_topic="수업에서는", loc="수업에서", gen="수업 계획의", topic="수업은", subject="수업이", object="수업을", space="수업 "),
        dict(record="공부 기록에", material="공부 자료에서", feedback="공부 피드백", evidence_table="공부 기록표", next_progress="다음 공부 범위", home_review_topic="집 공부는", loc_topic="공부할 때는", loc="공부 과정에서", gen="공부 과정의", topic="공부는", subject="공부가", object="공부를", space="공부 "),
        dict(record="복습 기록에", material="복습 자료에서", feedback="복습 피드백", evidence_table="복습 기록표", next_progress="다음 복습 범위", home_review_topic="가정에서 하는 복습은", loc_topic="복습에서는", loc="복습에서", gen="복습 과정의", topic="복습은", subject="복습이", object="복습을", space="복습 "),
    )[learning_variant]
    consultation_anchor = (
        f"{local}의 {config.grade_label} {config.subject} 상담",
        f"수업 조건을 확인하는 {local} {config.grade_label} {config.subject} 상담",
        f"학생 상황을 살피는 {local} {config.grade_label} {config.subject} 상담",
        f"상담 항목을 정리하는 {local} {config.grade_label} {config.subject} 상담",
        f"다음 계획을 비교하는 {local} {config.grade_label} {config.subject} 상담",
    )[consultation_variant]
    anchor_consultation_forms = {
        "before_long": (
            f"{consultation_anchor} 전에는",
            f"{consultation_anchor} 전에는",
            f"{consultation_anchor}에 앞서",
            f"{consultation_anchor} 전에는",
            f"{consultation_anchor} 전에는",
        )[consultation_variant],
        "before": f"{consultation_anchor} 전에",
        "after": (
            f"{consultation_anchor} 뒤에는",
            f"{consultation_anchor}{eul_reul(consultation_anchor)} 마친 뒤에는",
            f"{consultation_anchor}{eul_reul(consultation_anchor)} 마친 뒤에는",
            f"{consultation_anchor} 뒤에는",
            f"{consultation_anchor} 뒤에는",
        )[consultation_variant],
        "loc_topic": f"{consultation_anchor}에서는",
        "loc": f"{consultation_anchor}에서",
        "gen": f"{consultation_anchor}의",
        "material": f"{local} {config.grade_label} {config.subject} 상담 자료",
        "topic": f"{consultation_anchor}{eun_neun(consultation_anchor)}",
        "subject": f"{consultation_anchor}{i_ga(consultation_anchor)}",
        "object": f"{consultation_anchor}{eul_reul(consultation_anchor)}",
        "space": f"{consultation_anchor} ",
    }
    anchor_consultation_forms["class_time"] = (
        f"{local} {config.grade_label} {config.subject} 수업 시간"
    )
    anchor_consultation_forms["preparing_family"] = (
        f"{local} {config.grade_label} {config.subject} 상담을 준비하는 가정"
    )
    anchor_consultation_forms.update(
        {
            "priority": f"{local} {config.grade_label} {config.subject} 학습 우선순위",
            "diagnostic_material": f"{local} {config.grade_label} {config.subject} 진단 자료",
            "conditional_example": f"{local} {config.grade_label} {config.subject} 조건부 계획 예시",
            "actual_amount": f"{local} {config.grade_label} {config.subject} 실제 학습 분량",
            "schedule_amount": f"{local} {config.grade_label} {config.subject} 제안 계획의 주차별 분량",
            "week_four": f"{local} {config.grade_label} {config.subject} 계획의 넷째 주",
            "availability": f"{local} {config.grade_label} {config.subject} 수업 가능 여부",
            "class_assignment": f"{local} {config.grade_label} {config.subject} 실제 반 편성",
            "feedback_cycle": f"{local} {config.grade_label} {config.subject} 수업 피드백 주기",
            "proposed_plan": f"{local} {config.grade_label} {config.subject} 상담에서 제안된 계획이",
            "suitability": f"{local} {config.grade_label} {config.subject} 수업 적합성",
            "suitability_decision": f"{local} {config.grade_label} {config.subject} 수업 적합 여부",
        }
    )
    learning_anchor = (
        f"{local} {config.grade_label} {config.subject} 학습",
        f"현재 수준을 확인하는 {local} {config.grade_label} {config.subject} 학습",
        f"학생 자료를 살피는 {local} {config.grade_label} {config.subject} 수업",
        f"복습 기준을 정리하는 {local} {config.grade_label} {config.subject} 공부",
        f"기록을 살피는 {local} {config.grade_label} {config.subject} 학습",
    )[learning_variant]
    # Each grammatical form is explicit.  Slicing a generic prefix from forms
    # such as “공부할 때는” or “상담에 앞서” produced malformed public copy
    # (for example “상담 자리서” and “학생의 수학 공부할 때는”).
    anchor_learning_forms = {
        "record": f"{learning_anchor} 기록에",
        "material": f"{learning_anchor} 자료에서",
        "feedback": f"{learning_anchor} 피드백",
        "evidence_table": (
            f"{local} {config.grade_label} {config.subject} 학습 근거표",
            f"{local} {config.grade_label} {config.subject} 학습 기록표",
            f"{local} {config.grade_label} {config.subject} 수업 기록표",
            f"{local} {config.grade_label} {config.subject} 공부 기록표",
            f"{local} {config.grade_label} {config.subject} 복습 기록표",
        )[learning_variant],
        "loc_topic": f"{learning_anchor}에서는",
        "loc": f"{learning_anchor}에서",
        "gen": f"{learning_anchor}의",
        "topic": f"{learning_anchor}{eun_neun(learning_anchor)}",
        "subject": f"{learning_anchor}{i_ga(learning_anchor)}",
        "object": f"{learning_anchor}{eul_reul(learning_anchor)}",
        "space": f"{learning_anchor} ",
    }
    anchor_learning_forms["next_progress"] = (
        f"{local} {config.grade_label} {config.subject}의 다음 학습 진도",
        f"{local} {config.grade_label} {config.subject}의 이후 학습 진도",
        f"{local} {config.grade_label} {config.subject}의 다음 수업 진도",
        f"{local} {config.grade_label} {config.subject}의 다음 공부 범위",
        f"{local} {config.grade_label} {config.subject}의 다음 복습 범위",
    )[learning_variant]
    anchor_learning_forms["home_review_topic"] = (
        f"{local} {config.grade_label} 학생의 {config.subject} 가정 복습은",
        f"{local}에서 {config.subject}{eul_reul(config.subject)} 공부하는 {config.grade_label} 학생이 집에서 이어 갈 복습은",
        f"{local} {config.grade_label} {config.subject} 수업 뒤 가정 복습은",
        f"{local} {config.grade_label} 학생의 {config.subject} 집 공부는",
        f"{local} {config.grade_label} 학생이 가정에서 하는 {config.subject} 복습은",
    )[learning_variant]

    def soften(
        sentence: str,
        consultation_targets: dict[str, str] = consultation_forms,
        learning_targets: dict[str, str] = learning_forms,
    ) -> str:
        replacements = (
            (f"{consultation} 전에는", consultation_targets["before_long"]),
            (f"{consultation} 전에", consultation_targets["before"]),
            (f"{consultation} 뒤에는", consultation_targets["after"]),
            (f"{consultation}의 최근 기록", "최근 학습 기록"),
            (f"{consultation}의 우선순위", consultation_targets["priority"]),
            (f"{consultation}의 진단 자료", consultation_targets["diagnostic_material"]),
            (f"{consultation}의 조건부 예시", consultation_targets["conditional_example"]),
            (f"{consultation}의 실제 분량", consultation_targets["actual_amount"]),
            (f"{consultation}의 주차별 분량", consultation_targets["schedule_amount"]),
            (f"{consultation}의 넷째 주", consultation_targets["week_four"]),
            (f"{consultation}의 수업 가능 여부", consultation_targets["availability"]),
            (f"{consultation}의 실제 반 편성", consultation_targets["class_assignment"]),
            (f"{consultation}의 피드백 주기", consultation_targets["feedback_cycle"]),
            (f"{consultation}의 계획이", consultation_targets["proposed_plan"]),
            (f"{consultation}의 적합성", consultation_targets["suitability"]),
            (f"{consultation}의 적합 여부", consultation_targets["suitability_decision"]),
            (f"{consultation} 자료", consultation_targets["material"]),
            (f"{consultation}의 수업 시간", consultation_targets["class_time"]),
            (
                f"{consultation}{eul_reul(consultation)} 준비하는 가정",
                consultation_targets["preparing_family"],
            ),
            (f"{consultation}에서는", consultation_targets["loc_topic"]),
            (f"{consultation}에서", consultation_targets["loc"]),
            (f"{consultation}의", consultation_targets["gen"]),
            (f"{consultation}{eun_neun(consultation)}", consultation_targets["topic"]),
            (f"{consultation}{i_ga(consultation)}", consultation_targets["subject"]),
            (f"{consultation}{eul_reul(consultation)}", consultation_targets["object"]),
            (f"{consultation} ", consultation_targets["space"]),
            (f"{learning}의 가정 복습은", learning_targets["home_review_topic"]),
            (f"{learning}의 피드백", learning_targets["feedback"]),
            (f"{learning}의 근거표", learning_targets["evidence_table"]),
            (f"{learning}의 새 진도", learning_targets["next_progress"]),
            (f"{learning} 기록에", learning_targets["record"]),
            (f"{learning} 자료에서", learning_targets["material"]),
            (f"{learning}에서는", learning_targets["loc_topic"]),
            (f"{learning}에서", learning_targets["loc"]),
            (f"{learning}의", learning_targets["gen"]),
            (f"{learning}{eun_neun(learning)}", learning_targets["topic"]),
            (f"{learning}{i_ga(learning)}", learning_targets["subject"]),
            (f"{learning}{eul_reul(learning)}", learning_targets["object"]),
            (f"{learning} ", learning_targets["space"]),
        )
        for source, target in replacements:
            if sentence.startswith(source):
                return f"{target}{sentence[len(source):]}"
        return sentence

    localized: list[str] = []
    for sentence_index, sentence in enumerate(sentences):
        sentence = sentence.strip()
        if not sentence:
            continue
        if sentence_index > 0:
            sentence = soften(sentence)
        else:
            sentence = soften(
                sentence,
                anchor_consultation_forms,
                anchor_learning_forms,
            )
        localized.append(sentence)
    return " ".join(localized)


def build_manuscript(
    row: dict[str, str],
    config: CategoryConfig,
    selection: SourceSelection,
    index: int,
) -> dict[str, str]:
    del index  # Row order is validated separately; copy variation uses stable labels.
    local = row["근처 수업가능 동네"].strip()
    region, district = generator.service_area_parts(row)
    center = row.get("센터명", "").strip() or f"{local} 학습코칭센터"
    schools = school_names_for(row, config)
    if schools:
        named_schools = "·".join(schools[:2])
        if config.is_elementary:
            school_materials = f"{named_schools} 학습 자료"
            school_range = f"{named_schools}의 최근 단원 진도"
            school_heading = f"{named_schools} 학습 자료"
        else:
            school_materials = f"{named_schools} 관련 자료"
            school_range = f"{named_schools}의 최신 시험 범위"
            school_heading = f"{named_schools} 관련 자료"
    else:
        named_schools = ""
        if config.is_elementary:
            school_materials = "학생이 준비한 학교 학습 자료"
            school_range = "최근 단원 진도"
            school_heading = "학생의 실제 학교 학습 자료"
        else:
            school_materials = "학생이 준비한 학교 자료"
            school_range = "최신 시험 범위"
            school_heading = "학생의 실제 학교 자료"
    if config.is_elementary:
        recent_records = "최근 단원평가 자료와 학습지"
        diagnosis_records = "최근 단원평가 자료, 사용 교재, 학교 진도표"
        first_week_records = "단원평가 자료와 교재"
        schedule_text = "학교 학습 일정"
        range_text = "단원평가 범위"
        current_result = "현재 정답 수"
        school_reference_items = "최근 단원평가 자료, 학교 학습지, 사용 교재"
    else:
        recent_records = "최근 시험지와 교재"
        diagnosis_records = "최근 시험지, 사용 교재, 학교 범위표"
        first_week_records = "시험지와 교재"
        schedule_text = "학생의 시험 일정"
        range_text = "학교 범위"
        current_result = "현재 점수"
        school_reference_items = f"{school_range}, 교과서 진도, 학교 프린트"
    fee_reference = (
        "공개된 교습비 자료" if row.get("센터 교습비", "").strip() else "교습비 확인 방법"
    )
    title = f"{local} {config.subject_label}학원"
    area_label = " ".join(part for part in (region, district) if part)
    consultation_label = f"{local} {config.grade_label} {config.subject} 상담"
    learning_label = f"{local} {config.grade_label} {config.subject} 학습"
    comparison_label = f"{local} {config.grade_label} {config.subject} 수업 비교"
    primary, secondary, support = (
        selection.primary,
        selection.secondary,
        selection.support,
    )

    title_object = f"{title}{eul_reul(title)}"
    local_topic = f"{local}{eun_neun(local)}"
    local_subject = f"{local}{i_ga(local)}"
    grade_topic = f"{config.grade_label}{eun_neun(config.grade_label)}"
    subject_topic = f"{config.subject}{eun_neun(config.subject)}"
    subject_object = f"{config.subject}{eul_reul(config.subject)}"
    primary_subject = f"{primary.label}{i_ga(primary.label)}"
    primary_object = f"{primary.label}{eul_reul(primary.label)}"
    secondary_object = f"{secondary.label}{eul_reul(secondary.label)}"
    support_subject = f"{support.label}{i_ga(support.label)}"
    support_object = f"{support.label}{eul_reul(support.label)}"
    primary_and_secondary = f"{primary.label}·{secondary.label}"
    secondary_and_support = f"{secondary.label}·{support.label}"
    primary_or_secondary = f"{primary.label}{ina_na(primary.label)} {secondary.label}"
    checklist_note_open = choose(
        (
            f"상담 메모에는 {primary.label} 관련 질문과 {secondary.label} 관련 질문을 한 개씩 적어 두세요.",
            f"최근 교재를 펼쳐 {primary_object} 확인할 쪽과 {support_object} 다시 볼 쪽에 서로 다른 표시를 해 두세요.",
            f"학생에게 {primary.label} 관련 어려움을 한 문장으로 말하게 하고, {secondary.label}에서 확인할 내용을 이어 적어 보세요.",
            f"{school_materials}{eul_reul(school_materials)} 정리할 때는 {primary.label} 문항과 {support.label} 질문을 구분해 두세요.",
            f"질문 목록은 {primary.label}, {secondary.label}, {support.label} 순서로 나누고 가장 급한 항목에 표시해 두세요.",
            f"첫 주 계획을 비교하려면 ‘{primary.practice}’ 활동의 완료 기준과 ‘{secondary.practice}’ 활동의 확인 날짜를 따로 적어야 합니다.",
            f"상담 답변을 기록할 때는 {primary_object} 확인한 방법과 {support_object} 다시 살펴볼 방법을 서로 다른 칸에 적어 보세요.",
            f"통학 가능한 시간과 함께 {primary_object} 확인할 날짜, {secondary_object} 다시 볼 날짜도 미리 정리해 두세요.",
        ),
        config.category,
        local,
        "checklist-note-open",
    )
    checklist_note_close = choose(
        (
            f"{local}의 {config.grade_label} {config.subject} 상담 답변 옆에는 학생이 직접 할 행동과 확인 날짜를 함께 남기면 좋습니다.",
            f"{local}에서 {subject_object} 공부하는 {config.grade_label} 학생이 설명을 다시 말할 수 있는지 확인하면 이해한 범위를 구분하기 쉽습니다.",
            f"{local} {config.grade_label} 학생에게 제안된 {config.subject} 분량은 학교 일정과 집에서 가능한 시간에 맞는지 다시 대조하세요.",
            f"{local}의 {config.grade_label} {config.subject} 상담은 당일에 결정하지 않아도 되므로 답변을 적어 두고 다른 선택지와 차분히 비교하세요.",
            f"{center}에서 들은 운영 안내와 공개 자료가 일치하는지도 확인하세요.",
            f"{local} {config.grade_label} {config.subject} 학습에서 한 주 뒤 확인할 기준은 정답 수보다 설명과 재풀이가 유지되는지에 두는 편이 좋습니다.",
            f"{local}에서 정리한 질문이 많다면 이번 주에 확인할 것과 등록 전에 확인할 것을 나누어 물어보세요.",
            f"{local}에서 정리한 이 기록은 {subject_object} 공부하는 {config.grade_label} 학생에게 필요한 도움과 스스로 해 볼 범위를 구분하는 비교 자료가 됩니다.",
        ),
        config.category,
        local,
        "checklist-note-close",
    )
    evidence_note_open = choose(
        (
            f"근거를 정리할 때는 {primary.label}{gwa_wa(primary.label)} {secondary_object} 한 칸에 섞지 말고 각각 다른 문제로 확인하세요.",
            f"정답이 같은 두 문제라도 {primary_object} 설명한 내용과 {support_subject} 드러난 기록을 각각 살펴 원인을 따로 표시하세요.",
            f"학생이 스스로 설명한 문장은 {primary_object} 확인하는 자료로, 다시 풀어 본 흔적은 {secondary_object} 확인하는 자료로 나눠 보세요.",
            f"학교 자료에는 {primary_subject} 드러난 문항을, 현재 교재에는 {support_object} 다시 확인할 쪽을 표시해 두면 좋습니다.",
            f"오답 옆에는 {primary.label}에서 막힌 이유와 {secondary_object} 다시 확인할 날짜를 서로 다른 색으로 적어 보세요.",
            f"현재 교재에서는 {primary.label} 상태를 확인할 쪽과 {support.label}의 다음 행동을 기록할 쪽을 따로 골라 두는 편이 좋습니다.",
            f"한 번에 여러 원인을 붙이기보다 {primary.label}, {secondary.label}, {support.label} 가운데 가장 먼저 드러난 하나부터 기록하세요.",
            f"진단 표는 {primary.label}의 현재 상태, {secondary.label}의 재확인 여부, {support.label}의 다음 행동으로 나누어 작성할 수 있습니다.",
        ),
        config.category,
        local,
        "evidence-note-open",
    )
    evidence_note_close = choose(
        (
            "이렇게 나눈 기록은 다음 설명과 연습의 순서를 정하는 데 도움이 됩니다.",
            "자료마다 확인 날짜를 적어 두면 한 번의 실수와 반복되는 막힘을 구분하기 쉽습니다.",
            "학생이 말로 설명하지 못한 지점만 표시해도 상담 질문의 범위를 줄일 수 있습니다.",
            "기록을 세분하면 문제 수보다 어떤 도움이 필요한지를 먼저 이야기할 수 있습니다.",
            "고친 답이 실제 이해로 이어졌는지는 이후 재풀이 결과로 살펴보세요.",
            "정리한 표를 보여 주면 새 진도와 복습 가운데 어느 쪽을 먼저 볼지 질문하기 쉽습니다.",
            "완료한 내용과 아직 설명하지 못한 내용을 분리하면 첫 주 계획을 과도하게 잡는 일을 피할 수 있습니다.",
            "근거가 없는 항목은 추측하지 말고 학생이 실제로 남긴 자료부터 확인하는 편이 안전합니다.",
        ),
        config.category,
        local,
        "evidence-note-close",
    )

    meta = fit_meta(
        (
            f"{title} 선택 전 {primary_and_secondary}{eul_reul(secondary.label)} 점검하는 방법, {school_materials}{eul_reul(school_materials)} 활용한 조건부 4주 계획, 상담 체크리스트와 센터 위치를 확인하세요.",
            f"{area_label} {title} 안내입니다. {primary.label} 진단 근거와 {secondary_object} 보완하는 방법, {school_range} 확인, 추천 학생과 상담 전 준비 항목을 정리했습니다.",
            f"{title_object} 비교하는 학부모를 위해 {primary_and_secondary} 진단, {config.grade_label} 학교 자료 활용법, 상담에서 조정할 4주 계획과 확인 질문을 센터 정보와 함께 안내합니다.",
            f"{local}에서 {config.grade_label} {config.subject}학원을 알아볼 때 필요한 {primary.label} 점검법과 {school_materials} 준비, 첫 4주 학습 순서와 상담 질문을 센터 정보와 함께 정리했습니다.",
            f"{primary_subject} 자주 막히는 {config.grade_label} 학생이라면 무엇부터 확인해야 할까요? {title} 상담에 앞서 살펴볼 학습 기록, 학교 자료, 복습 계획과 방문 정보를 안내합니다.",
            f"{school_materials}{eul_reul(school_materials)} 바탕으로 {secondary_object} 점검하고 첫 학습 순서를 정하는 방법을 소개합니다. {title} 추천 대상과 상담 전 체크리스트도 함께 확인하세요.",
            f"{center} 방문 전 확인할 {config.grade_label} {config.subject} 학습 자료와 상담 질문을 모았습니다. {primary.label} 진단부터 가정 복습까지 {local} 학부모가 살펴볼 기준을 안내합니다.",
            f"조건부 4주 계획은 현재 기록을 정확히 읽는 데서 시작합니다. {title}에서 확인할 {primary_and_secondary} 학습 기준, 학교 자료 활용법, 상담 준비와 센터 위치를 정리했습니다.",
        ),
        title,
    )

    intro = [
        choose(
            (
                f"{title_object} 찾는다면 문제 수나 진도보다 최근 학습 기록에서 {primary.label} 상태를 먼저 확인하는 것이 핵심입니다. {consultation_label}에서는 학생의 자료로 {primary.check}를 확인하고 필요한 보완 범위를 정해야 합니다.",
                f"{title} 선택의 출발점은 {current_result} 자체보다 {primary.label}에서 반복되는 막힘을 찾는 일입니다. {consultation_label}에서는 학생이 {primary.check}를 {recent_records}에 표시하면 먼저 보완할 범위를 좁힐 수 있습니다.",
                f"{title}에 관한 짧은 답은 {primary.label} 진단부터 시작하라는 것입니다. {consultation_label}에서 {primary.check} 살펴보면 새 진도와 복습의 순서를 나누기 쉽습니다.",
            ),
            config.category,
            local,
            "intro-answer",
        ),
        choose(
            (
                f"{consultation_label} 전에는 {school_materials}{gwa_wa(school_materials)} {center} 위치를 먼저 확인하세요. {consultation_label}의 실제 반 편성, 수업 시간, 교재와 진도는 학생의 출발점에 따라 달라질 수 있으므로 최신 상담에서 확인해야 합니다.",
                f"{consultation_label}에서는 {school_materials}와 {center} 위치를 함께 확인해야 합니다. {consultation_label}의 수업 가능 여부와 일정은 상담 시점의 안내로 다시 확인하는 것이 안전합니다.",
                f"{area_label} {local}에서 {config.grade_label} {config.subject} 수업을 알아볼 때는 {school_range}{gwa_wa(school_range)} 학생이 틀린 문제를 함께 보아야 합니다. {consultation_label}에서는 {center}의 운영 일정과 {config.grade_label} 반 편성 여부를 방문 전에 확인해 계획과 실제 수업이 맞는지 점검해 주세요.",
            ),
            config.category,
            local,
            "intro-context",
        ),
    ]

    section_one = optional_third(
        choose(
            (
                f"{consultation_label}에서 가장 먼저 확인할 내용은 {primary.label}입니다. {consultation_label}에서는 최근 자료로 {primary.check}를 확인하고, 설명이 멈추는 지점을 첫 보완 대상으로 잡는 흐름이 적절합니다.",
                f"{learning_label}의 시작점은 진도를 앞당기는 일이 아니라 {primary.label}에서 드러난 빈칸을 찾는 일입니다. {consultation_label}에서는 학생이 {primary.check}를 직접 보여 주면 설명이 필요한 부분과 적절한 연습량을 정하기 쉽습니다.",
                f"{consultation_label}에서는 문제 수보다 ‘{primary_object} 어떻게 이해했는가’를 설명할 수 있어야 합니다. {consultation_label}의 최근 기록에서 {primary.check}를 먼저 살펴보는 데서 진단이 시작됩니다.",
            ),
            config.category,
            local,
            "direct-1",
        ),
        f"{learning_label}에서는 앞 단원과 새 단원을 연결하는 누적 이해가 중요하므로 {secondary.label}도 함께 봐야 합니다. {learning_label} 자료에서 {secondary.evidence}를 확인하면 단순 실수와 보완할 내용을 구분할 수 있습니다.",
        f"{consultation_label}의 우선순위를 한 번에 많이 정할 필요는 없습니다. {consultation_label}에서 {support.label}까지 살펴본 뒤 이번 주 확인 항목 하나와 다음 점검 날짜 하나를 학습 기록에 남기는 방식이 현실적입니다.",
        config=config,
        local=local,
        section_code="direct-answer",
    )

    section_two = optional_third(
        f"{consultation_label}의 진단 자료로는 {diagnosis_records}{i_ga(diagnosis_records)} 유용합니다. {consultation_label}에서는 {school_materials}{eul_reul(school_materials)} 살펴보며 {primary.evidence} 확인하면 {primary.label}에서 막힌 원인을 구체적으로 설명할 수 있습니다.",
        choose(
            (
                f"{learning_label}에서 틀린 문제는 {primary.label}, {secondary.label}, {support.label} 가운데 시작 원인을 하나씩 붙여 분류해 보세요. {consultation_label}에서 이 구분을 공유하면 보완 순서를 구체적으로 질문할 수 있습니다.",
                f"{learning_label}의 근거표에는 정답 여부만 쓰지 말고 {secondary.evidence}{eul_reul(secondary.evidence)} 함께 확인하는 편이 좋습니다. {learning_label}에서 같은 문제를 다시 해결했는지까지 확인해야 일시적인 수정과 재학습을 구분할 수 있습니다.",
                f"{consultation_label}에서 {school_materials}{eul_reul(school_materials)} 살피며 {support.evidence}{eul_reul(support.evidence)} 확인해 두세요. {consultation_label}에서는 확인한 내용을 바탕으로 설명이 필요한 부분과 혼자 연습할 부분을 나눌 수 있습니다.",
            ),
            config.category,
            local,
            "evidence-2",
        ) + f" {evidence_note_open} {evidence_note_close}",
        f"{local}에서 {config.grade_label} {config.subject} 수업을 비교할 때 {district}의 학교마다 진도와 평가 방식이 같다고 단정할 수는 없습니다. 따라서 {consultation_label} 계획은 {school_range}, 교과서 단원, 학생이 받은 안내문을 상담 시점에 다시 대조해야 합니다.",
        config=config,
        local=local,
        section_code="evidence-diagnosis",
    )

    section_three = optional_third(
        f"{consultation_label}{i_ga(consultation_label)} 특히 필요한 경우는 {primary.label}에서 같은 이유로 막히거나, {secondary_object} 보완하는 계획이 실행 기록으로 이어지지 않는 학생입니다. {consultation_label}에서 질문할 내용을 정리하지 못하는 학생은 진단 자료부터 준비해야 합니다.",
        choose(
            (
                f"반대로 {learning_label}에 필요한 진도와 학습량이 이미 분명하다면 시간표, 피드백 주기, 다시 확인하는 방식이 맞는지를 중심으로 비교할 수 있습니다. {consultation_label}에서는 학생의 현재 수준에 맞는 운영 여부를 우선 확인해야 합니다.",
                f"{local}에서 {config.grade_label} {config.subject} 수업을 비교하는 학부모는 학생이 혼자 해결할 수 있는 범위와 설명이 필요한 범위를 구분해 전달해 보세요. {consultation_label}에서 제안된 계획을 통학 동선과 주간 가능 시간까지 함께 보면 실행할 수 있는지 판단하기 쉽습니다.",
                f"{learning_label}에서는 분량만 늘리기보다 학교 진도와 이전 공백의 비중을 조절해야 합니다. {consultation_label}에서 학생의 목표, 남은 기간, 집에서 가능한 복습 시간을 함께 알려 주세요.",
            ),
            config.category,
            local,
            "recommended-2",
        ),
        f"{consultation_label}의 적합 여부는 학원 이름이나 한 번의 점수로 결정할 수 없습니다. {consultation_label}에서는 실제 자료에서 {support.check}를 살펴보고, 상담 뒤 제안된 계획을 지속할 수 있는지를 함께 판단해야 합니다.",
        config=config,
        local=local,
        section_code="recommended-student",
    )

    section_four = optional_third(
        choose(
            (
                f"{consultation_label}에서 4주 계획을 논의한다면 첫째 주에는 {first_week_records}{eul_reul(first_week_records)} 모아 {primary_and_secondary}에서 막힌 원인을 나눕니다. {learning_label}의 둘째 주에는 ‘{primary.practice}’를 시도한 뒤 학생이 해설 없이 설명할 수 있는지 확인합니다.",
                f"{learning_label}의 첫 주는 {primary_and_secondary} 상태를 실제 자료로 확인하는 기간으로 잡을 수 있습니다. {consultation_label}에서는 확인 결과를 바탕으로 둘째 주에 ‘{primary.practice}’를 연습하고, 풀이 이유를 학생의 말로 설명하게 합니다.",
                f"{consultation_label}에서 첫 두 주의 목표를 정할 때는 첫째 주에 {primary_and_secondary}에서 막힌 지점을 찾고, 둘째 주에 ‘{primary.practice}’를 적용해 설명과 풀이가 이어지는지 살펴보는 흐름을 검토할 수 있습니다.",
            ),
            config.category,
            local,
            "four-week-1",
        ),
        choose(
            (
                f"{consultation_label}에서 학교 일정과 첫 진단 결과를 확인한 뒤 셋째 주에는 ‘{secondary.practice}’와 ‘{support.practice}’를 번갈아 시도할 수 있습니다. {learning_label}의 넷째 주에는 ‘{primary.home_action}’를 실행한 결과를 보고 다음 범위를 조정합니다.",
                f"{learning_label}의 셋째 주에는 학생이 감당할 수 있는 분량 안에서 ‘{secondary.practice}’와 ‘{support.practice}’를 함께 점검합니다. {consultation_label}의 넷째 주에는 ‘{primary.home_action}’의 실행 여부를 살펴 유지할 내용과 바꿀 내용을 나눕니다.",
                f"{consultation_label}에서 후반 두 주의 순서를 학교 일정에 맞춰 조정해야 합니다. {learning_label}의 셋째 주에는 ‘{secondary.practice}’와 ‘{support.practice}’를 시도하고, 넷째 주에는 학생의 ‘{primary.home_action}’ 실행 여부를 살펴 다음 계획을 정할 수 있습니다.",
            ),
            config.category,
            local,
            "four-week-2",
        ),
        choose(
            (
                f"이 4주 흐름은 {consultation_label}에서 비교할 계획 예시일 뿐, 고정 수업 약속이 아닙니다. {consultation_label}의 주차별 분량과 순서는 {schedule_text}{gwa_wa(schedule_text)} 시작 수준을 확인한 뒤 달리 정해야 합니다.",
                f"{learning_label}의 4주 예시를 채택한다면 완료 기준도 함께 적어야 합니다. {learning_label}에서는 정답 수뿐 아니라 설명 가능 여부, 다시 푼 날짜, 같은 오류의 반복 여부를 남겨야 다음 조정 근거가 생깁니다.",
                f"{local}에서 다음 평가나 점검까지 4주가 남지 않은 {config.grade_label} 학생에게 {config.subject} 전 단원을 같은 비중으로 다루기 어렵습니다. {consultation_label}에서 {range_text}{gwa_wa(range_text)} 남은 날을 확인하고 {primary.label} 관련 우선 단원부터 좁혀 보세요.",
            ),
            config.category,
            local,
            "four-week-3",
        ),
        config=config,
        local=local,
        section_code="four-week-plan",
    )

    section_five = optional_third(
        f"{learning_label}의 피드백은 ‘잘했다’나 ‘더 풀자’보다 다음 행동이 보여야 합니다. {local}에서 {config.grade_label} {subject_object} 공부하는 학생에게는 {primary.evidence}를 짚고, 집에서는 ‘{primary.home_action}’를 다음 행동으로 정하는 방식이 구체적입니다.",
        choose(
            (
                f"학부모는 {learning_label} 기록에서 새로 배운 내용, 다시 볼 내용, 질문할 내용을 구분해 확인할 수 있습니다. {learning_label}을 돕는 가정에서는 ‘{secondary.home_action}’처럼 짧고 확인 가능한 행동을 정하는 편이 부담을 줄입니다.",
                f"{consultation_label} 뒤에는 학생이 받은 피드백을 한 문장으로 다시 말해 보게 해 주세요. {learning_label}에서는 학생이 ‘{support.home_action}’까지 스스로 정했는지 확인하면 수업 내용이 집 공부로 이어졌는지 살펴보기 쉽습니다.",
                f"{consultation_label}{eul_reul(consultation_label)} 준비하는 가정에서는 공부 시간을 감시하기보다 정한 완료 기준을 함께 확인하는 편이 좋습니다. {consultation_label}에서 ‘{secondary.home_action}’의 결과를 다음 상담 자료로 남기면 계획 조정에 사용할 수 있습니다.",
            ),
            config.category,
            local,
            "feedback-2",
        ),
        f"{consultation_label}의 피드백 주기와 전달 방식은 센터 운영에 따라 달라질 수 있습니다. {consultation_label}에서는 학생이 질문을 받는 시점, 틀린 문제를 다시 확인할 날짜, 학부모에게 공유되는 항목을 구체적으로 물어보세요.",
        config=config,
        local=local,
        section_code="feedback-home",
    )

    section_six = optional_third(
        f"{consultation_label} 전에는 {diagnosis_records}, 일주일 학습 가능 시간을 준비해 주세요. {consultation_label}에서 {primary.label} 관련 문제 두세 개와 스스로 잘 풀린 문제 한 개를 함께 고르면 출발점을 균형 있게 설명할 수 있습니다.",
        f"{consultation_label}에서는 실제 수업 가능 학년과 시간, 사용할 교재, 과제량, 결석 시 보완, 피드백 주기, 그리고 {fee_reference}까지 차례로 확인하세요. {center}까지의 통학 동선과 귀가 시간도 {learning_label}{eul_reul(learning_label)} 지속할 수 있는지 판단하는 항목입니다. {checklist_note_open} {checklist_note_close}",
        choose(
            (
                f"{consultation_label}에서 마지막으로 {school_range}{i_ga(school_range)} 반영되는지, {secondary_and_support}{eul_reul(support.label)} 어떤 자료로 다시 확인하는지 질문해 보세요. {consultation_label} 내용을 날짜와 함께 적어 두면 다른 선택지와 비교하기 쉽습니다.",
                f"{consultation_label}{i_ga(consultation_label)} 끝나면 제안된 첫 주 분량, 학생이 할 일, 다음 확인일을 한 줄씩 정리해 보세요. {consultation_label}의 계획이 실제 일정과 맞지 않으면 등록 전에 조정 가능 여부를 물어보는 것이 좋습니다.",
                f"{local}에서 {config.grade_label} {config.subject} 수업을 비교하는 학부모는 답변이 학생의 자료에 근거했는지 확인해 주세요. {consultation_label}에서 목표를 점수 약속으로 정하기보다 {primary_and_secondary}별 실행 기준으로 바꾸면 이후 점검이 명확해집니다.",
            ),
            config.category,
            local,
            "checklist-3",
        ),
        config=config,
        local=local,
        section_code="consult-checklist",
    )

    intro = diversify_paragraphs(
        intro,
        config=config,
        local=local,
        purpose="intro",
    )
    section_one = diversify_paragraphs(
        section_one,
        config=config,
        local=local,
        purpose="direct-answer",
    )
    section_two = diversify_paragraphs(
        section_two,
        config=config,
        local=local,
        purpose="evidence-diagnosis",
    )
    section_three = diversify_paragraphs(
        section_three,
        config=config,
        local=local,
        purpose="recommended-student",
    )
    section_four = diversify_paragraphs(
        section_four,
        config=config,
        local=local,
        purpose="conditional-four-week",
    )
    section_five = diversify_paragraphs(
        section_five,
        config=config,
        local=local,
        purpose="feedback-home",
    )
    section_six = diversify_paragraphs(
        section_six,
        config=config,
        local=local,
        purpose="consult-checklist",
    )

    sections = [
        (f"{consultation_label}, 무엇부터 확인할까요?", section_one),
        (
            f"{local}의 {school_heading}에서 찾는 {config.grade_label} {config.subject} 학습 근거"
            if schools
            else f"{local} 학생의 실제 학교 자료에서 찾는 {config.grade_label} {config.subject} 학습 근거",
            section_two,
        ),
        (f"{local} {config.grade_label} {config.subject}, 어떤 학생에게 맞을까요?", section_three),
        (f"상담에서 조정하는 {local} {config.grade_label} {config.subject} 4주 계획 예시", section_four),
        (f"{learning_label} 피드백을 집 공부로 잇는 법", section_five),
        (f"{consultation_label} 전 확인할 체크리스트", section_six),
    ]

    body_parts = [*intro]
    for heading, paragraphs in sections:
        body_parts.append(f"## {heading}")
        body_parts.extend(paragraphs)
    body = "\n\n".join(body_parts)

    faq_bank = [
        (
            f"{consultation_label}에서 {primary.label}{eun_neun(primary.label)} 어떻게 확인하나요?",
            f"{consultation_label} 전에는 {recent_records}에서 {primary.evidence}를 먼저 봅니다. {consultation_label}에서는 이 자료로 설명이 필요한 범위와 혼자 다시 풀 범위를 구분해 질문하면 됩니다.",
        ),
        (
            f"{consultation_label}에 {school_materials}{eul_reul(school_materials)} 가져가야 하나요?",
            f"{consultation_label}에서 필수라고 단정할 수는 없지만 {school_reference_items}{i_ga(school_reference_items)} 있으면 제안받은 계획을 더 구체적으로 비교할 수 있습니다. {consultation_label} 자료를 준비할 때 개인정보가 적힌 부분은 가리고 필요한 {config.subject} 학습 내용만 가져가도 됩니다.",
        ),
        (
            f"{learning_label}의 첫 4주는 어떻게 계획하나요?",
            f"{consultation_label}의 조건부 예시로 첫 주 진단, 둘째 주 ‘{primary.practice}’, 셋째 주 ‘{secondary.practice}’, 넷째 주 재확인 순서를 논의할 수 있습니다. {consultation_label}의 실제 분량은 {schedule_text}{gwa_wa(schedule_text)} 시작 수준을 확인한 뒤 정해야 합니다.",
        ),
        (
            f"{consultation_label}{eun_neun(consultation_label)} 어떤 학생에게 필요한가요?",
            f"{consultation_label}에서는 {primary_or_secondary}에서 같은 막힘이 반복되거나 계획은 있지만 완료 기록과 질문이 남지 않는 학생이 점검해 볼 수 있습니다. {consultation_label}의 적합성은 진단 자료와 통학 가능 시간까지 함께 보고 판단해야 합니다.",
        ),
        (
            f"{local}에서 {config.grade_label} {config.subject} 수업 등록 전 무엇을 확인해야 하나요?",
            f"{local} {config.grade_label} {config.subject} 수업 등록 전에는 {center}의 반 편성, 시작일, 교재, 과제, 보완 방식과 피드백 주기를 확인하세요. {consultation_label}에서는 학생 자료를 기준으로 시작 단원과 첫 주 분량을 어떻게 정하는지도 물어보는 것이 좋습니다.",
        ),
        (
            f"{learning_label}에서 {secondary.label}{eul_reul(secondary.label)} 보완하려면 무엇부터 해야 하나요?",
            f"{consultation_label}에서는 {secondary.evidence}를 확인해 현재 막힘의 원인을 먼저 나눕니다. 첫 실천 항목은 ‘{secondary.practice}’처럼 완료 여부를 확인할 수 있는 크기로 정하는 편이 좋습니다.",
        ),
        (
            f"{local} 학부모는 {config.grade_label} {config.subject} 피드백에서 무엇을 확인해야 하나요?",
            f"{consultation_label}에서는 맞고 틀린 결과보다 학생이 받은 설명과 다음 행동이 분명한지 확인하세요. {learning_label} 기록에 ‘{support.home_action}’의 실행 여부와 다음 확인일을 함께 남기면 가정에서도 흐름을 살펴볼 수 있습니다.",
        ),
        (
            f"{local} {config.grade_label} {config.subject} 교재와 진도는 어떻게 비교하나요?",
            f"{consultation_label}에서 현재 교재의 완료 단원, 학교 진도표, 선행과 복습의 비중을 함께 보여 주세요. {learning_label}에서 다음에 다룰 범위는 학교 일정과 이전 단원의 공백, 한 주에 가능한 분량을 확인한 뒤 정해야 합니다.",
        ),
        (
            f"{local} {config.grade_label} {config.subject} 수업 시간과 결석 시 보완 방식은 어떻게 확인하나요?",
            f"{consultation_label}의 수업 시간, 결석 시 보완, 과제 전달, 피드백 방식은 센터 운영과 반 편성에 따라 달라질 수 있습니다. {consultation_label} 전에 필요한 조건을 목록으로 적어 최신 안내와 하나씩 대조하세요.",
        ),
        (
            f"{local}에서 {config.grade_label} {config.subject} 통학과 비용은 어떻게 확인하나요?",
            f"{consultation_label} 전에는 {center}까지의 실제 이동 시간과 귀가 동선을 먼저 확인하세요. {consultation_label}에서는 {fee_reference}, 교재비 포함 여부, 수업 횟수와 변경 가능성을 최신 안내로 다시 확인해야 합니다.",
        ),
        (
            f"{consultation_label} 전에 질문 목록은 어떻게 만들까요?",
            f"{consultation_label}에 가져갈 질문은 {primary.label}, {secondary.label}, {support.label} 가운데 현재 가장 급한 순서로 적어 보세요. {learning_label} 자료에서 각 질문과 연결되는 실제 문제나 기록을 하나씩 골라 두면 답변을 비교하기 쉽습니다.",
        ),
        (
            f"{local} {config.grade_label} {config.subject} 복습은 집에서 어떻게 이어가나요?",
            f"{learning_label}의 가정 복습은 긴 분량보다 ‘{primary.home_action}’처럼 끝을 확인할 수 있는 행동으로 정합니다. {consultation_label}에서는 실행한 날짜와 어려웠던 지점을 다음 피드백에 어떻게 반영할지 물어보세요.",
        ),
    ]
    faq_offset = stable_number(config.category, local, "faq-bank") % len(faq_bank)
    faq_pairs = [
        faq_bank[(faq_offset + faq_index * 5) % len(faq_bank)]
        for faq_index in range(5)
    ]
    faq_pairs = [
        (
            question,
            localize_sentences(
                answer,
                config=config,
                local=local,
                purpose="faq",
                ordinal=faq_index,
            ),
        )
        for faq_index, (question, answer) in enumerate(faq_pairs)
    ]
    faq = "\n".join(
        f"Q{faq_index}. {question}\nA{faq_index}. {answer}"
        for faq_index, (question, answer) in enumerate(faq_pairs, 1)
    )

    scenarios = [
        f"가정 상담 장면: {local}의 한 학부모가 최근 자료 가운데 {primary.label}에서 막힌 문제와 {secondary.label} 관련 학습 기록을 나눠 가져오고, {consultation_label}에서 제안받을 첫 주 확인 항목과 다음 점검일을 질문하는 상황을 가정할 수 있습니다. 이 {consultation_label} 장면은 실제 이용 후기나 특정 학생의 성과가 아니며, 상담 준비용 가상 예시입니다. {local}에서 {config.grade_label} {subject_object} 공부하는 학생의 학습 결과는 출발점과 실천 정도에 따라 달라질 수 있습니다.",
        f"비교 상담 장면: {local}에서 {config.grade_label} {subject_object} 공부하는 학생의 학교 일정과 통학 가능 시간을 적어 두고 {center}에서 {support.label} 관련 기록을 확인하는 방식, 과제 조절 기준, 학부모 공유 항목을 묻는 장면을 생각해 볼 수 있습니다. 이 {consultation_label} 장면에서 비교할 항목은 학생의 실제 자료와 가능한 일정에 맞춰 조정해야 합니다.",
    ]

    summary_base = (
        f"{title} 선택 전 {primary.label}, {secondary.label}, {support.label}{eul_reul(support.label)} 최근 학습 기록에서 확인합니다. "
        f"{local}에서 {config.grade_label} {config.subject} 수업을 비교할 때에는 {school_materials}{gwa_wa(school_materials)} 센터 정보를 바탕으로 조건부 4주 계획, 추천 학생, 상담 체크리스트를 함께 점검합니다."
    )
    summary = localize_sentences(
        summary_base,
        config=config,
        local=local,
        purpose="quick-answer",
        ordinal=0,
    )
    return {
        "페이지타이틀": title,
        "메타설명": meta,
        "본문": body,
        "FAQ": faq,
        "학부모후기": "\n\n".join(scenarios),
        "JSON-LD 요약": summary,
    }


def split_sentences(value: str) -> list[str]:
    clean = re.sub(r"^##\s+", "", value, flags=re.MULTILINE)
    return [
        re.sub(r"\s+", " ", sentence).strip()
        for sentence in re.split(r"(?<=[.!?])\s+|\n+", clean)
        if len(re.sub(r"\s+", " ", sentence).strip()) >= 16
    ]


def masked_manuscript_shingles(
    manuscript: dict[str, str],
    row: dict[str, str],
    config: CategoryConfig,
) -> set[tuple[str, ...]]:
    title = manuscript["페이지타이틀"]
    authored = "\n".join(
        (
            f"{title} 핵심 답변",
            f"{title}에서 먼저 확인할 기준은 무엇인가요?",
            manuscript["JSON-LD 요약"],
            manuscript["본문"],
            manuscript["FAQ"],
            manuscript["학부모후기"],
        )
    )
    local = row["근처 수업가능 동네"].strip()
    mask_values = [
        local,
        row.get("지역", "").strip(),
        row.get("시or구", "").strip(),
        row.get("센터명", "").strip(),
        row.get("센터 주소", "").strip(),
        config.category,
        config.subject_label,
        f"{config.school_level} {config.grade_number}학년",
        config.grade_label,
        config.subject,
        *school_names_for(row, config),
    ]
    for token in sorted({item for item in mask_values if item}, key=len, reverse=True):
        authored = re.sub(re.escape(token), " 지역정보 ", authored, flags=re.I)
    authored = re.sub(r"https?://\S+", " 주소정보 ", authored)
    authored = re.sub(r"\b\d+(?:[-.:/]\d+)*\b", " 숫자정보 ", authored)
    tokens = re.findall(r"[가-힣A-Za-z]+", re.sub(r"\s+", " ", authored).lower())
    return {
        tuple(tokens[position : position + 5])
        for position in range(max(0, len(tokens) - 4))
    }


def validate_masked_similarity(
    records: list[tuple[str, set[tuple[str, ...]]]],
    limit: float = 0.75,
) -> tuple[float, tuple[str, str] | None, int]:
    # A 96-bucket one-permutation MinHash LSH keeps this all-page preflight fast.
    # With 24 bands of four buckets, a pair at the 0.75 hard limit has >99.98%
    # candidate probability; an additional 0.58 signature gate is conservative.
    bucket_count = 96
    band_size = 4
    hash_cache: dict[tuple[str, ...], int] = {}
    signatures: list[tuple[int, ...]] = []
    for label, shingles in records:
        if len(shingles) < 40:
            raise ValueError(f"마스킹 5-shingle 본문 길이 부족: {label}={len(shingles)}")
        buckets = [(1 << 64) - 1] * bucket_count
        for shingle in shingles:
            digest = hash_cache.get(shingle)
            if digest is None:
                digest = int.from_bytes(
                    hashlib.blake2b("\x1f".join(shingle).encode("utf-8"), digest_size=8).digest(),
                    "big",
                )
                hash_cache[shingle] = digest
            bucket = digest % bucket_count
            reduced = digest // bucket_count
            if reduced < buckets[bucket]:
                buckets[bucket] = reduced
        signatures.append(tuple(buckets))

    band_buckets: dict[tuple[int, tuple[int, ...]], list[int]] = defaultdict(list)
    candidate_pairs: set[tuple[int, int]] = set()
    for record_index, signature in enumerate(signatures):
        for band_start in range(0, bucket_count, band_size):
            band_number = band_start // band_size
            key = (band_number, signature[band_start : band_start + band_size])
            for prior_index in band_buckets[key]:
                candidate_pairs.add((prior_index, record_index))
            band_buckets[key].append(record_index)

    # Always cover close neighbours in stable order as a deterministic sanity set.
    for record_index in range(len(records)):
        for prior_index in range(max(0, record_index - 8), record_index):
            candidate_pairs.add((prior_index, record_index))

    maximum = 0.0
    maximum_pair: tuple[str, str] | None = None
    candidates_checked = 0
    for first_index, second_index in candidate_pairs:
        first_signature = signatures[first_index]
        second_signature = signatures[second_index]
        estimated = sum(
            first == second
            for first, second in zip(first_signature, second_signature, strict=True)
        ) / bucket_count
        if estimated < 0.58 and second_index - first_index > 8:
            continue
        first_label, first = records[first_index]
        second_label, second = records[second_index]
        intersection = len(first & second)
        union = len(first) + len(second) - intersection
        score = intersection / union if union else 1.0
        candidates_checked += 1
        if score > maximum:
            maximum = score
            maximum_pair = (first_label, second_label)
        if score >= limit:
            raise ValueError(
                f"마스킹 5-shingle 유사도 초과: {score:.4f} "
                f"{first_label} / {second_label}"
            )
    return maximum, maximum_pair, candidates_checked


def assert_no_source_sentence_copy(manuscript: dict[str, str], raw_html: str, title: str) -> None:
    source = re.sub(r"\s+", " ", visible_source_text(raw_html)).strip()
    generated = "\n".join(manuscript.values())
    for sentence in split_sentences(generated):
        normalized = re.sub(r"\s+", " ", sentence).strip()
        if len(normalized) >= 42 and normalized in source:
            raise ValueError(f"재사용 원문과 일치하는 문장 발견: {title}: {normalized[:80]}")


def build_category_manuscripts(
    rows: list[dict[str, str]],
    config: CategoryConfig,
) -> tuple[dict[str, dict[str, str]], dict[str, tuple[str, str, str]]]:
    raw_cells = read_workbook_cells(config.workbook_path)
    manuscripts: dict[str, dict[str, str]] = {}
    signal_log: dict[str, tuple[str, str, str]] = {}
    for index, (row, raw_html) in enumerate(zip(rows, raw_cells, strict=True)):
        local = row["근처 수업가능 동네"].strip()
        selection = select_signals(raw_html, config, local)
        manuscript = build_manuscript(row, config, selection, index)
        assert_no_source_sentence_copy(manuscript, raw_html, manuscript["페이지타이틀"])
        manuscripts[local] = manuscript
        signal_log[local] = selection.codes
    return manuscripts, signal_log


def validate_manuscripts(
    prepared: dict[str, dict[str, dict[str, str]]],
    rows: list[dict[str, str]],
) -> tuple[float, tuple[str, str] | None, int]:
    expected_locals = [row["근처 수업가능 동네"].strip() for row in rows]
    if len(expected_locals) != EXPECTED_PAGE_COUNT or len(set(expected_locals)) != EXPECTED_PAGE_COUNT:
        raise ValueError("센터 CSV 지역명은 371개이며 모두 고유해야 합니다.")

    all_titles: list[str] = []
    all_meta: list[str] = []
    all_faq_questions: list[str] = []
    all_manuscript_signatures: list[str] = []
    similarity_records: list[tuple[str, set[tuple[str, ...]]]] = []
    row_by_local = {row["근처 수업가능 동네"].strip(): row for row in rows}
    forbidden = ("AEO", "GEO", "SEO", "D열", "복사 원고", "키워드 삽입")
    blocked_copy_patterns = (
        r"학생(?:이|은)\s+학생이",
        r"학생과\s+학생이",
        r"학부모는\s+이를\s+학부모에게|학부모가\s+학부모에게",
        r"가정에서\s+(?:가정에서|다음\s+상담에서)",
        r"기록으로\s+꾸준히\s+기록",
        r"기록하기[^가-힣]{0,3}기록",
        r"확인할\s+순서를\s+정하는\s+순서",
        r"기준을\s+조정",
        r"다시\s+풀\s+문제의\s+개수를\s+다시",
        r"독해\s+근거\s+찾기을",
        r"확인한\s+다음\s+다음",
        r"질문을\s+질문하려면",
        r"문항별로\s+문항별",
        r"근거가\s+구체적인\s+근거",
        r"판단할\s+기준을\s+다시\s+판단",
        r"표시했는지를\s+표시하면",
        r"학교\s+범위표와\s+현재\s+진도와",
        r"등록\s+전에\s+다시\s+물을\s+질문",
        r"기록에서\s+계속\s+남길\s+항목",
        r"며칠\s+간격으로\s+확인한\s+기억\s+기록",
        r"다음\s+주에\s+줄일\s+부담",
        r"독해\s+근거\s+찾기의\s+근거",
        r"재학\s+중인\s+학교의\s+실제\s+자료",
        r"이\s+진단이\s+필요한\s+것으로\s+확인되면",
        r"첫\s+진단\s+결과와\s+학교\s+일정이\s+허용하는\s+경우",
        r"반영했습니다",
        r"근거를\s+근거로",
        r"상담\s+자료를\s+바탕으로[^.!?]{0,100}이\s+자료를\s+바탕으로",
        r"학생이[^.!?]{0,100}학생의\s+설명",
        r"시험\s+전\s+먼저\s+끝낼\s+범위를\s+먼저\s+확정",
        r"변화가\s+달라졌다면",
        r"상담\s+관점에서는,?\s*상담에서는",
        r"관련\s+자료를\s+기준으로[^.!?]{0,100}관련\s+자료",
        r"수업\s+비교를\s+검토할\s+때",
        r"수업\s+비교에서\s+가장\s+먼저\s+볼\s+답",
        r"상담\s+상황\s+예시\s*[:：]\s*상담\s+상황\s+예시",
        r"말이나\s+표시로",
        r"학원\s+관련[^.!?]{0,100}관련\s+자료",
        r"학습의\s+(?:학생이|꾸준히|직접|주말에|가정에서|현재\s+교재에서)",
        r"학습에서\s+가정\s+학습에서",
        r"설명한\s+뒤[^.!?]{0,120}살핀\s+뒤",
        r"다음\s+주\s+학습량을\s+조절할\s+기준을\s+정하면\s+다음\s+학습\s+분량",
        r"시험\s+전\s+먼저[^.!?]{0,100}먼저\s+찾아",
        r"(?:실제\s+수행량|실제로\s+끝낸\s+분량)[^.!?]{0,100}실제로",
        r"상담에서\s+상담을\s+시작하기\s+전",
        r"질문\s+목록을\s+만들기\s+전[^.!?]{0,100}질문을\s+모아",
        r"다음\s+확인일을\s+정할\s+때[^.!?]{0,100}다음\s+확인\s+때",
        r"(?:학습|이\s+학습|수업\s+계획|공부\s+과정|복습\s+과정)의\s+가정\s+복습",
        r"(?:상담|이\s+상담|첫\s+상담|진단\s+상담)의\s+(?:수업\s+시간|주차별\s+분량|넷째\s+주|실제\s+분량|실제\s+반\s+편성|수업\s+가능\s+여부|피드백\s+주기|조건부\s+예시|첫\s+주\s+확인\s+항목)",
        r"학생의\s+학습의|상담의\s+상담의|의\s+학습의",
        r"학생의\s+(?:수학|영어)\s+공부할\s+때는",
        r"상담\s+자리(?:친\s+뒤에는|서(?:\s|센터))",
        r"(?:중[1-3]|초[3-4])\s+(?:수학|영어)\s+학생에게",
        r"학생이\s+준비한\s+실제\s+학교\s+자료를\s+준비할\s+때",
        r"의\s+(?:중[1-3]|초[3-4])\s+(?:수학|영어)\s+학습의",
        r"(?:중[1-3]|초[3-4])\s+학생의\s+(?:수학|영어)\s+(?:수업\s+계획|공부\s+과정|복습\s+과정)의",
        r"학습에서는\s+학습량만",
        r"확인하기[’']를\s+실천했는지\s+확인",
        r"(?:중[1-3]|초[3-4])\s+학생[^.!?]{0,90}학생에게",
        r"상담\s+자리\s+안내가\s+특히\s+필요",
        r"확인했는지를\s+확인",
        r"수업\s+시간과\s+보완\s+방식은\s+같나요",
        r"(?:수학|영어)의\s+한\s+주\s+뒤\s+확인\s+기준",
        r"관련\s+자료[^.!?]{0,90}관련\s+범위",
        r"관련\s+자료[^.!?]{0,90}관련\s+(?:문항|질문)",
        r"자료를\s+준비할\s+때[^.!?]{0,120}학습\s+내용만\s+준비",
        r"기록되어\s+있는지도\s+남기는",
        r"복습에서는[^.!?]{0,120}누적\s+복습\s+습관도",
        r"같은\s+유형을\s+며칠\s+뒤\s+다시[^.!?]{0,140}다시\s+배울",
        r"학생이\s+준비한\s+실제\s+학교\s+자료[^.!?]{0,140}실제\s+문장",
        r"다음\s+행동을\s+점검하는[^.!?]{0,120}다음\s+행동",
        r"다시\s+확인할\s+날짜[^.!?]{0,120}다시\s+확인할\s+날짜",
    )

    for config in CONFIGS:
        if config.category not in prepared:
            continue
        manuscripts = prepared[config.category]
        if list(manuscripts) != expected_locals:
            raise ValueError(f"CSV 행 순서 대응 오류: {config.category}")
        for local, manuscript in manuscripts.items():
            title = manuscript["페이지타이틀"]
            required_keys = {"페이지타이틀", "메타설명", "본문", "FAQ", "학부모후기", "JSON-LD 요약"}
            if set(manuscript) != required_keys:
                raise ValueError(f"원고 키 오류: {title}")
            if title != f"{local} {config.subject_label}학원":
                raise ValueError(f"페이지 제목 오류: {title}")
            if not 80 <= len(manuscript["메타설명"]) <= 155:
                raise ValueError(f"메타설명 길이 오류: {title}={len(manuscript['메타설명'])}")
            intro, sections = generator.parse_body(manuscript["본문"])
            if len(intro) != 2 or len(sections) != 6:
                raise ValueError(f"본문 구조 오류: {title}, intro={len(intro)}, sections={len(sections)}")
            if any(not 2 <= len(paragraphs) <= 3 for _, paragraphs in sections):
                raise ValueError(f"H2 문단 수 오류: {title}")
            faqs = generator.parse_faq(manuscript["FAQ"])
            if len(faqs) != 5:
                raise ValueError(f"FAQ 개수 오류: {title}={len(faqs)}")
            reviews = generator.paragraph_list(manuscript["학부모후기"])
            if len(reviews) != 2:
                raise ValueError(f"상담 상황 예시 개수 오류: {title}={len(reviews)}")
            if not all(
                token in reviews[0]
                for token in (
                    "실제 이용 후기",
                    "특정 학생의 성과가 아니며",
                    "가상 예시",
                    f"{local}에서 {config.grade_label} {config.subject}{eul_reul(config.subject)} 공부하는 학생의 학습 결과는 출발점과 실천 정도에 따라 달라질 수 있습니다",
                )
            ):
                raise ValueError(f"첫 상담 예시 면책 문구 오류: {title}")
            if "학생의 실제 자료와 가능한 일정에 맞춰 조정해야 합니다" not in reviews[1]:
                raise ValueError(f"둘째 상담 예시 조건부 문구 오류: {title}")
            joined = "\n".join(manuscript.values())
            if any(term in joined for term in forbidden):
                raise ValueError(f"제작용 표현 노출: {title}")
            for pattern in blocked_copy_patterns:
                if re.search(pattern, joined):
                    raise ValueError(f"공개 문장 결합 오류: {title}: {pattern}")
            if config.subject == "수학" and (
                "교과서 예문을 바꿔 쓴 결과" in joined
                or "수업 전에 표시해 둔 낯선 표현" in joined
            ):
                raise ValueError(f"수학 문맥에 영어 근거 노출: {title}")
            if config.subject == "영어" and "문제를 시작하기 전에 적은 조건" in joined:
                raise ValueError(f"영어 문맥에 수학 근거 노출: {title}")

            page_sentences = split_sentences(joined)
            page_counts = Counter(page_sentences)
            repeated_in_page = [sentence for sentence, count in page_counts.items() if count > 1]
            if repeated_in_page:
                raise ValueError(f"페이지 내부 동일 문장: {title}: {repeated_in_page[:2]}")
            context_start_pattern = re.compile(
                rf"^(?:{re.escape(local)}\s+{re.escape(config.grade_label)}\s+"
                rf"{re.escape(config.subject)}\s+(?:상담|학습)|"
                rf"{re.escape(local)}의\s+{re.escape(config.grade_label)}\s+"
                rf"{re.escape(config.subject)}\s+(?:상담|학습))"
            )
            context_starts = sum(
                bool(context_start_pattern.search(sentence)) for sentence in page_sentences
            )
            # One natural page anchor per paragraph is intentional.  This cap
            # still rejects the earlier per-sentence keyword-prefix failure.
            if context_starts > 28:
                raise ValueError(
                    f"문두 지역·학년·과목 라벨 과다: {title}={context_starts}"
                )

            all_titles.append(title)
            all_meta.append(manuscript["메타설명"])
            all_faq_questions.extend(question for question, _ in faqs)
            all_manuscript_signatures.append(
                hashlib.sha256(joined.encode("utf-8")).hexdigest()
            )
            similarity_records.append(
                (
                    f"{config.category}/{local}",
                    masked_manuscript_shingles(
                        manuscript,
                        row_by_local[local],
                        config,
                    ),
                )
            )

    uniqueness_groups = {
        "title": all_titles,
        "meta": all_meta,
        "FAQ question": all_faq_questions,
        "manuscript": all_manuscript_signatures,
    }
    for label, values in uniqueness_groups.items():
        if len(values) != len(set(values)):
            duplicates = [value for value, count in Counter(values).items() if count > 1]
            raise ValueError(f"{label} 중복: {duplicates[:3]}")

    return validate_masked_similarity(similarity_records)


def representative_inventory() -> dict[int, str]:
    inventory: dict[int, str] = {}
    directory = generator.SITE / "assets" / "representative"
    for path in sorted(directory.glob("rep-*.*")):
        match = re.fullmatch(r"rep-(\d{3})\.(?:jpe?g|png|webp|gif)", path.name, re.IGNORECASE)
        if not match:
            continue
        number = int(match.group(1))
        if number in inventory:
            raise ValueError(f"대표 이미지 번호 중복: rep-{number:03d}")
        inventory[number] = f"assets/representative/{path.name}"
    expected = set(range(1, EXPECTED_PAGE_COUNT + 1))
    if set(inventory) != expected:
        raise ValueError(
            f"대표 이미지 인벤토리 오류: missing={sorted(expected - set(inventory))[:10]}, "
            f"extra={sorted(set(inventory) - expected)[:10]}"
        )
    return inventory


def representative_asset_for(
    config: CategoryConfig,
    inventory: dict[int, str],
    index: int,
) -> str:
    if not 0 <= index < EXPECTED_PAGE_COUNT:
        raise IndexError(f"대표 이미지 인덱스 범위 오류: {index}")
    number = ((index + config.representative_offset) % EXPECTED_PAGE_COUNT) + 1
    return inventory[number]


def validate_static_assets(rows: list[dict[str, str]], inventory: dict[int, str]) -> None:
    assignments: dict[str, list[str]] = {}
    for config in CONFIGS:
        values = [
            representative_asset_for(config, inventory, index)
            for index in range(EXPECTED_PAGE_COUNT)
        ]
        if len(set(values)) != EXPECTED_PAGE_COUNT:
            raise ValueError(f"카테고리 대표 이미지 순열 오류: {config.category}")
        assignments[config.category] = values
    for index, row in enumerate(rows):
        local_assets = {assignments[config.category][index] for config in CONFIGS}
        if len(local_assets) != len(CONFIGS):
            raise ValueError(f"같은 동네 카테고리 간 대표 이미지 중복: {row['근처 수업가능 동네']}")

    missing_maps: list[str] = []
    fallback_maps: list[str] = []
    for row in rows:
        map_asset = generator.find_map(row)
        local = row["근처 수업가능 동네"].strip()
        if not map_asset.startswith("assets/maps/"):
            fallback_maps.append(local)
        elif not (generator.SITE / map_asset).is_file():
            missing_maps.append(f"{local}:{map_asset}")
    if fallback_maps or missing_maps:
        raise ValueError(
            f"지도 371/371 대응 실패: fallback={fallback_maps[:10]}, missing={missing_maps[:10]}"
        )


def configure_generator(
    config: CategoryConfig,
    manuscripts: dict[str, dict[str, str]],
    inventory: dict[int, str],
    rows: list[dict[str, str]],
) -> None:
    generator.CATEGORY = config.category
    generator.SUBJECT_LABEL = config.subject_label
    generator.SUBJECT = config.subject
    generator.SUBJECT_EN = config.subject_en
    generator.FOCUS_LABEL = config.focus_label
    generator.GRADE_NUMBER = config.grade_number
    generator.GRADE_EN = config.grade_en
    generator.SCHOOL_LEVEL_NAME = config.school_level
    generator.SCHOOL_KEYS = (config.school_key,)
    generator.PUBLISH_DATE = CONTENT_DATE
    generator.MODIFIED_DATE = CONTENT_DATE
    generator.PRESERVE_SECTION_ORDER = True
    generator.load_manuscripts = lambda data=manuscripts: data
    generator.repeated_body_signatures = lambda _manuscripts: set()
    generator.representative_asset = (
        lambda index, cfg=config, assets=inventory: representative_asset_for(cfg, assets, index)
    )
    availability_key = f"가능학년\n({config.subject})"
    target_grade = config.grade_label

    def rows_for_category() -> list[dict[str, str]]:
        rendered_rows: list[dict[str, str]] = []
        for source_row in rows:
            row = dict(source_row)
            listed_grades = {
                re.sub(r"\s+", "", value)
                for value in re.findall(r"(?:초|중|고)\s*[1-6]", row.get(availability_key, ""))
            }
            if target_grade not in listed_grades:
                row[availability_key] = ""
            rendered_rows.append(row)
        return rendered_rows

    generator.load_center_rows = rows_for_category


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="스터디와와 학년·과목별 지역 페이지를 생성합니다."
    )
    parser.add_argument(
        "--category",
        action="append",
        choices=[config.category for config in CONFIGS],
        help="생성할 카테고리입니다. 여러 번 지정할 수 있으며, 생략하면 전체 카테고리를 생성합니다.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    selected_names = set(args.category or [config.category for config in CONFIGS])
    selected = [config for config in CONFIGS if config.category in selected_names]

    rows = generator.read_csv(generator.COMMON / "센터정보 정리.csv")
    if len(rows) != EXPECTED_PAGE_COUNT:
        raise ValueError(f"센터 CSV 개수 오류: {len(rows)}")
    # Pass normalized values to the shared renderer as well, so visible chips,
    # ItemList entities and schema school references all use the same facts.
    for row in rows:
        row["타깃학교\n(중)"] = "·".join(middle_school_names(row))
        row["타깃학교\n(초)"] = "·".join(elementary_school_names(row))
    generator.load_center_rows = lambda data=rows: [dict(row) for row in data]
    inventory = representative_inventory()
    validate_static_assets(rows, inventory)

    prepared: dict[str, dict[str, dict[str, str]]] = {}
    signal_logs: dict[str, dict[str, tuple[str, str, str]]] = {}
    for config in selected:
        manuscripts, signal_log = build_category_manuscripts(rows, config)
        prepared[config.category] = manuscripts
        signal_logs[config.category] = signal_log
    similarity_max, similarity_pair, similarity_candidates = validate_manuscripts(
        prepared,
        rows,
    )

    for config in selected:
        counts = Counter(code for codes in signal_logs[config.category].values() for code in codes)
        print(
            f"preflight category={config.category} pages={len(prepared[config.category])} "
            f"source_signals={dict(sorted(counts.items()))}"
        )
    print("preflight maps=371/371 representative_per_category=371 cross_category_local_overlap=0")
    print(
        f"preflight masked_5shingle_max={similarity_max:.4f} limit<0.75 "
        f"candidates={similarity_candidates} pair={similarity_pair}"
    )

    # Two passes ensure that reciprocal same-local category links are present even
    # when all category directories are created during this invocation.
    for pass_number in (1, 2):
        for config in selected:
            configure_generator(config, prepared[config.category], inventory, rows)
            generator.main()
        print(f"generation pass={pass_number}/2 categories={len(selected)}")


if __name__ == "__main__":
    main()
