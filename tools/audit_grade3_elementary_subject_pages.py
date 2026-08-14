from __future__ import annotations

import csv
import hashlib
import html
import json
import math
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, Iterator
from urllib.parse import quote, unquote, urljoin, urlparse
from xml.etree import ElementTree as ET

from openpyxl import load_workbook


SITE = Path(__file__).resolve().parents[1]
COMMON = SITE.parent / "참고자료" / "공통자료"
COMMON_CSV = COMMON / "센터정보 정리.csv"
SOURCE_DIR = SITE.parents[1] / "새 폴더"
SITEMAP = SITE / "sitemap.xml"
DOMAIN = "https://xn--2z1b50xixca111l.com"
DOMAIN_HOST = urlparse(DOMAIN).netloc
PARENT = "과목별학원"
SITE_NAME = "코칭아카데미"
EXPECTED_LOCALS = 371
MAX_EXAMPLES = 8
SOURCE_SENTENCE_MIN = 60
SOURCE_SHINGLE_WORDS = 12
SIMILARITY_LIMIT = 0.75


@dataclass(frozen=True)
class Category:
    slug: str
    level: str
    grade: int
    subject: str
    source_name: str

    @property
    def grade_token(self) -> str:
        return f"{self.level}{self.grade}"

    @property
    def label(self) -> str:
        return f"{self.grade_token} {self.subject}"

    @property
    def school_column(self) -> str:
        return f"타깃학교\n({'초' if self.level == '초' else '중'})"

    @property
    def availability_column(self) -> str:
        return f"가능학년\n({self.subject})"


CATEGORIES = (
    Category("중3수학학원", "중", 3, "수학", "중3 수학학원 원고.xlsx"),
    Category("중3영어학원", "중", 3, "영어", "중3 영어학원 원고.xlsx"),
    Category("초3수학학원", "초", 3, "수학", "초3 수학학원 원고.xlsx"),
    Category("초3영어학원", "초", 3, "영어", "초3 영어학원.xlsx"),
    Category("초4수학학원", "초", 4, "수학", "초4 수학학원 원고.xlsx"),
)

ALL_SUBJECT_CATEGORIES = (
    "고1수학학원",
    "고1영어학원",
    "고2수학학원",
    "고2영어학원",
    "중1수학학원",
    "중1영어학원",
    "중2수학학원",
    "중2영어학원",
    *(category.slug for category in CATEGORIES),
)

REQUIRED_DETAIL_TYPES = {
    "EducationalOrganization",
    "LocalBusiness",
    "Article",
    "Service",
    "FAQPage",
    "BreadcrumbList",
    "ItemList",
}
FORBIDDEN_SCHEMA_TYPES = {"Review", "AggregateRating"}
FORBIDDEN_SCHEMA_KEYS = {
    "aggregateRating",
    "bestRating",
    "ratingCount",
    "ratingValue",
    "review",
    "reviewCount",
    "reviewRating",
    "worstRating",
}
AUTHORING_PATTERNS = (
    ("D열 제작 메모", re.compile(r"D열")),
    ("원고 제작 메모", re.compile(r"(?:이|본|해당)?\s*원고(?:에|는|를|가|에서|로)?")),
    ("검색 최적화 제작어", re.compile(r"(?<![A-Za-z])(?:SEO|AEO|GEO)(?![A-Za-z])", re.I)),
    ("JSON-LD 제작어", re.compile(r"JSON\s*-?\s*LD", re.I)),
    ("메타 제작어", re.compile(r"메타\s*(?:설명|요소|태그|데이터)")),
    ("생성 프롬프트 흔적", re.compile(r"프롬프트|생성형\s*AI|챗\s*GPT|ChatGPT", re.I)),
    ("복사·재작성 메모", re.compile(r"다른\s*사이트|기존\s*글|복사한\s*내용|재작성한\s*내용")),
)
OVERCLAIM_PATTERNS = (
    (
        "성적·점수 상승 단정",
        re.compile(
            r"(?:성적|점수|등급)(?:이|가|을|를|은|는|으로|에서|까지|\s)*"
            r"(?:바로\s*)?(?:오르|올리|상승|향상|개선|회복|높이|끌어올리|완성|보장)"
        ),
    ),
    (
        "결과 보장 표현",
        re.compile(
            r"(?:100\s*%|무조건|반드시|확실히|완벽(?:히|하게)?|단기간에?).{0,32}"
            r"(?:성적|점수|등급|향상|상승|합격|성공|해결|보장)"
            r"|(?:성적|점수|등급|합격).{0,32}(?:100\s*%|무조건|반드시|보장)"
        ),
    ),
    ("하락 방지 단정", re.compile(r"(?:성적|점수)\s*하락을\s*(?:막|방지)")),
    ("과도한 최상급", re.compile(r"업계\s*1위|전국\s*1위|최고의\s*학원|유일한\s*학원|완벽한\s*성적")),
)
RAW_COPY_PATTERNS = (
    (
        "원본 article-* 클래스",
        re.compile(
            r'class=["\'][^"\']*\barticle-(?:main|hero|intro|section|local-feature-section|subject-card|target-card|closing)\b',
            re.I,
        ),
    ),
    ("원본 영문 표제", re.compile(r"LOCAL\s+ACADEMY\s+GUIDE", re.I)),
    (
        "참고 사이트 도메인",
        re.compile(r"(?:전문수업|전국학원|소수정예학원|코칭센터)\.(?:com|kr)|학습코칭\.kr", re.I),
    ),
)
DISCLOSURE_PATTERNS = (
    re.compile(r"실제.{0,45}(?:이용\s*)?후기.{0,80}(?:아니|않)"),
    re.compile(r"후기.{0,45}(?:아니|않).{0,100}(?:가상\s*)?예시"),
)
VARIABLE_RESULT_PATTERN = re.compile(r"학습\s*결과.{0,100}(?:달라질|다를)\s*수")
VIRTUAL_EXAMPLE_PATTERN = re.compile(r"(?:가상|상담\s*상황)\s*예시")
NEGATIVE_SUPPORT_PATTERN = re.compile(r"미지원|지원하지\s*않|수업\s*불가|제공하지\s*않|운영하지\s*않")
SUPPORT_CONFIRM_PATTERN = re.compile(
    r"(?:실제\s*)?수업\s*가능\s*(?:학년|여부|시간|일정)?.{0,70}(?:상담|확인)"
    r"|(?:상담|문의).{0,70}(?:실제\s*)?수업\s*가능"
)

TOKEN_RE = re.compile(r"[가-힣]+|[a-z0-9]+(?:[-._][a-z0-9]+)*", re.I)
SENTENCE_SPLIT_RE = re.compile(r"(?:\r?\n)+|(?<=[.!?])\s+")
VOID_ELEMENTS = {
    "area", "base", "br", "col", "embed", "hr", "img", "input",
    "link", "meta", "param", "source", "track", "wbr",
}
SKIP_TEXT_TAGS = {"script", "style", "template", "noscript"}
AUTHORED_CLASSES = {"answer-section", "manuscript-section", "faq-list", "review-grid"}


def normalize_text(value: object) -> str:
    text = unicodedata.normalize("NFKC", html.unescape(str(value or "")))
    text = text.replace("\u200b", "").replace("\ufeff", "")
    return re.sub(r"\s+", " ", text).strip()


def token_fingerprint(value: str) -> str:
    return " ".join(token.lower() for token in TOKEN_RE.findall(normalize_text(value)))


def split_sentences(value: str) -> list[str]:
    return [normalize_text(part) for part in SENTENCE_SPLIT_RE.split(value) if normalize_text(part)]


def slug_ko(value: str) -> str:
    value = re.sub(r"\s+", "", value.strip())
    return re.sub(r'[\\/:*?"<>|#%&+]', "", value)


def canonical_school(value: str) -> str:
    value = normalize_text(value)
    if value.endswith(("초등학교", "고등학교")):
        return value[:-3]
    if value.endswith("중학교"):
        return value[:-2]
    return value


def split_school_items(value: str) -> list[str]:
    result: list[str] = []
    # 공통 CSV는 학교를 쉼표·가운뎃점뿐 아니라 공백만으로 나눈 행도 있다.
    for part in re.split(r"[,/·.\s]+", value or ""):
        name = canonical_school(part)
        if name and name not in result:
            result.append(name)
    return result


def split_grade_items(value: str) -> list[str]:
    return [part for part in re.split(r"[,/·.\s]+", normalize_text(value)) if part]


def absolute_url(path: str) -> str:
    return DOMAIN + quote(path, safe="/:#?&=%")


def category_path(category: Category) -> str:
    return f"/{PARENT}/{category.slug}/"


def detail_path(category: Category, local: str) -> str:
    return f"/{PARENT}/{category.slug}/{slug_ko(local)}/"


def schema_types(node: dict[str, Any]) -> set[str]:
    value = node.get("@type")
    if isinstance(value, str):
        return {value}
    if isinstance(value, list):
        return {item for item in value if isinstance(item, str)}
    return set()


def json_walk(value: Any) -> Iterator[Any]:
    yield value
    if isinstance(value, dict):
        for child in value.values():
            yield from json_walk(child)
    elif isinstance(value, list):
        for child in value:
            yield from json_walk(child)


def schema_items(node: dict[str, Any]) -> list[dict[str, Any]]:
    value = node.get("itemListElement", [])
    if not isinstance(value, list):
        return []
    return [item for item in value if isinstance(item, dict)]


def nonempty(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    if isinstance(value, (list, dict, tuple, set)):
        return bool(value)
    return True


@dataclass
class Node:
    tag: str
    attrs: dict[str, str] = field(default_factory=dict)
    parent: Node | None = None
    children: list[Node | str] = field(default_factory=list)

    @property
    def classes(self) -> set[str]:
        return set(self.attrs.get("class", "").split())

    def ancestors(self) -> Iterator[Node]:
        current = self.parent
        while current is not None:
            yield current
            current = current.parent


class DOMParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.root = Node("document")
        self.stack = [self.root]

    def handle_starttag(self, tag: str, attrs_list: list[tuple[str, str | None]]) -> None:
        tag = tag.lower()
        attrs = {key.lower(): value or "" for key, value in attrs_list}
        node = Node(tag, attrs, self.stack[-1])
        self.stack[-1].children.append(node)
        if tag not in VOID_ELEMENTS:
            self.stack.append(node)

    def handle_startendtag(self, tag: str, attrs_list: list[tuple[str, str | None]]) -> None:
        self.handle_starttag(tag, attrs_list)
        if tag.lower() not in VOID_ELEMENTS:
            self.handle_endtag(tag)

    def handle_endtag(self, tag: str) -> None:
        tag = tag.lower()
        for index in range(len(self.stack) - 1, 0, -1):
            if self.stack[index].tag == tag:
                del self.stack[index:]
                return

    def handle_data(self, data: str) -> None:
        self.stack[-1].children.append(data)


def parse_dom(source: str) -> Node:
    parser = DOMParser()
    parser.feed(source)
    parser.close()
    return parser.root


def iter_nodes(node: Node) -> Iterator[Node]:
    for child in node.children:
        if isinstance(child, Node):
            yield child
            yield from iter_nodes(child)


def nodes_by_tag(root: Node, tag: str) -> list[Node]:
    return [node for node in iter_nodes(root) if node.tag == tag]


def nodes_by_class(root: Node, class_name: str) -> list[Node]:
    return [node for node in iter_nodes(root) if class_name in node.classes]


def node_text(node: Node, *, skip_hidden: bool = False) -> str:
    parts: list[str] = []

    def visit(current: Node | str) -> None:
        if isinstance(current, str):
            parts.append(current)
            return
        if current.tag in SKIP_TEXT_TAGS:
            return
        style = current.attrs.get("style", "").replace(" ", "").lower()
        hidden = (
            "display:none" in style
            or "visibility:hidden" in style
            or "hidden" in current.attrs
            or current.attrs.get("aria-hidden", "").lower() == "true"
        )
        if skip_hidden and hidden:
            return
        for child in current.children:
            visit(child)

    visit(node)
    return normalize_text(" ".join(parts))


def first_descendant(node: Node, tag: str) -> Node | None:
    return next((child for child in iter_nodes(node) if child.tag == tag), None)


def is_descendant_of_class(node: Node, classes: set[str]) -> bool:
    return bool(node.classes & classes) or any(parent.classes & classes for parent in node.ancestors())


@dataclass
class ParsedPage:
    path: Path
    source: str
    root: Node
    title: str
    descriptions: list[str]
    canonicals: list[str]
    og_urls: list[str]
    h1s: list[str]
    graph: list[dict[str, Any]]
    json_values: list[Any]


@dataclass
class Audit:
    failures: Counter[str] = field(default_factory=Counter)
    examples: dict[str, list[str]] = field(default_factory=lambda: defaultdict(list))
    checks: Counter[str] = field(default_factory=Counter)

    def fail(self, code: str, path: Path | str, detail: str) -> None:
        self.failures[code] += 1
        if len(self.examples[code]) >= MAX_EXAMPLES:
            return
        if isinstance(path, Path):
            try:
                label = path.resolve().relative_to(SITE.resolve()).as_posix()
            except (OSError, ValueError):
                label = str(path)
        else:
            label = path
        self.examples[code].append(f"{label}: {detail}")

    def checked(self, code: str, count: int = 1) -> None:
        self.checks[code] += count

    def finish(self, similarity: SimilarityResult | None = None) -> int:
        print("=== 중3·초3·초4 과목별학원 신규 5카테고리 독립 엄격 감사 ===")
        print("checks " + " ".join(f"{key}={self.checks[key]}" for key in sorted(self.checks)))
        if similarity is not None:
            print(
                "similarity "
                f"documents={similarity.documents} candidates={similarity.candidates} "
                f"max={similarity.maximum:.4f} limit<{SIMILARITY_LIMIT:.2f}"
            )
            if similarity.maximum_pair:
                print(f"similarity_max_pair={' <> '.join(similarity.maximum_pair)}")
        if self.failures:
            print("AUDIT_FAILED " + " ".join(
                f"{code}={self.failures[code]}" for code in sorted(self.failures)
            ))
            for code in sorted(self.failures):
                print(f"\n[{code}] examples")
                for example in self.examples[code]:
                    print(f"- {example}")
            return 1
        print("AUDIT_OK failures=0")
        return 0


def parse_page(path: Path, audit: Audit) -> ParsedPage | None:
    try:
        source = path.read_text(encoding="utf-8")
    except (OSError, UnicodeError) as exc:
        audit.fail("html_read", path, str(exc))
        return None
    try:
        root = parse_dom(source)
    except Exception as exc:
        audit.fail("html_parse", path, str(exc))
        return None
    titles = [node_text(node) for node in nodes_by_tag(root, "title")]
    descriptions = [
        normalize_text(node.attrs.get("content"))
        for node in nodes_by_tag(root, "meta")
        if node.attrs.get("name", "").lower() == "description"
    ]
    canonicals = [
        node.attrs.get("href", "").strip()
        for node in nodes_by_tag(root, "link")
        if "canonical" in node.attrs.get("rel", "").lower().split()
    ]
    og_urls = [
        node.attrs.get("content", "").strip()
        for node in nodes_by_tag(root, "meta")
        if node.attrs.get("property", "").lower() == "og:url"
    ]
    h1s = [node_text(node) for node in nodes_by_tag(root, "h1")]
    json_values: list[Any] = []
    graph: list[dict[str, Any]] = []
    json_nodes = [
        node for node in nodes_by_tag(root, "script")
        if node.attrs.get("type", "").lower() == "application/ld+json"
    ]
    if not json_nodes:
        audit.fail("jsonld_missing", path, "application/ld+json 블록 없음")
    for index, node in enumerate(json_nodes, 1):
        raw = "".join(child for child in node.children if isinstance(child, str)).strip()
        try:
            value = json.loads(raw)
        except (TypeError, ValueError) as exc:
            audit.fail("jsonld_parse", path, f"block={index}: {exc}")
            continue
        json_values.append(value)
        if isinstance(value, dict) and isinstance(value.get("@graph"), list):
            candidates = value["@graph"]
        elif isinstance(value, list):
            candidates = value
        else:
            candidates = [value]
        graph.extend(item for item in candidates if isinstance(item, dict))
    return ParsedPage(
        path=path,
        source=source,
        root=root,
        title=titles[0] if len(titles) == 1 else "",
        descriptions=descriptions,
        canonicals=canonicals,
        og_urls=og_urls,
        h1s=h1s,
        graph=graph,
        json_values=json_values,
    )


def read_rows(audit: Audit) -> list[dict[str, str]]:
    if not COMMON_CSV.is_file():
        audit.fail("common_csv_missing", COMMON_CSV, "센터정보 정리.csv 없음")
        return []
    try:
        with COMMON_CSV.open("r", encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.DictReader(handle))
    except (OSError, UnicodeError, csv.Error) as exc:
        audit.fail("common_csv_read", COMMON_CSV, str(exc))
        return []
    required_columns = {
        "근처 수업가능 동네", "동 영어", "지역", "시or구", "센터명", "센터 주소",
        "타깃학교\n(초)", "타깃학교\n(중)",
        "가능학년\n(영어)", "가능학년\n(수학)",
    }
    missing_columns = required_columns - set(rows[0] if rows else {})
    if missing_columns:
        audit.fail("common_csv_columns", COMMON_CSV, f"missing={sorted(missing_columns)}")
    locals_ = [normalize_text(row.get("근처 수업가능 동네")) for row in rows]
    if len(rows) != EXPECTED_LOCALS or len(set(locals_)) != EXPECTED_LOCALS or "" in locals_:
        audit.fail(
            "common_csv_contract",
            COMMON_CSV,
            f"rows={len(rows)} unique_locals={len(set(locals_))} blanks={locals_.count('')}",
        )
    slugs = [slug_ko(local) for local in locals_]
    if len(set(slugs)) != len(slugs) or any(not slug for slug in slugs):
        audit.fail("common_slug_contract", COMMON_CSV, f"slugs={len(slugs)} unique={len(set(slugs))}")
    return rows


def parse_iso_date(value: Any) -> date | None:
    if not isinstance(value, str) or not value.strip():
        return None
    raw = value.strip().replace("Z", "+00:00")
    try:
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}", raw):
            return date.fromisoformat(raw)
        return datetime.fromisoformat(raw).date()
    except ValueError:
        return None


def page_main_text(page: ParsedPage) -> str:
    mains = nodes_by_tag(page.root, "main")
    return node_text(mains[0], skip_hidden=True) if len(mains) == 1 else ""


def authored_text(page: ParsedPage) -> str:
    mains = nodes_by_tag(page.root, "main")
    if len(mains) != 1:
        return ""
    values: list[str] = []
    for node in iter_nodes(mains[0]):
        if node.classes & AUTHORED_CLASSES:
            if any(parent.classes & AUTHORED_CLASSES for parent in node.ancestors()):
                continue
            value = node_text(node, skip_hidden=True)
            if value:
                values.append(value)
    return normalize_text(" ".join(values))


def has_noindex(page: ParsedPage) -> bool:
    for node in nodes_by_tag(page.root, "meta"):
        if node.attrs.get("name", "").lower() not in {"robots", "googlebot", "bingbot"}:
            continue
        if "noindex" in node.attrs.get("content", "").lower():
            return True
    return bool(re.search(r"x-robots-tag.{0,80}noindex", page.source, re.I | re.S))


def validate_metadata(
    page: ParsedPage,
    expected_title: str,
    expected_h1: str,
    expected_canonical: str,
    audit: Audit,
) -> None:
    if page.title != expected_title:
        audit.fail("title_exact", page.path, f"expected={expected_title!r} actual={page.title!r}")
    if page.h1s != [expected_h1]:
        audit.fail("h1_exact", page.path, f"expected={[expected_h1]!r} actual={page.h1s!r}")
    if page.canonicals != [expected_canonical]:
        audit.fail("canonical_exact", page.path, f"expected={[expected_canonical]!r} actual={page.canonicals!r}")
    if page.og_urls != [expected_canonical]:
        audit.fail("og_url_exact", page.path, f"expected={[expected_canonical]!r} actual={page.og_urls!r}")
    if len(page.descriptions) != 1 or len(page.descriptions[0]) < 60:
        audit.fail("description_contract", page.path, f"values={page.descriptions!r}")
    if has_noindex(page):
        audit.fail("noindex_forbidden", page.path, "371/371 제공·색인 정책에 noindex가 노출됨")


def breadcrumb_labels(node: Node) -> list[str]:
    result: list[str] = []
    for item in iter_nodes(node):
        if item.tag not in {"a", "span", "li"}:
            continue
        value = node_text(item)
        if value and value not in {"/", "›", ">", "»"} and value not in result:
            result.append(value)
    return result


def list_item_name(item: dict[str, Any]) -> str:
    if isinstance(item.get("name"), str):
        return normalize_text(item["name"])
    nested = item.get("item")
    if isinstance(nested, dict):
        return normalize_text(nested.get("name"))
    return ""


def validate_breadcrumb(
    page: ParsedPage,
    category: Category,
    expected_h1: str,
    expected_canonical: str,
    audit: Audit,
) -> None:
    visible = nodes_by_class(page.root, "breadcrumb")
    expected_labels = ["홈", PARENT, category.slug, expected_h1]
    if len(visible) != 1:
        audit.fail("breadcrumb_visible_count", page.path, f"count={len(visible)}")
    else:
        labels = breadcrumb_labels(visible[0])
        if labels != expected_labels:
            audit.fail("breadcrumb_visible_exact", page.path, f"expected={expected_labels!r} actual={labels!r}")
    candidates = [node for node in page.graph if "BreadcrumbList" in schema_types(node)]
    if len(candidates) != 1:
        audit.fail("breadcrumb_schema_count", page.path, f"count={len(candidates)}")
        return
    items = schema_items(candidates[0])
    names = [normalize_text(item.get("name")) for item in items]
    urls = [item.get("item") or item.get("url") for item in items]
    expected_urls = [
        absolute_url("/"),
        absolute_url(f"/{PARENT}/"),
        absolute_url(category_path(category)),
        expected_canonical,
    ]
    if names != expected_labels or urls != expected_urls:
        audit.fail(
            "breadcrumb_schema_exact",
            page.path,
            f"names={names!r} urls={urls!r}",
        )
    positions = [item.get("position") for item in items]
    if positions != [1, 2, 3, 4]:
        audit.fail("breadcrumb_schema_positions", page.path, f"positions={positions!r}")


def validate_schema(
    page: ParsedPage,
    category: Category,
    row: dict[str, str],
    expected_h1: str,
    expected_canonical: str,
    audit: Audit,
) -> None:
    found_types: set[str] = set()
    for node in page.graph:
        found_types.update(schema_types(node))
    if missing := REQUIRED_DETAIL_TYPES - found_types:
        audit.fail("schema_required_types", page.path, f"missing={sorted(missing)} found={sorted(found_types)}")

    webpages = [node for node in page.graph if "WebPage" in schema_types(node)]
    if len(webpages) != 1:
        audit.fail("schema_webpage_count", page.path, f"count={len(webpages)}")
    else:
        webpage = webpages[0]
        if webpage.get("url") != expected_canonical or webpage.get("@id") != f"{expected_canonical}#webpage":
            audit.fail("schema_webpage_identity", page.path, f"url={webpage.get('url')!r} id={webpage.get('@id')!r}")
        for key in ("about", "mentions", "hasPart"):
            if not nonempty(webpage.get(key)):
                audit.fail("schema_webpage_semantics", page.path, f"WebPage.{key} 누락/비어 있음")

    articles = [node for node in page.graph if "Article" in schema_types(node)]
    if len(articles) != 1:
        audit.fail("schema_article_count", page.path, f"count={len(articles)}")
    else:
        article = articles[0]
        if normalize_text(article.get("headline")) != expected_h1:
            audit.fail("schema_article_headline", page.path, f"headline={article.get('headline')!r}")
        for key in ("about", "mentions", "hasPart", "articleSection"):
            if not nonempty(article.get(key)):
                audit.fail("schema_article_semantics", page.path, f"Article.{key} 누락/비어 있음")
        published = parse_iso_date(article.get("datePublished"))
        modified = parse_iso_date(article.get("dateModified"))
        if published is None or modified is None or published > modified or modified > date.today():
            audit.fail(
                "schema_article_dates",
                page.path,
                f"datePublished={article.get('datePublished')!r} dateModified={article.get('dateModified')!r}",
            )

    services = [node for node in page.graph if "Service" in schema_types(node)]
    if len(services) != 1:
        audit.fail("schema_service_count", page.path, f"count={len(services)}")
    else:
        service = services[0]
        if service.get("url") != expected_canonical:
            audit.fail("schema_service_url", page.path, f"url={service.get('url')!r}")
        for key in ("about", "mentions", "offers"):
            if not nonempty(service.get(key)):
                audit.fail("schema_service_semantics", page.path, f"Service.{key} 누락/비어 있음")
        audience = service.get("audience") if isinstance(service.get("audience"), dict) else {}
        role = normalize_text(audience.get("educationalRole"))
        expected_role = f"{'초등학교' if category.level == '초' else '중학교'} {category.grade}학년 학생"
        if role != expected_role:
            audit.fail("schema_service_audience", page.path, f"expected={expected_role!r} actual={role!r}")

    organizations = [
        node for node in page.graph
        if schema_types(node) & {"EducationalOrganization", "LocalBusiness"}
    ]
    combined = [
        node for node in organizations
        if {"EducationalOrganization", "LocalBusiness"}.issubset(schema_types(node))
    ]
    if len(combined) != 1:
        audit.fail("schema_organization_combined", page.path, f"count={len(combined)}")
    else:
        organization = combined[0]
        if normalize_text(organization.get("name")) != normalize_text(row.get("센터명")):
            audit.fail("schema_center_fact", page.path, f"name={organization.get('name')!r}")
        if not nonempty(organization.get("makesOffer")):
            audit.fail("schema_makes_offer", page.path, "EducationalOrganization/LocalBusiness.makesOffer 누락")
        address = organization.get("address") if isinstance(organization.get("address"), dict) else {}
        if normalize_text(address.get("streetAddress")) != normalize_text(row.get("센터 주소")):
            audit.fail("schema_address_fact", page.path, f"streetAddress={address.get('streetAddress')!r}")

    for value in page.json_values:
        for node in json_walk(value):
            if not isinstance(node, dict):
                continue
            if illegal_types := schema_types(node) & FORBIDDEN_SCHEMA_TYPES:
                audit.fail("schema_review_forbidden", page.path, f"types={sorted(illegal_types)}")
            if illegal_keys := set(node) & FORBIDDEN_SCHEMA_KEYS:
                audit.fail("schema_rating_forbidden", page.path, f"keys={sorted(illegal_keys)}")


def visible_faqs(page: ParsedPage) -> list[tuple[str, str]]:
    details = [node for node in nodes_by_tag(page.root, "details") if "faq-item" in node.classes]
    result: list[tuple[str, str]] = []
    for detail in details:
        summary = first_descendant(detail, "summary")
        if summary is None:
            continue
        question = node_text(summary)
        answer_nodes = [
            node for node in iter_nodes(detail)
            if node.tag == "p" and not any(parent.tag == "summary" for parent in node.ancestors())
        ]
        answer = normalize_text(" ".join(node_text(node) for node in answer_nodes))
        if question and answer:
            result.append((question, answer))
    return result


def normalize_schema_text(value: Any) -> str:
    return normalize_text(re.sub(r"<[^>]+>", " ", str(value or "")))


def validate_faq(page: ParsedPage, audit: Audit) -> None:
    visible = visible_faqs(page)
    if len(visible) < 4:
        audit.fail("faq_visible_count", page.path, f"count={len(visible)}")
    questions = [question for question, _ in visible]
    if len(questions) != len(set(questions)):
        audit.fail("faq_question_duplicate_within", page.path, f"questions={questions!r}")
    for index, (question, answer) in enumerate(visible, 1):
        if not question.endswith("?"):
            audit.fail("faq_question_form", page.path, f"index={index} question={question!r}")
        if len(answer) < 70 or len(split_sentences(answer)) < 2:
            audit.fail("faq_answer_depth", page.path, f"index={index} chars={len(answer)} sentences={len(split_sentences(answer))}")
    faq_nodes = [node for node in page.graph if "FAQPage" in schema_types(node)]
    if len(faq_nodes) != 1:
        audit.fail("faq_schema_count", page.path, f"count={len(faq_nodes)}")
        return
    entities = faq_nodes[0].get("mainEntity", [])
    if not isinstance(entities, list):
        entities = [entities]
    schema_pairs: list[tuple[str, str]] = []
    for entity in entities:
        if not isinstance(entity, dict):
            continue
        answer = entity.get("acceptedAnswer")
        text = answer.get("text") if isinstance(answer, dict) else ""
        schema_pairs.append((normalize_schema_text(entity.get("name")), normalize_schema_text(text)))
    if visible != schema_pairs:
        audit.fail(
            "faq_visible_schema_mismatch",
            page.path,
            f"visible_count={len(visible)} schema_count={len(schema_pairs)} sample={visible[:1]!r}|{schema_pairs[:1]!r}",
        )


def validate_quick_answer(page: ParsedPage, expected_h1: str, audit: Audit) -> None:
    sections = nodes_by_class(page.root, "answer-section")
    if len(sections) != 1:
        audit.fail("quick_answer_section", page.path, f"count={len(sections)}")
        return
    headings = [node_text(node) for node in iter_nodes(sections[0]) if node.tag == "h2"]
    if headings != [f"{expected_h1} 핵심 답변"]:
        audit.fail("quick_answer_heading", page.path, f"headings={headings!r}")
    items = [node for node in iter_nodes(sections[0]) if "answer-item" in node.classes]
    answers = [node_text(node) for node in iter_nodes(sections[0]) if "a" in node.classes]
    if len(items) != 1 or len(answers) != 1 or len(answers[0]) < 80:
        audit.fail("quick_answer_item", page.path, f"items={len(items)} answers={answers!r}")
        return
    schema_descriptions = {
        normalize_text(node.get("description"))
        for node in page.graph
        if schema_types(node) & {"Article", "Service"} and nonempty(node.get("description"))
    }
    if answers[0] not in schema_descriptions:
        audit.fail("quick_answer_schema_summary", page.path, f"answer={answers[0]!r}")


def resolve_local_asset(page: Path, value: str) -> Path | None:
    raw = value.strip().split()[0] if value.strip() else ""
    if not raw or raw.startswith(("data:", "blob:")):
        return None
    parsed = urlparse(raw)
    if parsed.scheme in {"http", "https"}:
        if parsed.netloc != DOMAIN_HOST:
            return None
        target = SITE / unquote(parsed.path).lstrip("/")
    elif raw.startswith("/"):
        target = SITE / unquote(parsed.path).lstrip("/")
    else:
        target = page.parent / unquote(parsed.path)
    try:
        resolved = target.resolve()
        resolved.relative_to(SITE.resolve())
    except (OSError, ValueError):
        return None
    return resolved


def expected_map(row: dict[str, str]) -> Path | None:
    raw = normalize_text(row.get("동 영어"))
    candidates = (raw, raw.replace(" ", "-"), raw.replace(" ", ""), raw.replace("_", "-"))
    for base in candidates:
        for extension in (".jpg", ".jpeg", ".png", ".webp"):
            path = SITE / "assets" / "maps" / f"{base}{extension}"
            if path.is_file():
                return path.resolve()
    return None


def positive_dimension(node: Node, key: str) -> bool:
    value = node.attrs.get(key, "")
    return value.isdigit() and int(value) > 0


def validate_images(
    page: ParsedPage,
    row: dict[str, str],
    expected_h1: str,
    audit: Audit,
) -> Path | None:
    images = nodes_by_tag(page.root, "img")
    roles = {
        "representative": [
            node for node in images
            if "/representative/" in node.attrs.get("src", "").replace("\\", "/")
            or "대표" in node.attrs.get("alt", "")
        ],
        "body": [
            node for node in images
            if "/centers/common/" in node.attrs.get("src", "").replace("\\", "/")
            or "본문" in node.attrs.get("alt", "")
        ],
        "map": [
            node for node in images
            if "/maps/" in node.attrs.get("src", "").replace("\\", "/")
            or "지도" in node.attrs.get("alt", "")
        ],
    }
    words = {"representative": "대표", "body": "본문", "map": "지도"}
    for role, candidates in roles.items():
        if len(candidates) != 1:
            audit.fail("image_role_count", page.path, f"role={role} count={len(candidates)}")
            continue
        image = candidates[0]
        if not positive_dimension(image, "width") or not positive_dimension(image, "height"):
            audit.fail("image_dimensions", page.path, f"role={role} attrs={image.attrs!r}")
        alt = normalize_text(image.attrs.get("alt"))
        if expected_h1 not in alt or words[role] not in alt:
            audit.fail("image_alt", page.path, f"role={role} alt={alt!r}")
        target = resolve_local_asset(page.path, image.attrs.get("src", ""))
        if target is None or not target.is_file():
            audit.fail("image_missing", page.path, f"role={role} src={image.attrs.get('src')!r}")

    representative: Path | None = None
    if len(roles["representative"]) == 1:
        image = roles["representative"][0]
        target = resolve_local_asset(page.path, image.attrs.get("src", ""))
        pool = (SITE / "assets" / "representative").resolve()
        try:
            if target is None:
                raise ValueError("경로 해석 실패")
            target.relative_to(pool)
            if not target.is_file() or target.suffix.lower() not in {".gif", ".jpg", ".jpeg", ".png", ".webp"}:
                raise ValueError("대표 이미지 풀의 실파일이 아님")
        except ValueError as exc:
            audit.fail("representative_rule", page.path, f"target={target} reason={exc}")
        else:
            representative = target
        style = image.attrs.get("style", "").replace(" ", "").lower()
        if "display:none" not in style or "loading" in image.attrs:
            audit.fail("representative_loading_rule", page.path, f"attrs={image.attrs!r}")

    if len(roles["body"]) == 1:
        image = roles["body"][0]
        target = resolve_local_asset(page.path, image.attrs.get("src", ""))
        filename = "seoul6839.webp" if normalize_text(row.get("지역")) == "서울" else "local6839.webp"
        expected = (SITE / "assets" / "centers" / "common" / filename).resolve()
        if target != expected:
            audit.fail("body_image_rule", page.path, f"actual={target} expected={expected}")
        picture = next((parent for parent in image.ancestors() if parent.tag == "picture"), None)
        sources = [node for node in iter_nodes(picture) if node.tag == "source"] if picture else []
        mobile = filename.replace(".webp", "-mobile.webp")
        if not any(
            (asset := resolve_local_asset(page.path, node.attrs.get("srcset", ""))) is not None
            and asset.name == mobile and asset.is_file()
            for node in sources
        ):
            audit.fail("body_mobile_image_rule", page.path, f"expected={mobile}")

    if len(roles["map"]) == 1:
        target = resolve_local_asset(page.path, roles["map"][0].attrs.get("src", ""))
        expected = expected_map(row)
        if expected is None or target != expected:
            audit.fail("map_image_rule", page.path, f"actual={target} expected={expected}")

    for node in images + nodes_by_tag(page.root, "source"):
        attribute = "srcset" if node.tag == "source" else "src"
        value = node.attrs.get(attribute, "")
        parsed = urlparse(value.split()[0] if value else "")
        is_local = not parsed.scheme or parsed.netloc == DOMAIN_HOST
        if value and is_local:
            target = resolve_local_asset(page.path, value)
            if target is None or not target.is_file():
                audit.fail("image_missing", page.path, f"{attribute}={value!r}")
    return representative


def build_school_patterns(rows: list[dict[str, str]]) -> dict[str, re.Pattern[str] | None]:
    result: dict[str, re.Pattern[str] | None] = {}
    for level, column in (("초", "타깃학교\n(초)"), ("중", "타깃학교\n(중)")):
        names = {
            name for row in rows for name in split_school_items(row.get(column, ""))
            if len(name) >= 2
        }
        result[level] = (
            re.compile(
                rf"(?<![가-힣A-Za-z0-9])({'|'.join(re.escape(name) for name in sorted(names, key=len, reverse=True))})(?:학교)?(?![가-힣A-Za-z0-9])"
            )
            if names else None
        )
    return result


def validate_facts_schools_support(
    page: ParsedPage,
    category: Category,
    row: dict[str, str],
    school_pattern: re.Pattern[str] | None,
    audit: Audit,
) -> None:
    visible = page_main_text(page)
    if not visible:
        audit.fail("main_output_empty", page.path, "공개 main 텍스트가 비어 있음")
        return
    region = normalize_text(row.get("지역"))
    district = normalize_text(row.get("시or구"))
    if normalize_text(row.get("센터 주소")).startswith("세종특별자치시"):
        region, district = "세종특별자치시", "세종시"
    elif district.endswith(("로", "길")):
        district = ""
    facts = {
        "근처 수업가능 동네": normalize_text(row.get("근처 수업가능 동네")),
        "지역": region,
        "시or구": district,
        "센터명": normalize_text(row.get("센터명")),
        "센터 주소": normalize_text(row.get("센터 주소")),
    }
    for key, value in facts.items():
        if value and value not in visible:
            audit.fail("csv_visible_fact", page.path, f"missing {key}={value!r}")

    allowed = set(split_school_items(row.get(category.school_column, "")))
    school_lists = [
        node for node in page.graph
        if "ItemList" in schema_types(node)
        and ("schools" in str(node.get("@id", "")).lower() or "학교" in normalize_text(node.get("name")))
    ]
    if len(school_lists) != 1:
        audit.fail("school_itemlist_count", page.path, f"count={len(school_lists)}")
        schema_names: set[str] = set()
    else:
        schema_names = {
            canonical_school(list_item_name(item)) for item in schema_items(school_lists[0])
            if list_item_name(item) and list_item_name(item) != "상담 시 학교 확인"
        }
        if schema_names != allowed:
            audit.fail("school_itemlist_exact", page.path, f"expected={sorted(allowed)} actual={sorted(schema_names)}")

    chip_containers = nodes_by_class(page.root, "chip-list")
    chip_names: set[str] = set()
    for container in chip_containers:
        for node in iter_nodes(container):
            if node.tag in {"span", "li", "a"}:
                name = canonical_school(node_text(node))
                if name and name != "상담 시 학교 확인":
                    chip_names.add(name)
    if chip_names != allowed:
        audit.fail("school_visible_exact", page.path, f"expected={sorted(allowed)} actual={sorted(chip_names)}")
    if not allowed and "상담 시 학교 확인" not in visible:
        audit.fail("school_empty_fallback", page.path, "학교 CSV 공란인데 상담 확인 고지 없음")

    mentioned = {
        canonical_school(match.group(1)) for match in school_pattern.finditer(visible)
    } if school_pattern else set()
    unexpected = mentioned - allowed
    if unexpected:
        audit.fail("school_whitelist", page.path, f"unexpected={sorted(unexpected)} allowed={sorted(allowed)}")
    missing = allowed - mentioned
    if missing:
        audit.fail("school_missing", page.path, f"missing={sorted(missing)}")
    suffix = "초등학교" if category.level == "초" else "중학교"
    explicit = {
        canonical_school(match.group(1))
        for match in re.finditer(
            rf"(?<![가-힣A-Za-z0-9])([가-힣A-Za-z0-9·]{{2,24}}{suffix})(?![가-힣A-Za-z0-9])",
            visible,
        )
    }
    unknown_explicit = explicit - allowed
    if unknown_explicit:
        audit.fail("school_whitelist", page.path, f"unknown_explicit={sorted(unknown_explicit)}")

    available = normalize_text(row.get(category.availability_column))
    grade_items = set(split_grade_items(available))
    supported = category.grade_token in grade_items
    expected_availability = available if supported else "상담 시 확인"
    if f"{category.subject} {expected_availability}" not in visible:
        audit.fail(
            "availability_csv_fact",
            page.path,
            f"missing={category.subject} {expected_availability}",
        )
    property_values = [node for node in page.graph if "PropertyValue" in schema_types(node)]
    matching = [
        node for node in property_values
        if normalize_text(node.get("name")) == f"{category.subject} 수업 가능 학년"
    ]
    if len(matching) != 1:
        audit.fail("availability_schema_count", page.path, f"count={len(matching)}")
    elif normalize_text(matching[0].get("value")) != expected_availability:
        audit.fail(
            "availability_schema_fact",
            page.path,
            f"expected={expected_availability!r} actual={matching[0].get('value')!r}",
        )
    if not supported and not SUPPORT_CONFIRM_PATTERN.search(visible):
        audit.fail(
            "support_confirmation_branch",
            page.path,
            f"{category.grade_token}이 CSV 가능학년에 없지만 실제 가능 학년·일정 상담 확인 고지가 없음",
        )
    if NEGATIVE_SUPPORT_PATTERN.search(visible):
        audit.fail("support_negative_forbidden", page.path, "371/371 제공 정책과 충돌하는 미지원·수업 불가 문구")
    audit.checked("support_listed" if supported else "support_consultation_branch")


def internal_target(page: Path, canonical: str, href: str) -> tuple[str, Path] | None:
    href = href.strip()
    if not href or href.startswith(("#", "tel:", "sms:", "mailto:", "javascript:", "data:")):
        return None
    direct = urlparse(href)
    if direct.scheme in {"http", "https"} and direct.netloc != DOMAIN_HOST:
        return None
    joined = urlparse(urljoin(canonical, href))
    if joined.netloc and joined.netloc != DOMAIN_HOST:
        return None
    route = unquote(joined.path)
    if not route.startswith("/"):
        route = "/" + route
    target = SITE / route.lstrip("/")
    if route.endswith("/"):
        target = target / "index.html"
    elif not target.suffix and target.is_dir():
        target = target / "index.html"
    normalized_route = route
    if normalized_route == "/index.html":
        normalized_route = "/"
    elif normalized_route.endswith("/index.html"):
        normalized_route = normalized_route[:-len("index.html")]
    try:
        target = target.resolve()
        target.relative_to(SITE.resolve())
    except (OSError, ValueError):
        pass
    return normalized_route, target


def validate_internal_links(
    page: ParsedPage,
    category: Category,
    local: str,
    expected_canonical: str,
    audit: Audit,
) -> None:
    paths: set[str] = set()
    for anchor in nodes_by_tag(page.root, "a"):
        target = internal_target(page.path, expected_canonical, anchor.attrs.get("href", ""))
        if target is None:
            continue
        route, file_path = target
        paths.add(route)
        if not file_path.is_file():
            audit.fail("broken_internal_link", page.path, f"href={anchor.attrs.get('href')!r} target={file_path}")
    required = {"/", f"/{PARENT}/", category_path(category)}
    if missing := required - paths:
        audit.fail("internal_link_structure", page.path, f"missing={sorted(missing)}")
    same_category_prefix = category_path(category)
    if not any(path.startswith(same_category_prefix) and path != detail_path(category, local) for path in paths):
        audit.fail("internal_link_peer_local", page.path, "같은 카테고리의 다른 지역 상세 링크 없음")
    same_local_routes = {
        f"/{PARENT}/{slug}/{slug_ko(local)}/" for slug in ALL_SUBJECT_CATEGORIES if slug != category.slug
    }
    if not paths & same_local_routes:
        audit.fail("internal_link_cross_category", page.path, "같은 지역의 다른 과목·학년 링크 없음")
    related = [
        node for node in page.graph
        if "ItemList" in schema_types(node) and "related" in str(node.get("@id", "")).lower()
    ]
    if len(related) != 1 or len(schema_items(related[0])) < 4:
        audit.fail("internal_link_itemlist", page.path, f"count={len(related)} items={len(schema_items(related[0])) if related else 0}")


@dataclass(frozen=True)
class SourceRef:
    category: str
    row: int
    local: str

    def label(self) -> str:
        return f"{self.category}:xlsx-row-{self.row}:{self.local}"


def build_geo_pattern(rows: list[dict[str, str]]) -> re.Pattern[str] | None:
    terms: set[str] = set()
    for row in rows:
        for key in ("근처 수업가능 동네", "지역", "시or구", "센터명"):
            if value := normalize_text(row.get(key)):
                terms.add(value)
    if not terms:
        return None
    alternatives = "|".join(re.escape(term) for term in sorted(terms, key=len, reverse=True))
    return re.compile(rf"(?<![가-힣A-Za-z0-9])(?:{alternatives})(?![가-힣A-Za-z0-9])", re.I)


def mask_copy_entities(value: str, geo_pattern: re.Pattern[str] | None) -> str:
    value = normalize_text(value)
    value = re.sub(
        r"(?:와와학습코칭센터|스터디와와|코칭아카데미|코칭센터)(?:\s*[가-힣A-Za-z0-9()]+점)?",
        " ORGTOKEN ",
        value,
        flags=re.I,
    )
    value = re.sub(r"(?<![가-힣A-Za-z0-9])(?:초|중|고)\s*[1-6](?:\s*학년)?(?![가-힣A-Za-z0-9])", " GRADETOKEN ", value)
    value = re.sub(r"(?<![가-힣A-Za-z0-9])(?:초등학교|중학교|고등학교)\s*[1-6]\s*학년(?![가-힣A-Za-z0-9])", " GRADETOKEN ", value)
    value = re.sub(r"(?<![가-힣])(?:국어|영어|수학|과학|사회|영수)(?![가-힣])", " SUBJECTTOKEN ", value)
    if geo_pattern is not None:
        value = geo_pattern.sub(" LOCTOKEN ", value)
    return token_fingerprint(value)


def shingle_digest(tokens: tuple[str, ...]) -> int:
    payload = "\x1f".join(tokens).encode("utf-8")
    return int.from_bytes(hashlib.blake2b(payload, digest_size=8).digest(), "big")


def iter_source_shingles(fingerprint: str) -> Iterator[tuple[int, str]]:
    words = fingerprint.split()
    for index in range(max(0, len(words) - SOURCE_SHINGLE_WORDS + 1)):
        shingle = tuple(words[index:index + SOURCE_SHINGLE_WORDS])
        yield shingle_digest(shingle), " ".join(shingle)


def source_public_text(fragment: str) -> str:
    root = parse_dom(fragment)
    mains = nodes_by_tag(root, "main")
    return node_text(mains[0], skip_hidden=True) if mains else node_text(root, skip_hidden=True)


def load_source_indexes(
    rows: list[dict[str, str]],
    geo_pattern: re.Pattern[str] | None,
    audit: Audit,
) -> tuple[dict[str, SourceRef], dict[str, SourceRef], dict[int, SourceRef]]:
    raw_sentences: dict[str, SourceRef] = {}
    masked_sentences: dict[str, SourceRef] = {}
    shingles: dict[int, SourceRef] = {}
    for category in CATEGORIES:
        path = SOURCE_DIR / category.source_name
        if not path.is_file():
            audit.fail("source_xlsx_missing", path, "첨부 원고 파일 없음")
            continue
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
            worksheet = workbook.active
            values = [
                (excel_row, row[0])
                for excel_row, row in enumerate(worksheet.iter_rows(values_only=True), 1)
                if row and isinstance(row[0], str) and row[0].strip()
            ]
            workbook.close()
        except Exception as exc:
            audit.fail("source_xlsx_read", path, str(exc))
            continue
        if len(values) != EXPECTED_LOCALS:
            audit.fail("source_row_count", path, f"expected={EXPECTED_LOCALS} actual={len(values)}")
        for position, (excel_row, fragment) in enumerate(values, 1):
            local = normalize_text(rows[position - 1].get("근처 수업가능 동네")) if position <= len(rows) else ""
            ref = SourceRef(category.slug, excel_row, local)
            try:
                public = source_public_text(fragment)
            except Exception as exc:
                audit.fail("source_html_parse", path, f"row={excel_row}: {exc}")
                continue
            if not public:
                audit.fail("source_public_empty", path, f"row={excel_row} local={local}")
                continue
            audit.checked("source_documents")
            for sentence in split_sentences(public):
                raw = token_fingerprint(sentence)
                masked = mask_copy_entities(sentence, geo_pattern)
                if len(raw) >= SOURCE_SENTENCE_MIN:
                    raw_sentences.setdefault(raw, ref)
                if len(masked) >= SOURCE_SENTENCE_MIN:
                    masked_sentences.setdefault(masked, ref)
            fingerprint = mask_copy_entities(public, geo_pattern)
            for digest, _ in iter_source_shingles(fingerprint):
                shingles.setdefault(digest, ref)
    if not raw_sentences or not shingles:
        audit.fail("source_index_empty", SOURCE_DIR, f"sentences={len(raw_sentences)} shingles={len(shingles)}")
    audit.checked("source_sentence_fingerprints", len(set(raw_sentences) | set(masked_sentences)))
    audit.checked("source_shingle_fingerprints", len(shingles))
    return raw_sentences, masked_sentences, shingles


def validate_source_rewrite(
    value: str,
    path: Path,
    geo_pattern: re.Pattern[str] | None,
    raw_sentences: dict[str, SourceRef],
    masked_sentences: dict[str, SourceRef],
    source_shingles: dict[int, SourceRef],
    audit: Audit,
) -> None:
    sentence_hits: list[str] = []
    for sentence in split_sentences(value):
        raw = token_fingerprint(sentence)
        masked = mask_copy_entities(sentence, geo_pattern)
        ref = raw_sentences.get(raw) if len(raw) >= SOURCE_SENTENCE_MIN else None
        if ref is None and len(masked) >= SOURCE_SENTENCE_MIN:
            ref = masked_sentences.get(masked)
        if ref is not None:
            sentence_hits.append(f"{ref.label()} :: {sentence[:120]}")
    if sentence_hits:
        audit.fail("source_sentence_reuse", path, f"matches={len(sentence_hits)} sample={sentence_hits[:2]!r}")
    shingle_hits: dict[int, tuple[SourceRef, str]] = {}
    for digest, sample in iter_source_shingles(mask_copy_entities(value, geo_pattern)):
        ref = source_shingles.get(digest)
        if ref is not None:
            shingle_hits.setdefault(digest, (ref, sample))
    if shingle_hits:
        examples = [f"{ref.label()} :: {sample}" for ref, sample in list(shingle_hits.values())[:2]]
        audit.fail("source_12word_shingle_reuse", path, f"matches={len(shingle_hits)} sample={examples!r}")


def validate_visible_safety(page: ParsedPage, category: Category, audit: Audit) -> None:
    visible = page_main_text(page)
    authored = authored_text(page)
    if not authored:
        audit.fail("authored_output_empty", page.path, "답변·본문·FAQ·상담 예시 텍스트 없음")
        return
    for name, pattern in AUTHORING_PATTERNS:
        if match := pattern.search(visible):
            audit.fail("authoring_trace", page.path, f"{name}: {match.group(0)!r}")
    for name, pattern in OVERCLAIM_PATTERNS:
        if match := pattern.search(authored):
            audit.fail("outcome_guarantee", page.path, f"{name}: {match.group(0)!r}")
    for name, pattern in RAW_COPY_PATTERNS:
        if match := pattern.search(page.source):
            audit.fail("raw_copy_trace", page.path, f"{name}: {match.group(0)!r}")
    if not any(pattern.search(visible) for pattern in DISCLOSURE_PATTERNS):
        audit.fail("consultation_example_not_review", page.path, "실제 후기·성과가 아니라는 고지 없음")
    if not VIRTUAL_EXAMPLE_PATTERN.search(visible):
        audit.fail("consultation_example_virtual", page.path, "가상/상담 상황 예시 고지 없음")
    if not VARIABLE_RESULT_PATTERN.search(visible):
        audit.fail("consultation_example_result_variable", page.path, "학습 결과 개인차 고지 없음")
    unsafe_reviews = [
        sentence for sentence in split_sentences(visible)
        if "후기" in sentence and not re.search(r"(?:아니|않|예시)", sentence)
    ]
    if unsafe_reviews:
        audit.fail("visible_review_claim", page.path, f"sample={unsafe_reviews[:2]!r}")
    target = category.grade_token
    other_grades = {
        match.group(0).replace(" ", "")
        for match in re.finditer(r"(?<![가-힣A-Za-z0-9])(?:초|중|고)\s*[1-6](?![가-힣A-Za-z0-9])", authored)
        if match.group(0).replace(" ", "") != target
    }
    if other_grades:
        audit.fail("other_grade_authored_exposure", page.path, f"target={target} found={sorted(other_grades)}")


@dataclass
class ContentRecord:
    label: str
    path: Path
    blocks: list[str]
    paragraphs: list[str]
    sections: list[str]
    mask_values: list[str]
    shingles: set[tuple[str, ...]] = field(default_factory=set)


def authored_units(page: ParsedPage) -> tuple[list[str], list[str], list[str]]:
    mains = nodes_by_tag(page.root, "main")
    if len(mains) != 1:
        return [], [], []
    blocks: list[str] = []
    paragraphs: list[str] = []
    sections: list[str] = []
    section_classes = {"answer-item", "manuscript-card", "review-card", "faq-item"}
    for node in iter_nodes(mains[0]):
        if node.tag in {"h2", "h3", "p", "li", "summary"} and is_descendant_of_class(node, AUTHORED_CLASSES):
            value = node_text(node, skip_hidden=True)
            if value:
                blocks.append(value)
                if node.tag == "p" and len(value) >= 90 and len(split_sentences(value)) >= 2:
                    paragraphs.append(value)
        if node.classes & section_classes:
            value = node_text(node, skip_hidden=True)
            if len(value) >= 120 and len(split_sentences(value)) >= 2:
                sections.append(value)
    return blocks, paragraphs, sections


def mask_content(record: ContentRecord) -> str:
    value = normalize_text(" \n ".join(record.blocks)).lower()
    for raw in sorted({normalize_text(item).lower() for item in record.mask_values if normalize_text(item)}, key=len, reverse=True):
        value = re.sub(re.escape(raw), " entitytoken ", value, flags=re.I)
    value = re.sub(r"(?<![가-힣a-z0-9])(?:초|중|고)\s*[1-6](?:\s*학년)?(?![가-힣a-z0-9])", " gradetoken ", value)
    value = re.sub(r"(?<![가-힣])(?:국어|영어|수학|과학|사회|영수)(?![가-힣])", " subjecttoken ", value)
    return token_fingerprint(value)


def make_similarity_shingles(record: ContentRecord) -> set[tuple[str, ...]]:
    tokens = mask_content(record).split()
    return {tuple(tokens[index:index + 5]) for index in range(max(0, len(tokens) - 4))}


@dataclass
class SimilarityResult:
    documents: int
    candidates: int
    maximum: float
    maximum_pair: tuple[str, str] | None


def validate_exact_content_units(records: list[ContentRecord], audit: Audit) -> None:
    for kind, getter in (("paragraph", lambda record: record.paragraphs), ("section", lambda record: record.sections)):
        occurrences: dict[str, list[str]] = defaultdict(list)
        for record in records:
            for index, value in enumerate(getter(record), 1):
                occurrences[normalize_text(value)].append(f"{record.label}#{index}")
        duplicates = [
            (value, labels) for value, labels in occurrences.items() if len(labels) > 1
        ]
        for value, labels in sorted(duplicates, key=lambda item: (-len(item[1]), item[0])):
            within = len({label.split("#", 1)[0] for label in labels}) == 1
            audit.fail(
                f"duplicate_multisentence_{kind}_{'within' if within else 'cross'}",
                "content",
                f"occurrences={len(labels)} pages={labels[:6]!r} text={value[:220]!r}",
            )
    documents: dict[str, list[str]] = defaultdict(list)
    for record in records:
        value = normalize_text(" \n ".join(record.blocks))
        if value:
            documents[value].append(record.label)
    for value, labels in documents.items():
        if len(labels) > 1:
            audit.fail("duplicate_exact_authored_document", "content", f"pages={labels[:6]!r} text={value[:220]!r}")


def audit_similarity(records: list[ContentRecord], audit: Audit) -> SimilarityResult:
    for record in records:
        record.shingles = make_similarity_shingles(record)
        if len(record.shingles) < 40:
            audit.fail("similarity_content_too_short", record.path, f"5-shingles={len(record.shingles)}")
    usable = [record for record in records if record.shingles]
    frequencies: Counter[tuple[str, ...]] = Counter()
    for record in usable:
        frequencies.update(record.shingles)
    ordered = sorted(usable, key=lambda record: (len(record.shingles), record.label))
    postings: dict[tuple[str, ...], list[int]] = defaultdict(list)
    candidate_count = 0
    maximum = 0.0
    maximum_pair: tuple[str, str] | None = None
    for index, record in enumerate(ordered):
        size = len(record.shingles)
        ordered_tokens = sorted(record.shingles, key=lambda token: (frequencies[token], token))
        prefix_length = size - math.ceil(SIMILARITY_LIMIT * size) + 1
        minimum_prior_size = math.ceil(SIMILARITY_LIMIT * size)
        candidates: set[int] = set()
        for token in ordered_tokens[:prefix_length]:
            candidates.update(
                prior for prior in postings.get(token, [])
                if len(ordered[prior].shingles) >= minimum_prior_size
            )
        for prior in candidates:
            other = ordered[prior]
            intersection = len(record.shingles & other.shingles)
            union = size + len(other.shingles) - intersection
            score = intersection / union if union else 1.0
            candidate_count += 1
            if score > maximum:
                maximum = score
                maximum_pair = (other.label, record.label)
            if score >= SIMILARITY_LIMIT:
                audit.fail(
                    "masked_5shingle_similarity",
                    "content",
                    f"score={score:.4f} pages={other.label!r}, {record.label!r}",
                )
        for token in record.shingles:
            postings[token].append(index)
    return SimilarityResult(len(usable), candidate_count, maximum, maximum_pair)


def validate_category_hub(
    category: Category,
    rows: list[dict[str, str]],
    audit: Audit,
) -> tuple[str, str, str, str] | None:
    path = SITE / PARENT / category.slug / "index.html"
    if not path.is_file():
        audit.fail("category_hub_missing", path, "신규 카테고리 허브 없음")
        return None
    page = parse_page(path, audit)
    if page is None:
        return None
    canonical = absolute_url(category_path(category))
    title = f"{category.slug} | {SITE_NAME}"
    validate_metadata(page, title, category.slug, canonical, audit)
    found: set[str] = set()
    for node in page.graph:
        found.update(schema_types(node))
    if missing := {"CollectionPage", "BreadcrumbList", "ItemList"} - found:
        audit.fail("category_hub_schema", path, f"missing={sorted(missing)}")
    expected_urls = {absolute_url(detail_path(category, row["근처 수업가능 동네"])) for row in rows}
    item_lists = [node for node in page.graph if "ItemList" in schema_types(node)]
    matching = []
    for node in item_lists:
        urls = {item.get("url") or item.get("item") for item in schema_items(node)}
        if urls == expected_urls:
            matching.append(node)
    if len(matching) != 1 or len(schema_items(matching[0])) != EXPECTED_LOCALS:
        audit.fail("category_hub_itemlist", path, f"matching={len(matching)}")
    anchor_urls: set[str] = set()
    for anchor in nodes_by_tag(page.root, "a"):
        target = internal_target(path, canonical, anchor.attrs.get("href", ""))
        if target and target[0].startswith(category_path(category)) and target[0] != category_path(category):
            anchor_urls.add(absolute_url(target[0]))
            if not target[1].is_file():
                audit.fail("category_hub_broken_link", path, f"href={anchor.attrs.get('href')!r}")
    if anchor_urls != expected_urls:
        audit.fail(
            "category_hub_links",
            path,
            f"missing={len(expected_urls - anchor_urls)} extra={len(anchor_urls - expected_urls)}",
        )
    audit.checked("category_hubs")
    description = page.descriptions[0] if len(page.descriptions) == 1 else ""
    return title, category.slug, canonical, description


def validate_parent_hub(audit: Audit) -> None:
    path = SITE / PARENT / "index.html"
    if not path.is_file():
        audit.fail("parent_hub_missing", path, "과목별학원 부모 허브 없음")
        return
    page = parse_page(path, audit)
    if page is None:
        return
    canonical = absolute_url(f"/{PARENT}/")
    expected_paths = {f"/{PARENT}/{slug}/" for slug in ALL_SUBJECT_CATEGORIES}
    grids = nodes_by_class(page.root, "category-grid")
    if len(grids) != 1:
        audit.fail("parent_hub_cards", path, f"category-grid count={len(grids)}")
    else:
        paths: list[str] = []
        for anchor in [node for node in iter_nodes(grids[0]) if node.tag == "a"]:
            target = internal_target(path, canonical, anchor.attrs.get("href", ""))
            if target:
                paths.append(target[0])
                if not target[1].is_file():
                    audit.fail("parent_hub_broken_link", path, f"href={anchor.attrs.get('href')!r}")
        if len(paths) != len(expected_paths) or set(paths) != expected_paths:
            audit.fail("parent_hub_cards", path, f"count={len(paths)} expected={len(expected_paths)} paths={paths!r}")
    expected_urls = {absolute_url(value) for value in expected_paths}
    matches = []
    for node in [item for item in page.graph if "ItemList" in schema_types(item)]:
        urls = {item.get("url") or item.get("item") for item in schema_items(node)}
        if urls == expected_urls:
            matches.append(node)
    if len(matches) != 1 or len(schema_items(matches[0])) != len(expected_urls):
        audit.fail("parent_hub_itemlist", path, f"matching={len(matches)} expected_items={len(expected_urls)}")
    audit.checked("parent_hub")


def parse_sitemap(audit: Audit) -> list[str]:
    if not SITEMAP.is_file():
        audit.fail("sitemap_missing", SITEMAP, "sitemap.xml 없음")
        return []
    try:
        root = ET.parse(SITEMAP).getroot()
    except (ET.ParseError, OSError) as exc:
        audit.fail("sitemap_parse", SITEMAP, str(exc))
        return []
    locations = [normalize_text(node.text) for node in root.findall(".//{*}loc") if normalize_text(node.text)]
    if len(locations) != len(set(locations)):
        audit.fail("sitemap_duplicate", SITEMAP, f"count={len(locations)} unique={len(set(locations))}")
    return locations


def validate_sitemap(rows: list[dict[str, str]], expected: set[str], audit: Audit) -> None:
    locations = parse_sitemap(audit)
    prefixes = tuple(category_path(category) for category in CATEGORIES)
    actual = {
        location for location in locations
        if unquote(urlparse(location).path).startswith(prefixes)
    }
    if actual != expected:
        audit.fail(
            "sitemap_scope_exact",
            SITEMAP,
            f"expected={len(expected)} actual={len(actual)} missing={len(expected - actual)} extra={len(actual - expected)}",
        )
    target_count = len(CATEGORIES) * (len(rows) + 1)
    if len(expected) != target_count:
        audit.fail("sitemap_expected_contract", SITEMAP, f"expected_set={len(expected)} target={target_count}")
    if absolute_url(f"/{PARENT}/") not in set(locations):
        audit.fail("sitemap_parent_hub", SITEMAP, "과목별학원 부모 허브 URL 누락")
    audit.checked("sitemap_urls", len(actual))


def main() -> int:
    audit = Audit()
    rows = read_rows(audit)
    if not rows:
        return audit.finish()

    geo_pattern = build_geo_pattern(rows)
    school_patterns = build_school_patterns(rows)
    raw_source_sentences, masked_source_sentences, source_shingles = load_source_indexes(
        rows,
        geo_pattern,
        audit,
    )
    expected_source_docs = len(CATEGORIES) * EXPECTED_LOCALS
    if audit.checks["source_documents"] != expected_source_docs:
        audit.fail(
            "source_document_coverage",
            SOURCE_DIR,
            f"actual={audit.checks['source_documents']} expected={expected_source_docs}",
        )

    expected_canonicals: set[str] = set()
    metadata: dict[str, list[tuple[str, str]]] = {
        "title": [], "h1": [], "canonical": [], "description": [],
    }
    records: list[ContentRecord] = []
    representative_assignments: dict[str, dict[str, Path]] = defaultdict(dict)

    for category in CATEGORIES:
        hub = validate_category_hub(category, rows, audit)
        if hub is not None:
            for key, value in zip(metadata, hub):
                metadata[key].append((f"hub:{category.slug}", value))
            expected_canonicals.add(hub[2])

        root = SITE / PARENT / category.slug
        expected_slugs = {slug_ko(row["근처 수업가능 동네"]) for row in rows}
        detail_files = sorted(root.glob("*/index.html")) if root.is_dir() else []
        actual_slugs = {path.parent.name for path in detail_files}
        if actual_slugs != expected_slugs or len(detail_files) != EXPECTED_LOCALS:
            audit.fail(
                "detail_output_set",
                root,
                f"expected={EXPECTED_LOCALS} actual_files={len(detail_files)} actual_slugs={len(actual_slugs)} "
                f"missing={len(expected_slugs - actual_slugs)} extra={len(actual_slugs - expected_slugs)} "
                f"missing_sample={sorted(expected_slugs - actual_slugs)[:5]!r}",
            )

        for row in rows:
            local = normalize_text(row.get("근처 수업가능 동네"))
            path = root / slug_ko(local) / "index.html"
            if not path.is_file():
                audit.fail("detail_output_missing", path, "371/371 제공 대상 상세 페이지 없음")
                continue
            page = parse_page(path, audit)
            if page is None:
                continue
            expected_h1 = f"{local} {category.label}학원"
            expected_title = f"{expected_h1} | {SITE_NAME}"
            canonical = absolute_url(detail_path(category, local))
            validate_metadata(page, expected_title, expected_h1, canonical, audit)
            validate_breadcrumb(page, category, expected_h1, canonical, audit)
            validate_schema(page, category, row, expected_h1, canonical, audit)
            validate_quick_answer(page, expected_h1, audit)
            validate_faq(page, audit)
            validate_visible_safety(page, category, audit)
            representative = validate_images(page, row, expected_h1, audit)
            if representative is not None:
                representative_assignments[category.slug][local] = representative
            validate_facts_schools_support(
                page,
                category,
                row,
                school_patterns.get(category.level),
                audit,
            )
            validate_internal_links(page, category, local, canonical, audit)

            authored = authored_text(page)
            if authored:
                validate_source_rewrite(
                    authored,
                    path,
                    geo_pattern,
                    raw_source_sentences,
                    masked_source_sentences,
                    source_shingles,
                    audit,
                )
            blocks, paragraphs, sections = authored_units(page)
            if not blocks:
                audit.fail("authored_blocks_empty", path, "중복 감사 대상 블록 없음")
            records.append(ContentRecord(
                label=f"{category.slug}/{local}",
                path=path,
                blocks=blocks,
                paragraphs=paragraphs,
                sections=sections,
                mask_values=[
                    local,
                    normalize_text(row.get("지역")),
                    normalize_text(row.get("시or구")),
                    normalize_text(row.get("센터명")),
                    normalize_text(row.get("센터 주소")),
                    category.slug,
                    category.label,
                    category.grade_token,
                    f"{'초등학교' if category.level == '초' else '중학교'} {category.grade}학년",
                    category.subject,
                    *split_school_items(row.get(category.school_column, "")),
                ],
            ))

            description = page.descriptions[0] if len(page.descriptions) == 1 else ""
            values = (
                ("title", page.title),
                ("h1", page.h1s[0] if len(page.h1s) == 1 else ""),
                ("canonical", page.canonicals[0] if len(page.canonicals) == 1 else ""),
                ("description", description),
            )
            for key, value in values:
                metadata[key].append((f"{category.slug}/{local}", value))
            expected_canonicals.add(canonical)
            audit.checked("detail_pages")

    representative_root = SITE / "assets" / "representative"
    representative_pool = {
        path.resolve() for path in representative_root.iterdir()
        if path.is_file() and path.suffix.lower() in {".gif", ".jpg", ".jpeg", ".png", ".webp"}
    } if representative_root.is_dir() else set()
    if len(representative_pool) != EXPECTED_LOCALS:
        audit.fail("representative_pool", representative_root, f"actual={len(representative_pool)} expected={EXPECTED_LOCALS}")
    for category in CATEGORIES:
        assignments = representative_assignments.get(category.slug, {})
        paths = set(assignments.values())
        if len(assignments) != EXPECTED_LOCALS or len(paths) != EXPECTED_LOCALS:
            audit.fail(
                "representative_category_unique",
                SITE / PARENT / category.slug,
                f"assignments={len(assignments)} unique={len(paths)} expected={EXPECTED_LOCALS}",
            )
        elif paths != representative_pool:
            audit.fail(
                "representative_category_pool",
                SITE / PARENT / category.slug,
                f"missing={len(representative_pool - paths)} outside={len(paths - representative_pool)}",
            )
    for row in rows:
        local = normalize_text(row.get("근처 수업가능 동네"))
        paths = [representative_assignments.get(category.slug, {}).get(local) for category in CATEGORIES]
        present = [path for path in paths if path is not None]
        if len(present) == len(CATEGORIES) and len(set(present)) != len(CATEGORIES):
            audit.fail("representative_locality_distinct", f"representatives/{local}", f"files={[path.name for path in present]!r}")
    audit.checked("representative_assignments", sum(len(value) for value in representative_assignments.values()))

    expected_metadata = len(CATEGORIES) * (EXPECTED_LOCALS + 1)
    for key, values in metadata.items():
        groups: dict[str, list[str]] = defaultdict(list)
        for label, value in values:
            if value:
                groups[value].append(label)
        for value, labels in groups.items():
            if len(labels) > 1:
                audit.fail(f"{key}_unique", "metadata", f"pages={labels[:6]!r} value={value!r}")
        if len(values) != expected_metadata:
            audit.fail(f"{key}_coverage", "metadata", f"actual={len(values)} expected={expected_metadata}")
        audit.checked(f"unique_{key}", len(groups))

    validate_parent_hub(audit)
    validate_sitemap(rows, expected_canonicals, audit)

    expected_records = len(CATEGORIES) * EXPECTED_LOCALS
    similarity: SimilarityResult | None = None
    if len(records) != expected_records:
        audit.fail("content_coverage", "content", f"records={len(records)} expected={expected_records}")
    else:
        validate_exact_content_units(records, audit)
        similarity = audit_similarity(records, audit)
        audit.checked("similarity_documents", similarity.documents)

    return audit.finish(similarity)


if __name__ == "__main__":
    sys.exit(main())
